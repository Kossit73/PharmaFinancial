"""Streamlit web application for the Pharmaceuticals financial model."""
from __future__ import annotations

import copy
import csv
import hashlib
import io
import json
import re
from datetime import datetime
from pathlib import Path
from collections.abc import Iterable, Mapping, Sequence
import math
from typing import TYPE_CHECKING, Any, Callable, Dict, List, Optional, Tuple, cast

try:  # pragma: no cover - allow importing helper functions without Streamlit installed
    import streamlit as st
except Exception:  # pragma: no cover - lightweight stub for non-Streamlit environments
    class _StreamlitStub:
        session_state: dict = {}

        def __getattr__(self, name: str):  # noqa: D401 - simple runtime guard
            def _missing(*args, **kwargs):
                raise RuntimeError(
                    "Streamlit runtime is required for UI operations."
                )

            return _missing

    st = _StreamlitStub()  # type: ignore
    st.sidebar = _StreamlitStub()  # type: ignore[attr-defined]

if TYPE_CHECKING:  # pragma: no cover - typing aid only
    from streamlit.delta_generator import DeltaGenerator
else:  # pragma: no cover - used when Streamlit isn't fully available
    DeltaGenerator = Any  # type: ignore[misc]

from .ai import AIInsights
from .debt import amortise_entries
from .inputs import DebtEntry, ModelInputs, parse_inputs
from .model import (
    CASH_FLOW_BEGIN_COLUMN,
    CASH_FLOW_END_COLUMN,
    CASH_FLOW_NET_COLUMN,
    FinancialModel,
    FinancialOutputs,
)
from .report import collect_report_sections, generate_report
from .table import Table

try:  # pragma: no cover - executed in environments with pandas available
    import pandas as pd
except Exception:  # pragma: no cover - fallback for environments without pandas
    pd = None  # type: ignore

try:  # pragma: no cover - optional dependency for charting
    import plotly.express as px
except Exception:  # pragma: no cover - gracefully degrade when Plotly missing
    px = None  # type: ignore

try:  # pragma: no cover - optional dependency for Excel ingestion
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - import guard when package missing
    load_workbook = None  # type: ignore

try:  # pragma: no cover - optional dependency for Word ingestion
    from docx import Document
except Exception:  # pragma: no cover - import guard when package missing
    Document = None  # type: ignore

try:  # pragma: no cover - optional dependency for PDF ingestion
    from PyPDF2 import PdfReader
except Exception:  # pragma: no cover - import guard when package missing
    PdfReader = None  # type: ignore

# ---------------------------------------------------------------------------
# Module level caches
# ---------------------------------------------------------------------------

DEFAULT_INPUT_PATH = Path(__file__).resolve().parent / "data" / "default_inputs.json"
DEFAULT_INPUT_JSON = DEFAULT_INPUT_PATH.read_text(encoding="utf-8")
DEFAULT_RISK_CATEGORIES = ["inherent", "climate", "political"]
DEPRECIATION_METHOD_LABELS = {
    "straight_line": "Straight Line",
    "reducing_balance": "Reducing Balance",
}
DEPRECIATION_METHOD_OPTIONS = list(DEPRECIATION_METHOD_LABELS.values())
DEPRECIATION_LABEL_TO_VALUE = {
    label: value for value, label in DEPRECIATION_METHOD_LABELS.items()
}

MAX_VISIBLE_INVENTORY_ROWS = 2
MAX_VISIBLE_INFLATION_ROWS = 2
MAX_VISIBLE_RISK_ROWS = 2
MAX_VISIBLE_RECEIVABLE_ROWS = 2
MAX_VISIBLE_COMMISSION_ROWS = 6
MAX_VISIBLE_COST_ROWS = 6
MIN_PROJECTION_YEAR = 1900
MAX_PROJECTION_YEAR = 2300
SCENARIO_TOOL_LABELS = {
    "decision_tree": "Decision Tree Tools",
    "stress_testing": "Stress Testing",
    "backtesting": "Backtesting",
    "walk_forward": "Walk-forward Testing",
    "driver_based": "Driver-based Modeling",
    "real_options": "Real Options Analysis (ROA)",
}
SCENARIO_TOOL_ALIASES = {
    "walk_forward_testing": "walk_forward",
    "driver_based_modeling": "driver_based",
    "real_options_analysis": "real_options",
}

AI_PROVIDER_OPTIONS = ["OpenAI", "Azure OpenAI", "Anthropic", "Vertex AI", "Custom"]
ML_METHOD_LABELS = {
    "linear_regression": "Linear Regression",
    "cagr": "Compound Annual Growth",
    "moving_average": "Moving Average",
}
ML_LABEL_TO_CODE = {label: code for code, label in ML_METHOD_LABELS.items()}
GEN_AI_FEATURE_LABELS = {
    "summary": "Executive Summary",
    "risk_review": "Risk Review",
    "cash_flow_highlight": "Cash Flow Highlights",
}
GEN_AI_LABEL_TO_CODE = {label: code for code, label in GEN_AI_FEATURE_LABELS.items()}

_INPUT_CACHE: dict[str, ModelInputs] = {}
_MODEL_CACHE: dict[str, tuple["FinancialModel", "FinancialOutputs"]] = {}


def _rerun() -> None:
    """Trigger a Streamlit rerun using the available API.

    Streamlit has exposed multiple rerun helpers across releases. Newer
    versions ship :func:`st.rerun` while older builds provide
    :func:`st.experimental_rerun`.  When the dashboard runs outside of the
    Streamlit runtime (for example during unit tests or direct script
    execution) calling either helper raises ``StreamlitAPIException``.  The
    function therefore attempts each available helper and swallows runtime
    errors so the rest of the UI logic can continue gracefully.
    """

    candidates = []
    rerun = getattr(st, "rerun", None)
    if callable(rerun):  # pragma: no cover - depends on Streamlit version
        candidates.append(rerun)

    legacy = getattr(st, "experimental_rerun", None)
    if callable(legacy):  # pragma: no cover - depends on Streamlit version
        candidates.append(legacy)

    for trigger in candidates:
        try:
            trigger()
            break
        except Exception:  # pragma: no cover - runtime specific failures
            continue


def _ensure_widget_default(key: str, value: object) -> None:
    """Initialise a widget's session value without overwriting edits."""

    try:
        if key not in st.session_state:
            st.session_state[key] = value
    except Exception:  # pragma: no cover - depends on Streamlit runtime
        pass


def _set_widget_value(key: str, value: object) -> None:
    """Synchronise a widget's value in ``st.session_state``.

    Streamlit widgets retain their initial value once instantiated. To ensure
    computed, read-only fields such as the total revenue and cost columns stay
    in sync with their drivers (production units, price, and cost inputs), the
    application updates the backing ``session_state`` entry prior to rendering
    the widget.  When running outside of a Streamlit session—such as during
    automated tests—the assignment may fail, so the helper guards against
    runtime-specific exceptions.
    """

    try:
        st.session_state[key] = value
    except Exception:  # pragma: no cover - depends on Streamlit runtime
        pass


def _json_default(value: object) -> object:
    """Best-effort JSON serialiser for hashing Streamlit payloads."""

    if isinstance(value, (set, tuple)):
        return list(value)
    if isinstance(value, Path):
        return str(value)
    raise TypeError(f"Object of type {type(value).__name__} is not JSON serialisable")


def _payload_digest(payload: Mapping[str, object]) -> str:
    """Return a stable digest representing the current payload contents."""

    serialised = json.dumps(payload, sort_keys=True, default=_json_default)
    return hashlib.sha256(serialised.encode("utf-8")).hexdigest()


def _clone_payload(payload: Mapping[str, object]) -> dict[str, object]:
    """Return a deep copy of ``payload`` suitable for workspace storage."""

    return json.loads(json.dumps(payload, default=_json_default))


def _generate_workspace_label(existing: Mapping[str, object], prefix: str = "Workspace") -> str:
    """Return a workspace label that does not collide with ``existing`` keys."""

    index = 1
    existing_keys = {str(key) for key in existing.keys()}
    while True:
        candidate = f"{prefix} {index}"
        if candidate not in existing_keys:
            return candidate
        index += 1


def _normalise_payload(payload: Mapping[str, object]) -> Mapping[str, object]:
    """Normalise payloads prior to parsing to guarantee stable caching."""

    return json.loads(json.dumps(payload, sort_keys=True, default=_json_default))


def _cached_parse_inputs(payload: Mapping[str, object]) -> tuple[ModelInputs, str]:
    """Parse inputs with caching so repeated reruns avoid recomputation."""

    digest = _payload_digest(payload)
    cached = _INPUT_CACHE.get(digest)
    if cached is None:
        normalised = _normalise_payload(payload)
        cached = parse_inputs(normalised)
        _INPUT_CACHE[digest] = cached
    return cached, digest


def _cached_model_run(inputs: ModelInputs, digest: str) -> tuple[FinancialModel, FinancialOutputs]:
    """Return cached model/output pairs for the provided payload digest."""

    cached = _MODEL_CACHE.get(digest)
    if cached is not None:
        return cached

    model = FinancialModel(inputs)
    outputs = model.run()
    _MODEL_CACHE[digest] = (model, outputs)
    return model, outputs


def _scenario_options(payload: Mapping[str, object]) -> List[str]:
    """Return the available scenario labels including the base case."""

    options: List[str] = ["Base"]
    scenarios = payload.get("scenarios")
    if not isinstance(scenarios, Mapping):
        return options

    for name in scenarios.keys():
        label = str(name).strip()
        if not label:
            continue
        if label.lower() == "base":
            options[0] = label
        elif label not in options:
            options.append(label)

    return options


def _scenario_slug(name: str) -> str:
    """Return a URL-safe slug for ``name`` used in widget keys and filenames."""

    slug = re.sub(r"[^0-9a-z]+", "_", name.strip().lower())
    slug = re.sub(r"_+", "_", slug).strip("_")
    return slug or "base"


def _ensure_scenario_payload(
    selected_scenario: str,
    snapshot: Mapping[str, object],
    base_model: FinancialModel,
    base_outputs: FinancialOutputs,
) -> tuple[FinancialModel, FinancialOutputs]:
    """Return model results for ``selected_scenario`` built from ``snapshot`` payload."""

    scenario_name = (selected_scenario or "Base").strip()
    if not scenario_name:
        scenario_name = "Base"

    if scenario_name.lower() == "base":
        return base_model, base_outputs

    scenarios = snapshot.get("scenarios", {})
    matched: Mapping[str, object] | None = None
    if isinstance(scenarios, Mapping):
        for name, values in scenarios.items():
            if str(name).strip().lower() == scenario_name.lower():
                if isinstance(values, Mapping):
                    matched = values
                break

    if matched is None:
        return base_model, base_outputs

    payload = _clone_payload(snapshot)

    inflation = matched.get("inflation")
    if isinstance(inflation, Iterable) and not isinstance(inflation, (str, bytes)):
        payload["inflation_series"] = [float(value) for value in inflation]

    interest = matched.get("interest")
    if isinstance(interest, Iterable) and not isinstance(interest, (str, bytes)):
        interest_values = [float(value) for value in interest if value is not None]
        if interest_values:
            financing = dict(payload.get("financing", {}))
            financing["discount_rate"] = float(interest_values[0])
            payload["financing"] = financing

    inputs, digest = _cached_parse_inputs(payload)
    return _cached_model_run(inputs, digest)


def _generate_excel_bytes(
    model: FinancialModel, outputs: FinancialOutputs, scenario_name: str
) -> bytes:
    """Return an Excel workbook representing the model results."""

    sections = collect_report_sections(model, outputs)
    report_name = "pharma_financial_model"
    slug = _scenario_slug(scenario_name)
    if slug and slug != "base":
        report_name = f"{report_name}_{slug}"
    data, _mime, _filename = generate_report(sections, "Excel", report_name=report_name)
    return data


def _render_projection_horizon(payload: dict) -> None:
    """Allow users to adjust the model start and end years via dropdowns."""

    current_years = [
        _parse_year_value(year, default=0)
        for year in payload.get("years", [])
        if year is not None
    ]

    if not current_years:
        fallback_start = datetime.now().year
        fallback_end = fallback_start + 9
        current_years = list(range(fallback_start, fallback_end + 1))
        payload["years"] = current_years

    current_start = int(current_years[0])
    current_end = int(current_years[-1])

    base_min = min(MIN_PROJECTION_YEAR, current_start, current_end)
    base_max = max(MAX_PROJECTION_YEAR, current_start, current_end)
    year_options = list(range(base_min, base_max + 1))

    start_index = year_options.index(current_start) if current_start in year_options else 0
    end_index = year_options.index(current_end) if current_end in year_options else len(year_options) - 1

    cols = st.columns([1, 1, 1])
    start_year = cols[0].selectbox(
        "Start Year",
        year_options,
        index=start_index,
        key="horizon_start_year",
        help="Select the first projection year for the financial model.",
    )
    end_year = cols[1].selectbox(
        "End Year",
        year_options,
        index=end_index,
        key="horizon_end_year",
        help="Select the final projection year for the financial model.",
    )

    if end_year < start_year:
        cols[2].error("End year must be greater than or equal to start year.")
        return

    if start_year == current_start and end_year == current_end:
        return

    new_years = list(range(int(start_year), int(end_year) + 1))
    payload["years"] = new_years
    labels = [str(year) for year in new_years]
    _align_payload_horizon(payload, labels, len(new_years), update_years=True)
    _initialise_session_payload(payload)
    _rerun()


def _select_or_create_option(
    container: Any,
    label: str,
    options: Sequence[str],
    key_prefix: str,
    current_value: str | None = None,
) -> str:
    """Render a select box that allows choosing from ``options`` or a custom value."""

    cleaned = [str(option).strip() for option in options if str(option).strip()]
    seen: dict[str, None] = {value: None for value in cleaned}
    ordered_options = list(seen.keys())

    value = (current_value or "").strip()
    if value and value not in seen:
        ordered_options.append(value)

    if not ordered_options:
        ordered_options.append("")

    option_list = ordered_options + ["Add new…"]
    default_index = option_list.index(value) if value in option_list else len(option_list) - 1

    selection = container.selectbox(
        label,
        option_list,
        index=default_index,
        key=f"{key_prefix}_select",
    )

    if selection == "Add new…":
        custom_default = "" if value in ordered_options else value
        custom_value = container.text_input(
            f"{label} (custom)",
            value=custom_default,
            key=f"{key_prefix}_custom",
        )
        return custom_value.strip()

    return selection.strip()


def _parse_year_value(label: str | int | float | None, default: int = 0) -> int:
    """Extract an integer year from ``label`` when possible."""

    if label is None:
        return default

    if isinstance(label, (int, float)):
        try:
            return int(label)
        except Exception:  # pragma: no cover - defensive conversion
            return default

    match = re.search(r"-?\d+", str(label))
    if match:
        try:
            return int(match.group())
        except Exception:  # pragma: no cover - defensive conversion
            return default

    return default


UTILITY_FLOAT_FIELDS = [
    "electricity_per_day",
    "electricity_rate",
    "water_per_day",
    "water_rate",
    "steam_per_hour",
    "steam_rate",
]

UTILITY_INT_FIELDS = [
    "electricity_days",
    "water_days",
    "steam_days",
    "steam_hours",
]


def _streamlit_runtime_exists() -> bool:
    """Return ``True`` when the Streamlit runtime has been initialised."""

    try:  # pragma: no cover - depends on Streamlit internals
        from streamlit.runtime import exists
    except Exception:  # pragma: no cover - runtime API unavailable
        exists = None  # type: ignore[assignment]

    if exists is not None:
        try:  # pragma: no cover - defensive against older Streamlit versions
            return bool(exists())
        except Exception:
            return False

    try:  # pragma: no cover - fallback for older Streamlit versions
        from streamlit.runtime.scriptrunner import get_script_run_ctx
    except Exception:
        get_script_run_ctx = None  # type: ignore[assignment]

    if get_script_run_ctx is None:
        try:  # pragma: no cover - older Streamlit versions
            from streamlit.scriptrunner import get_script_run_ctx
        except Exception:
            return False

    return get_script_run_ctx() is not None


def main() -> None:
    if not _streamlit_runtime_exists():  # pragma: no cover - requires Streamlit runner
        raise RuntimeError(
            "Streamlit runtime is not initialised. Launch the app with "
            "`streamlit run streamlit_app.py` to enable interactive inputs."
        )

    st.set_page_config(
        page_title="Pharmaceuticals Financial Model",
        page_icon="💊",
        layout="wide",
    )

    st.title("Pharmaceuticals Financial Model")
    st.caption(
        "Interactive financial modelling environment covering statements, "
        "scenario analysis, and Monte Carlo simulation."
    )

    config_container = st.container()
    download_container = st.container()

    inputs, digest = _resolve_inputs(config_container)
    model, outputs = _cached_model_run(inputs, digest)

    _render_excel_model_download(download_container, model, outputs)

    tabs = st.tabs(
        [
            "Input Landing Page",
            "Key Metrics Dashboard",
            "Financial Performance",
            "Financial Position",
            "Cash Flow Statement",
            "Sensitivity Analysis",
            "Scenario / IFs Analysis",
            "Monte Carlo Simulation",
            "Break-even & Payback",
        ]
    )

    with tabs[0]:
        _render_inputs_tab(inputs, model, outputs)
    with tabs[1]:
        _render_dashboard_tab(model, outputs)
    with tabs[2]:
        _render_income_statement(model, outputs)
    with tabs[3]:
        _render_statement_tab("Statement of Financial Position", outputs.balance_sheet)
    with tabs[4]:
        _render_statement_tab("Statement of Cash Flows", outputs.cash_flow)
    with tabs[5]:
        _render_sensitivity(outputs)
    with tabs[6]:
        _render_scenarios(outputs)
    with tabs[7]:
        _render_monte_carlo(outputs)
    with tabs[8]:
        _render_break_even(outputs)


def _resolve_inputs(container: DeltaGenerator) -> tuple[ModelInputs, str]:
    with container:
        if "input_payload" not in st.session_state:
            _initialise_session_payload(json.loads(DEFAULT_INPUT_JSON))

        payload = st.session_state["input_payload"]

    _ai_settings_to_payload(st.session_state.get("ai_settings", {}), payload)
    rows = st.session_state.setdefault(
        "core_assumption_rows", _payload_to_core_rows(payload)
    )
    synced_rows = _sync_core_rows_from_widgets(rows)
    if synced_rows != rows:
        st.session_state["core_assumption_rows"] = synced_rows
        rows = synced_rows
    _core_rows_to_payload(rows, payload)

    commission_rows = st.session_state.setdefault(
        "commission_rows", _payload_to_commission_rows(payload)
    )
    synced_commission = _sync_commission_rows_from_widgets(commission_rows)
    if synced_commission != commission_rows:
        st.session_state["commission_rows"] = synced_commission
        commission_rows = synced_commission
    _commission_rows_to_payload(commission_rows, payload)

    utility_entries = st.session_state.setdefault(
        "utility_entries", _payload_to_utility_entries(payload)
    )
    synced_utilities = _sync_utility_entries_from_widgets(utility_entries)
    if synced_utilities != utility_entries:
        st.session_state["utility_entries"] = synced_utilities
        utility_entries = synced_utilities
    _utility_entries_to_payload(utility_entries, payload)

    receivable_rows = st.session_state.setdefault(
        "receivable_rows", _payload_to_receivable_rows(payload)
    )
    synced_receivables = _sync_receivable_rows_from_widgets(receivable_rows, payload)
    if synced_receivables != receivable_rows:
        st.session_state["receivable_rows"] = synced_receivables
        receivable_rows = synced_receivables
    _receivable_rows_to_payload(receivable_rows, payload)

    inventory_rows = st.session_state.setdefault(
        "inventory_rows", _payload_to_inventory_rows(payload)
    )
    synced_inventory = _sync_inventory_rows_from_widgets(inventory_rows, payload)
    if synced_inventory != inventory_rows:
        st.session_state["inventory_rows"] = synced_inventory
        inventory_rows = synced_inventory
    _inventory_rows_to_payload(inventory_rows, payload)

    direct_rows = st.session_state.setdefault(
        "direct_labor_rows",
        _mapping_to_rows(payload.get("labor", {}).get("direct", {}), "Role", "Annual Cost"),
    )
    synced_direct = _sync_labor_rows_from_widgets("direct_labor_rows", direct_rows)
    if synced_direct != direct_rows:
        st.session_state["direct_labor_rows"] = synced_direct
        direct_rows = synced_direct
    _labor_rows_to_payload("direct", direct_rows, payload)

    indirect_rows = st.session_state.setdefault(
        "indirect_labor_rows",
        _mapping_to_rows(payload.get("labor", {}).get("indirect", {}), "Role", "Annual Cost"),
    )
    synced_indirect = _sync_labor_rows_from_widgets("indirect_labor_rows", indirect_rows)
    if synced_indirect != indirect_rows:
        st.session_state["indirect_labor_rows"] = synced_indirect
        indirect_rows = synced_indirect
    _labor_rows_to_payload("indirect", indirect_rows, payload)

    cost_rows = st.session_state.setdefault(
        "fixed_variable_rows", _payload_to_fixed_variable_rows(payload)
    )
    synced_costs = _sync_fixed_variable_rows_from_widgets(cost_rows)
    if synced_costs != cost_rows:
        st.session_state["fixed_variable_rows"] = synced_costs
        cost_rows = synced_costs
    _fixed_variable_rows_to_payload(cost_rows, payload)

    break_even_rows = st.session_state.setdefault(
        "break_even_rows", _payload_to_break_even_rows(payload)
    )
    synced_break_even = _sync_break_even_rows_from_widgets(break_even_rows)
    if synced_break_even != break_even_rows:
        st.session_state["break_even_rows"] = synced_break_even
        break_even_rows = synced_break_even
    _break_even_rows_to_payload(break_even_rows, payload)

    depreciation_rows = st.session_state.setdefault(
        "depreciation_rows", _payload_to_depreciation_rows(payload)
    )
    synced_depreciation = _sync_depreciation_rows_from_widgets(depreciation_rows, payload)
    if synced_depreciation != depreciation_rows:
        st.session_state["depreciation_rows"] = synced_depreciation
        depreciation_rows = synced_depreciation
    _depreciation_rows_to_payload(depreciation_rows, payload)

    inflation_rows = st.session_state.setdefault(
        "inflation_rows", _payload_to_inflation_rows(payload)
    )
    synced_inflation = _sync_inflation_rows_from_widgets(inflation_rows, payload)
    if synced_inflation != inflation_rows:
        st.session_state["inflation_rows"] = synced_inflation
        inflation_rows = synced_inflation
    _inflation_rows_to_payload(inflation_rows, payload)

    risk_rows = st.session_state.setdefault(
        "risk_rows", _payload_to_risk_rows(payload)
    )
    synced_risk = _sync_risk_rows_from_widgets(risk_rows, payload)
    if synced_risk != risk_rows:
        st.session_state["risk_rows"] = synced_risk
        risk_rows = synced_risk
    _risk_rows_to_payload(risk_rows, payload)

    sensitivity_rows = st.session_state.setdefault(
        "sensitivity_rows", _payload_to_sensitivity_rows(payload)
    )
    synced_sensitivity = _sync_sensitivity_rows_from_widgets(sensitivity_rows)
    if synced_sensitivity != sensitivity_rows:
        st.session_state["sensitivity_rows"] = synced_sensitivity
        sensitivity_rows = synced_sensitivity
    _sensitivity_rows_to_payload(sensitivity_rows, payload)

    senior_debt_rows = st.session_state.setdefault(
        "senior_debt_rows", _payload_to_debt_rows(payload, "senior_debt")
    )
    synced_senior = _sync_debt_rows_from_widgets("senior_debt", senior_debt_rows)
    if synced_senior != senior_debt_rows:
        st.session_state["senior_debt_rows"] = synced_senior
        senior_debt_rows = synced_senior
    _debt_rows_to_payload(senior_debt_rows, payload, "senior_debt")

    revolver_rows = st.session_state.setdefault(
        "revolver_rows", _payload_to_debt_rows(payload, "revolver")
    )
    synced_revolver = _sync_debt_rows_from_widgets("revolver", revolver_rows)
    if synced_revolver != revolver_rows:
        st.session_state["revolver_rows"] = synced_revolver
        revolver_rows = synced_revolver
    _debt_rows_to_payload(revolver_rows, payload, "revolver")

    overdraft_rows = st.session_state.setdefault(
        "overdraft_rows", _payload_to_debt_rows(payload, "overdraft")
    )
    synced_overdraft = _sync_debt_rows_from_widgets("overdraft", overdraft_rows)
    if synced_overdraft != overdraft_rows:
        st.session_state["overdraft_rows"] = synced_overdraft
        overdraft_rows = synced_overdraft
    _debt_rows_to_payload(overdraft_rows, payload, "overdraft")

    tax_rows = st.session_state.setdefault(
        "tax_entries", _payload_to_tax_entries(payload)
    )
    synced_tax = _sync_tax_entries_from_widgets(tax_rows)
    if synced_tax != tax_rows:
        st.session_state["tax_entries"] = synced_tax
        tax_rows = synced_tax
    tax_payload = payload.setdefault("tax", {})
    if isinstance(tax_payload, dict):
        tax_payload["rate"] = float(
            _get_widget_number(
                "tax_base_rate", tax_payload.get("rate", 0.0), float
            )
        )
        tax_payload["timing_adjustment"] = float(
            _get_widget_number(
                "tax_timing", tax_payload.get("timing_adjustment", 0.0), float
            )
        )
    _tax_entries_to_payload(
        tax_rows,
        tax_payload,
        payload.get("years", []),
        float(tax_payload.get("rate", 0.0)),
    )

    #
    # Commit the latest widget-driven edits back into session state before
    # parsing the model inputs.  Cloning the payload ensures we hold a stable
    # copy that no longer aliases the temporary dictionaries used during the
    # Streamlit render cycle.  Without this step the financial engine could
    # observe a stale view of the assumptions whenever subsequent helpers
    # mutated the shared dictionary in-place after caching had already
    # occurred.  By persisting a deep copy we guarantee that parse_inputs and
    # downstream consumers always receive the exact data visible in the input
    # tables.
    #
    committed_payload = _clone_payload(payload)
    st.session_state["input_payload"] = committed_payload

    inputs, digest = _cached_parse_inputs(committed_payload)
    st.session_state["input_fingerprint"] = digest
    return inputs, digest


def _load_payload_from_bytes(data: bytes, suffix: str) -> Mapping[str, object]:
    """Load a payload mapping from uploaded file bytes."""

    suffix = suffix or ".json"
    if suffix in {".json", ""}:
        return _load_payload_from_text(data.decode("utf-8"))
    if suffix == ".csv":
        return _load_payload_from_csv(data)
    if suffix in {".xlsx", ".xls"}:
        return _load_payload_from_excel(data)
    if suffix == ".docx":
        return _load_payload_from_docx(data)
    if suffix == ".pdf":
        return _load_payload_from_pdf(data)
    raise ValueError(f"Unsupported file type: {suffix}")


def _load_payload_from_text(text: str) -> Mapping[str, object]:
    stripped = text.strip()
    if not stripped:
        raise ValueError("Uploaded file was empty.")
    try:
        return json.loads(stripped)
    except json.JSONDecodeError as exc:  # pragma: no cover - invalid user input
        raise ValueError("Uploaded document does not contain valid JSON assumptions.") from exc


def _extract_json_fragment(text: str) -> str:
    if "{" in text and "}" in text:
        start = text.find("{")
        end = text.rfind("}")
        if end > start:
            return text[start : end + 1]
    return text


def _load_payload_from_csv(data: bytes) -> Mapping[str, object]:
    text = data.decode("utf-8-sig")
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        fragment = _extract_json_fragment(text)
        if fragment and fragment != text:
            return _load_payload_from_text(fragment)

        reader = csv.reader(io.StringIO(text))
        cells: list[str] = []
        for row in reader:
            cells.extend(cell for cell in row if cell is not None)
        joined = _extract_json_fragment("".join(cells).strip())
        if not joined:
            raise ValueError("CSV file did not contain any usable JSON text.")
        return _load_payload_from_text(joined)


def _load_payload_from_excel(data: bytes) -> Mapping[str, object]:
    if load_workbook is None:  # pragma: no cover - optional dependency path
        raise ValueError("Excel support requires the 'openpyxl' package to be installed.")

    workbook = load_workbook(filename=io.BytesIO(data), data_only=True)
    text_parts: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                text_parts.append(str(value))
    combined = _extract_json_fragment("\n".join(text_parts).strip())
    if not combined:
        raise ValueError("Excel file did not contain any readable text.")
    return _load_payload_from_text(combined)


def _load_payload_from_docx(data: bytes) -> Mapping[str, object]:
    if Document is None:  # pragma: no cover - optional dependency path
        raise ValueError("Word support requires the 'python-docx' package to be installed.")

    document = Document(io.BytesIO(data))
    text = _extract_json_fragment(
        "\n".join(paragraph.text for paragraph in document.paragraphs).strip()
    )
    if not text:
        raise ValueError("Word document did not contain any readable text.")
    return _load_payload_from_text(text)


def _load_payload_from_pdf(data: bytes) -> Mapping[str, object]:
    if PdfReader is None:  # pragma: no cover - optional dependency path
        raise ValueError("PDF support requires the 'PyPDF2' package to be installed.")

    reader = PdfReader(io.BytesIO(data))
    text_parts: list[str] = []
    for page in reader.pages:
        extracted = page.extract_text() or ""
        text_parts.append(extracted)
    combined = _extract_json_fragment("\n".join(text_parts).strip())
    if not combined:
        raise ValueError("PDF file did not contain any readable text.")
    return _load_payload_from_text(combined)


def _render_excel_model_download(
    container: DeltaGenerator, base_model: FinancialModel, base_outputs: FinancialOutputs
) -> None:
    with container:
        st.markdown("### Excel Model Download")

        payload = st.session_state.get("input_payload") or {}
        scenario_options = _scenario_options(payload)
        stored_selection = st.session_state.get("excel_scenario_selection")
        default_index = 0
        if isinstance(stored_selection, str) and stored_selection in scenario_options:
            default_index = scenario_options.index(stored_selection)

        selected_scenario = st.selectbox(
            "Select scenario for Excel export",
            scenario_options,
            index=default_index,
            key="excel_scenario_selection",
        )

        snapshot = st.session_state.get("input_snapshot")
        if snapshot is None:
            snapshot = copy.deepcopy(payload)
            st.session_state["input_snapshot"] = snapshot

        model, results = _ensure_scenario_payload(
            selected_scenario, snapshot, base_model, base_outputs
        )
        st.session_state["model_results"] = (model, results)

        excel_map: Dict[str, bytes] = st.session_state.setdefault("excel_bytes_map", {})
        excel_bytes = excel_map.get(selected_scenario)

        model.scenario = selected_scenario

        download_container = st.container()
        with download_container:
            if not excel_bytes:
                if st.button(
                    "Prepare Excel Model",
                    key=f"prepare_excel_{selected_scenario.lower()}",
                ):
                    with st.spinner("Preparing Excel workbook..."):
                        excel_bytes = _generate_excel_bytes(
                            model, results, selected_scenario
                        )
                    excel_map[selected_scenario] = excel_bytes
                    st.session_state.excel_bytes_map = excel_map
            if excel_bytes:
                st.download_button(
                    "Download Excel Model",
                    data=excel_bytes,
                    file_name="Ecommerce_Financial_Model.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                if st.button(
                    "Clear Prepared Excel",
                    key=f"clear_excel_{selected_scenario.lower()}",
                ):
                    excel_map.pop(selected_scenario, None)
                    st.session_state.excel_bytes_map = excel_map
                    excel_bytes = None
            if not excel_bytes:
                st.info("Click 'Prepare Excel Model' to generate the workbook for download.")


def _render_inputs_tab(
    inputs: ModelInputs, base_model: FinancialModel, base_outputs: FinancialOutputs
) -> None:
    payload = st.session_state["input_payload"]

    st.markdown("### AI & Machine Learning Configuration")
    _render_ai_settings(payload)
    _ai_settings_to_payload(st.session_state.get("ai_settings", {}), payload)

    st.markdown("### Projection Horizon")
    _render_projection_horizon(payload)

    st.subheader("Core Assumptions")
    rows: List[dict] = st.session_state.get("core_assumption_rows", [])
    _prime_core_widget_state(rows)

    if not rows:
        st.info("No core assumptions configured. Use the form below to add entries.")

    production_estimate = payload.get("production_estimate", {})
    total_unit_defaults = payload.get("total_production_units", {})
    capacity_defaults = payload.get("production_capacity", {})
    inflation_factors = _inflation_factors_from_payload(payload)
    risk_factors = _risk_factors_from_payload(payload)

    updated_rows: list[dict] = []
    for index, row in enumerate(rows):
        container = st.container()
        with container:
            row_product = str(row.get("Product", ""))
            default_units = float(row.get("Total Production Units", 0.0))
            if default_units == 0.0:
                if row_product in total_unit_defaults:
                    default_units = float(total_unit_defaults[row_product])
                elif isinstance(production_estimate, Mapping) and row_product in production_estimate:
                    default_units = sum(
                        float(value)
                        for value in production_estimate.get(row_product, [])
                    )
            default_capacity = float(row.get("Max Capacity", 0.0))
            if default_capacity == 0.0 and row_product in capacity_defaults:
                default_capacity = float(capacity_defaults[row_product])

            cols = st.columns([3, 2, 2, 2, 2, 2, 2, 2, 2, 1])
            desc_key = f"core_desc_{index}"
            prod_key = f"core_prod_{index}"
            sell_key = f"core_sell_{index}"
            freight_key = f"core_freight_{index}"
            markup_key = f"core_markup_{index}"
            units_key = f"core_units_{index}"
            capacity_key = f"core_capacity_{index}"

            _ensure_widget_default(desc_key, row.get("Product", ""))
            _ensure_widget_default(prod_key, float(row.get("Production Cost", 0.0)))
            _ensure_widget_default(sell_key, float(row.get("Selling Price", 0.0)))
            _ensure_widget_default(freight_key, float(row.get("Freight Cost", 0.0)))
            _ensure_widget_default(markup_key, float(row.get("Markup", 0.0)))
            _ensure_widget_default(units_key, default_units)
            _ensure_widget_default(capacity_key, default_capacity)

            description = cols[0].text_input(
                "Description",
                value=str(st.session_state.get(desc_key, "")),
                key=desc_key,
                help="Name of the product or assumption this row represents.",
            )
            production = cols[1].number_input(
                "Production Cost",
                value=float(st.session_state.get(prod_key, 0.0)),
                key=prod_key,
                step=0.001,
                format="%.4f",
            )
            selling = cols[2].number_input(
                "Selling Price",
                value=float(st.session_state.get(sell_key, 0.0)),
                key=sell_key,
                step=0.001,
                format="%.4f",
            )
            freight = cols[3].number_input(
                "Freight Cost",
                value=float(st.session_state.get(freight_key, 0.0)),
                key=freight_key,
                step=0.001,
                format="%.4f",
            )
            markup = cols[4].number_input(
                "Markup",
                value=float(st.session_state.get(markup_key, 0.0)),
                key=markup_key,
                step=0.01,
                format="%.2f",
            )
            total_units = cols[5].number_input(
                "Total Production Units",
                value=float(st.session_state.get(units_key, default_units)),
                key=units_key,
                step=1.0,
                format="%.4f",
                min_value=0.0,
            )
            max_capacity = cols[6].number_input(
                "Max Capacity",
                value=float(st.session_state.get(capacity_key, default_capacity)),
                key=capacity_key,
                step=1.0,
                format="%.4f",
                min_value=0.0,
            )

            clamped_units = float(total_units)
            if max_capacity > 0.0 and clamped_units > max_capacity + 1e-9:
                cols[5].error("Capacity exceeded")
                clamped_units = max_capacity

            scaled_series = _scaled_production_series(
                description.strip(),
                clamped_units,
                payload.get("years", []),
                production_estimate,
            )
            total_revenue = float(clamped_units) * float(selling)
            total_cost = float(clamped_units) * (
                float(production) + float(freight) + float(markup)
            )

            revenue_key = f"core_revenue_{index}"
            cost_key = f"core_cost_{index}"
            _set_widget_value(revenue_key, total_revenue)
            _set_widget_value(cost_key, total_cost)

            cols[7].number_input(
                "Total Revenue",
                key=revenue_key,
                format="%.4f",
                disabled=True,
            )
            cols[8].number_input(
                "Total Cost",
                key=cost_key,
                format="%.4f",
                disabled=True,
            )
            if cols[9].button("Remove", key=f"core_remove_{index}"):
                del rows[index]
                st.session_state["core_assumption_rows"] = rows
                _prime_core_widget_state(rows)
                _rerun()

        updated_rows.append(
            {
                "Product": description.strip(),
                "Production Cost": production,
                "Selling Price": selling,
                "Freight Cost": freight,
                "Markup": markup,
                "Total Production Units": clamped_units,
                "Max Capacity": max_capacity,
                "Total Revenue": total_revenue,
                "Total Cost": total_cost,
            }
        )

    if updated_rows != rows:
        st.session_state["core_assumption_rows"] = updated_rows
        _prime_core_widget_state(updated_rows)

    st.markdown("#### Add a core assumption")
    with st.form("add_core_assumption"):
        new_description = st.text_input(
            "Description", key="core_new_description", help="Label for the new row."
        )
        new_production = st.number_input(
            "Production Cost", value=0.0, step=0.001, format="%.4f", key="core_new_prod"
        )
        new_selling = st.number_input(
            "Selling Price", value=0.0, step=0.001, format="%.4f", key="core_new_sell"
        )
        new_freight = st.number_input(
            "Freight Cost", value=0.0, step=0.001, format="%.4f", key="core_new_freight"
        )
        new_markup = st.number_input(
            "Markup", value=0.0, step=0.01, format="%.2f", key="core_new_markup"
        )
        new_units = st.number_input(
            "Total Production Units",
            value=0.0,
            step=1.0,
            format="%.4f",
            key="core_new_units",
            min_value=0.0,
        )
        new_capacity = st.number_input(
            "Max Capacity",
            value=0.0,
            step=1.0,
            format="%.4f",
            key="core_new_capacity",
            min_value=0.0,
        )
        submitted = st.form_submit_button("Add")

    if submitted:
        if not new_description.strip():
            st.warning("Description is required to add a core assumption.")
        else:
            clamped_units = new_units
            if new_capacity > 0.0 and new_units > new_capacity + 1e-9:
                st.error("Capacity exceeded")
                clamped_units = new_capacity
            total_revenue = clamped_units * new_selling
            total_cost = clamped_units * (new_production + new_freight + new_markup)
            rows.append(
                {
                    "Product": new_description.strip(),
                    "Production Cost": new_production,
                    "Selling Price": new_selling,
                    "Freight Cost": new_freight,
                    "Markup": new_markup,
                    "Total Production Units": clamped_units,
                    "Max Capacity": new_capacity,
                    "Total Revenue": total_revenue,
                    "Total Cost": total_cost,
                }
            )
            st.session_state["core_assumption_rows"] = rows
            _prime_core_widget_state(rows)
            for key in (
                "core_new_description",
                "core_new_prod",
                "core_new_sell",
                "core_new_freight",
                "core_new_markup",
                "core_new_units",
                "core_new_capacity",
            ):
                st.session_state.pop(key, None)
            _rerun()

    st.markdown("### Distributors Commission Input Table")
    _render_distributor_commission(payload)

    st.markdown("### Direct Labour Structure")
    _render_labor_section("direct", "direct_labor_rows", payload)

    st.markdown("### Indirect Labour Structure")
    _render_labor_section("indirect", "indirect_labor_rows", payload)

    st.markdown("### Fixed & Variable Costs Input Table")
    _render_fixed_variable_costs(payload)

    _render_utility_schedule(payload)

    st.markdown("### Accounts Receivable Input Table")
    _render_receivable_inputs(payload)

    st.markdown("### Inventory & Accounts Payable Input Table")
    _render_inventory_inputs(payload)

    st.markdown("### Fixed Assets Schedule")
    _render_depreciation_schedule(payload)

    st.markdown("### Cost & Financing Assumptions")
    _render_cost_and_financing(payload)

    st.markdown("### Tax Schedule")
    _render_tax_schedule(payload)

    st.markdown("### Inflation Schedule")
    _render_inflation_schedule(payload)

    st.markdown("### Risk Schedule")
    _render_risk_schedule(payload)

    st.markdown("### AI & Machine Learning Summary")
    _render_ai_summary(payload)

    _core_rows_to_payload(st.session_state.get("core_assumption_rows", []), payload)
    _commission_rows_to_payload(st.session_state.get("commission_rows", []), payload)
    _utility_entries_to_payload(st.session_state.get("utility_entries", []), payload)
    _receivable_rows_to_payload(st.session_state.get("receivable_rows", []), payload)
    _inventory_rows_to_payload(st.session_state.get("inventory_rows", []), payload)
    _depreciation_rows_to_payload(st.session_state.get("depreciation_rows", []), payload)
    _risk_rows_to_payload(st.session_state.get("risk_rows", []), payload)
    _inflation_rows_to_payload(st.session_state.get("inflation_rows", []), payload)
    _debt_rows_to_payload(st.session_state.get("senior_debt_rows", []), payload, "senior_debt")
    _debt_rows_to_payload(st.session_state.get("revolver_rows", []), payload, "revolver")
    _debt_rows_to_payload(st.session_state.get("overdraft_rows", []), payload, "overdraft")
    st.session_state["input_payload"] = payload


def _render_dashboard_tab(model: FinancialModel, outputs: FinancialOutputs) -> None:
    income = _with_year(outputs.income_statement)
    supports_plotly = px is not None and pd is not None

    if not supports_plotly:
        st.warning(
            "Plotly visualisations unavailable. Displaying financial metrics as tables instead."
        )
        st.dataframe(income, use_container_width=True)
    else:
        col1, col2 = st.columns(2)
        with col1:
            fig_revenue = px.line(income, x="Year", y="Net Revenue", title="Net Revenue")
            st.plotly_chart(fig_revenue, use_container_width=True)
        with col2:
            fig_ebitda = px.line(income, x="Year", y="EBITDA", title="EBITDA")
            st.plotly_chart(fig_ebitda, use_container_width=True)

    st.markdown("### Investment Metrics")
    metric_pairs = _extract_metric_pairs(outputs.summary_metrics)
    if not metric_pairs:
        st.info("No investment metrics were generated for the current assumptions.")
    else:
        metric_cols = st.columns(len(metric_pairs))
        for col, (name, value) in zip(metric_cols, metric_pairs):
            with col:
                formatted = _format_number(value)
                st.metric(label=name, value=formatted)

    st.markdown("### Goal Seek Metric")
    goal_data = _ensure_dataframe(outputs.goal_seek)
    if isinstance(goal_data, list):
        if not goal_data:
            st.caption("No goal seek configuration provided in the assumptions.")
        else:
            st.dataframe(goal_data, use_container_width=True)
    else:
        if hasattr(goal_data, "empty") and getattr(goal_data, "empty"):
            st.caption("No goal seek configuration provided in the assumptions.")
        elif pd is not None and isinstance(goal_data, pd.DataFrame):
            display = goal_data.copy()
            index_name = display.index.name or "Metric"
            display = display.reset_index().rename(columns={index_name: "Metric"})
            metric_columns = st.columns(len(display))
            for column, (_, row) in zip(metric_columns, display.iterrows()):
                label = str(row.get("Metric", "Goal"))
                actual_value = float(row.get("Actual", float("nan")))
                target_value = float(row.get("Target", float("nan")))
                delta_value = actual_value - target_value
                with column:
                    st.metric(
                        label=f"{label} Actual",
                        value=_format_number(actual_value),
                        delta=_format_number(delta_value),
                    )
            st.dataframe(display, use_container_width=True)
        else:
            st.dataframe(goal_data, use_container_width=True)

    st.markdown("### Working Capital Schedule")
    try:
        working_capital = model.working_capital_schedule()
        st.dataframe(_with_year(working_capital), use_container_width=True)
        st.caption(
            "Working capital balances reconcile receivables, inventory, and payables "
            "with the statement of financial position while showing year-over-year "
            "changes."
        )
    except Exception as exc:  # pragma: no cover - defensive user feedback
        st.warning(f"Unable to compute working capital schedule: {exc}")

    st.markdown("### Inventory Schedule")
    try:
        inventory_table = model.inventory_schedule()
        st.dataframe(_with_year(inventory_table), use_container_width=True)
        st.caption(
            "Inventory is derived as cost of sales divided by calendar days and "
            "multiplied by the configured inventory days, matching the balance "
            "sheet totals."
        )
    except Exception as exc:  # pragma: no cover - defensive user feedback
        st.warning(f"Unable to compute inventory schedule: {exc}")

    st.markdown("### Key Analysis Dashboard")
    if not supports_plotly:
        st.info(
            "Install pandas and plotly to view charts. Displaying analytical tables instead."
        )

        st.markdown("#### Sensitivity Analysis")
        if outputs.sensitivity_results:
            for variable, table in outputs.sensitivity_results.items():
                st.markdown(f"- **{variable}**")
                st.dataframe(_ensure_dataframe(table), use_container_width=True)
        else:
            st.caption("No sensitivity configurations provided.")

        st.markdown("#### Scenario / IFs Analysis")
        if outputs.scenario_results:
            for name, table in outputs.scenario_results.items():
                st.markdown(f"- **{name}**")
                st.dataframe(_with_year(table), use_container_width=True)
        else:
            st.caption("No scenarios configured in the assumptions.")

        st.markdown("#### Break-even Analysis")
        st.dataframe(_ensure_dataframe(outputs.break_even), use_container_width=True)

        st.markdown("#### Payback Schedule")
        st.dataframe(_with_year(outputs.payback), use_container_width=True)

        st.markdown("#### Discounted Payback Schedule")
        st.dataframe(_with_year(outputs.discounted_payback), use_container_width=True)

        st.markdown("#### AI & Machine Learning Insights")
        _render_ai_dashboard(outputs.ai_insights)
        return

    # Sensitivity Analysis charts
    if outputs.sensitivity_results:
        st.markdown("#### Sensitivity Analysis")
        for variable, table in outputs.sensitivity_results.items():
            frame = _ensure_dataframe(table)
            if isinstance(frame, pd.DataFrame):
                frame = frame.reset_index()
            else:
                frame = pd.DataFrame(frame)

            x_column = "Multiplier" if "Multiplier" in frame.columns else frame.columns[0]
            melted = frame.melt(
                id_vars=[x_column],
                value_vars=[col for col in ["NPV", "IRR"] if col in frame.columns],
                var_name="Metric",
                value_name="Value",
            )
            title = f"Sensitivity: {variable.replace('_', ' ').title()}"
            fig = px.line(
                melted,
                x=x_column,
                y="Value",
                color="Metric",
                markers=True,
                title=title,
            )
            st.plotly_chart(fig, use_container_width=True)
    else:
        st.caption("No sensitivity configurations provided.")

    # Scenario / IFs Analysis charts
    if outputs.scenario_results:
        st.markdown("#### Scenario / IFs Analysis")
        scenario_frames: list[pd.DataFrame] = []
        for name, table in outputs.scenario_results.items():
            frame = _with_year(table)
            if isinstance(frame, pd.DataFrame):
                scenario_frame = frame.copy()
            else:
                scenario_frame = pd.DataFrame(frame)
            if "Year" not in scenario_frame.columns:
                scenario_frame = scenario_frame.reset_index().rename(columns={"index": "Year"})
            scenario_frame["Scenario"] = name
            scenario_frames.append(scenario_frame)

        if scenario_frames:
            combined = pd.concat(scenario_frames, ignore_index=True)
            if "Net Revenue" in combined.columns:
                fig = px.line(
                    combined,
                    x="Year",
                    y="Net Revenue",
                    color="Scenario",
                    title="Scenario Net Revenue",
                )
                st.plotly_chart(fig, use_container_width=True)
            if "Net Income" in combined.columns:
                fig_income = px.line(
                    combined,
                    x="Year",
                    y="Net Income",
                    color="Scenario",
                    title="Scenario Net Income",
                )
                st.plotly_chart(fig_income, use_container_width=True)
        else:
            st.caption("No scenarios configured in the assumptions.")
    else:
        st.caption("No scenarios configured in the assumptions.")

    # Break-even chart
    st.markdown("#### Break-even Analysis")
    break_even_df = _ensure_dataframe(outputs.break_even)
    if isinstance(break_even_df, pd.DataFrame):
        break_even_frame = break_even_df.reset_index().rename(columns={"index": "Product"})
    else:
        break_even_frame = pd.DataFrame(break_even_df)
    y_column = "Break-even Units" if "Break-even Units" in break_even_frame.columns else break_even_frame.columns[-1]
    fig_break_even = px.bar(
        break_even_frame,
        x="Product",
        y=y_column,
        title="Break-even Units by Product",
    )
    st.plotly_chart(fig_break_even, use_container_width=True)

    # Payback charts
    st.markdown("#### Payback Schedule")
    payback_df = _with_year(outputs.payback)
    if isinstance(payback_df, pd.DataFrame):
        payback_frame = payback_df
    else:
        payback_frame = pd.DataFrame(payback_df)
    fig_payback = px.line(
        payback_frame,
        x="Year",
        y="Cumulative",
        title="Cumulative Payback",
        markers=True,
    )
    st.plotly_chart(fig_payback, use_container_width=True)

    discounted_df = _with_year(outputs.discounted_payback)
    if isinstance(discounted_df, pd.DataFrame):
        discounted_frame = discounted_df
    else:
        discounted_frame = pd.DataFrame(discounted_df)
    fig_discounted = px.line(
        discounted_frame,
        x="Year",
        y="Cumulative",
        title="Discounted Cumulative Payback",
        markers=True,
    )
    st.plotly_chart(fig_discounted, use_container_width=True)

    st.markdown("#### AI & Machine Learning Insights")
    _render_ai_dashboard(outputs.ai_insights)


def _render_statement_tab(title: str, table) -> None:
    st.subheader(title)
    if isinstance(table, Table):
        display = table.rounded(0)
    elif pd is not None and isinstance(table, pd.DataFrame):
        display = table.round(0)
    else:
        display = table

    display_with_year = _with_year(display)
    st.dataframe(display_with_year, use_container_width=True)

    if px is None or pd is None:
        st.caption("Install pandas and plotly to unlock interactive analytics for this statement.")
        return

    if isinstance(display_with_year, pd.DataFrame):
        frame = display_with_year
    else:
        frame = pd.DataFrame(display_with_year)

    numeric_columns = [
        column
        for column in frame.columns
        if column != "Year" and pd.api.types.is_numeric_dtype(frame[column])
    ]
    if not numeric_columns:
        return

    headline_columns = numeric_columns[:4]
    if headline_columns:
        st.markdown("#### Analytical Trends")
        columns = st.columns(min(len(headline_columns), 2))
        for idx, column in enumerate(headline_columns):
            with columns[idx % len(columns)]:
                fig = px.line(
                    frame,
                    x="Year",
                    y=column,
                    markers=True,
                    title=f"{column} Trend",
                )
                st.plotly_chart(fig, use_container_width=True)

    if len(numeric_columns) > len(headline_columns):
        remaining = numeric_columns[len(headline_columns) :]
        melt_frame = frame.melt(id_vars=["Year"], value_vars=remaining, var_name="Metric", value_name="Value")
        fig = px.line(
            melt_frame,
            x="Year",
            y="Value",
            color="Metric",
            markers=True,
            title="Additional Statement Metrics",
        )
        st.plotly_chart(fig, use_container_width=True)


def _render_income_statement(model: FinancialModel, outputs: FinancialOutputs) -> None:
    st.subheader("Statement of Financial Performance")
    rounded_income = outputs.income_statement.rounded(0, exclude_keywords=("Margin", "Return"))
    income_frame = _with_year(rounded_income)
    if pd is not None and isinstance(income_frame, pd.DataFrame):
        display_frame = income_frame.drop(columns=["Depreciation"], errors="ignore")
    elif isinstance(income_frame, list):
        display_frame = [
            {key: value for key, value in row.items() if key != "Depreciation"}
            for row in income_frame
        ]
    else:
        display_frame = income_frame
    st.dataframe(display_frame, use_container_width=True)

    if px is not None and pd is not None:
        if isinstance(display_frame, pd.DataFrame):
            frame = display_frame
        else:
            frame = pd.DataFrame(display_frame)
        if "Year" in frame.columns:
            line_metrics = [
                column
                for column in ["Net Revenue", "Gross Profit", "EBITDA", "Net Income"]
                if column in frame.columns
            ]
            if line_metrics:
                st.markdown("#### Profit & Loss Trends")
                trend_data = frame[["Year", *line_metrics]]
                trend_frame = trend_data.melt(id_vars=["Year"], var_name="Metric", value_name="Value")
                fig_income = px.line(
                    trend_frame,
                    x="Year",
                    y="Value",
                    color="Metric",
                    markers=True,
                    title="Income Statement Highlights",
                )
                st.plotly_chart(fig_income, use_container_width=True)

            if "EBITDA Margin" in frame.columns:
                fig_margin = px.line(
                    frame,
                    x="Year",
                    y="EBITDA Margin",
                    markers=True,
                    title="EBITDA Margin",
                )
                st.plotly_chart(fig_margin, use_container_width=True)

    st.markdown("#### Gross Revenue Schedule")
    try:
        revenue_schedule = model.revenue_schedule()
    except Exception as exc:  # pragma: no cover - defensive guard for runtime issues
        st.warning(f"Unable to calculate gross revenue schedule: {exc}")
    else:
        revenue_frame = _with_year(revenue_schedule)
        st.dataframe(revenue_frame, use_container_width=True)
        st.caption(
            "Gross Revenue is decomposed into product-level sales, distributor commissions, "
            "and resulting net revenue."
        )
        if px is not None and pd is not None:
            if isinstance(revenue_frame, pd.DataFrame):
                frame = revenue_frame
            else:
                frame = pd.DataFrame(revenue_frame)
            numeric_columns = [
                column
                for column in frame.columns
                if column not in {"Year", "Product"} and pd.api.types.is_numeric_dtype(frame[column])
            ]
            if "Product" in frame.columns and numeric_columns:
                st.markdown("##### Revenue by Product")
                product_frame = frame.melt(
                    id_vars=["Year", "Product"],
                    value_vars=numeric_columns,
                    var_name="Metric",
                    value_name="Value",
                )
                fig_product = px.bar(
                    product_frame,
                    x="Year",
                    y="Value",
                    color="Product",
                    facet_row="Metric",
                    barmode="stack",
                    title="Revenue Composition",
                )
                st.plotly_chart(fig_product, use_container_width=True)
            elif numeric_columns:
                st.markdown("##### Revenue Drivers")
                melt_frame = frame.melt(id_vars=["Year"], value_vars=numeric_columns, var_name="Metric", value_name="Value")
                fig_revenue = px.line(
                    melt_frame,
                    x="Year",
                    y="Value",
                    color="Metric",
                    markers=True,
                    title="Revenue Schedule",
                )
                st.plotly_chart(fig_revenue, use_container_width=True)

    st.markdown("#### Total Expenses Schedule")
    try:
        expense_schedule = model.cost_structure()
    except Exception as exc:  # pragma: no cover - defensive guard for runtime issues
        st.warning(f"Unable to calculate total expenses schedule: {exc}")
        return

    expense_frame = _with_year(expense_schedule)
    st.dataframe(expense_frame, use_container_width=True)
    st.caption(
        "Total Expenses comprise raw materials, utilities, direct labour, cost of sales, "
        "and general & administrative costs."
    )
    if px is not None and pd is not None:
        if isinstance(expense_frame, pd.DataFrame):
            frame = expense_frame
        else:
            frame = pd.DataFrame(expense_frame)
        numeric_columns = [
            column
            for column in frame.columns
            if column != "Year" and pd.api.types.is_numeric_dtype(frame[column])
        ]
        if numeric_columns:
            trend_frame = frame.melt(id_vars=["Year"], value_vars=numeric_columns, var_name="Expense", value_name="Value")
            fig_expense = px.area(
                trend_frame,
                x="Year",
                y="Value",
                color="Expense",
                groupnorm="fraction",
                title="Expense Mix Over Time",
            )
            st.plotly_chart(fig_expense, use_container_width=True)


def _render_ai_dashboard(ai_insights: Optional[AIInsights]) -> None:
    if ai_insights is None:
        st.caption(
            "AI configuration not available. Enable AI enhancements on the Input Landing Page "
            "to unlock machine-generated commentary."
        )
        return

    if ai_insights.ml_forecast is not None:
        st.dataframe(_with_year(ai_insights.ml_forecast), use_container_width=True)
    else:
        st.caption(
            "Machine-learning forecasts are unavailable. Adjust the forecast horizon or ensure "
            "net revenue data is present."
        )

    summary = (ai_insights.generative_summary or "").strip()
    if summary:
        st.write(summary)
    else:
        st.caption("No generative summary returned for the current configuration.")

    if ai_insights.metadata:
        with st.expander("AI Metadata"):
            st.json(ai_insights.metadata)


def _render_sensitivity(outputs: FinancialOutputs) -> None:
    st.subheader("Sensitivity Analysis")
    payload = st.session_state.get("input_payload")
    if payload is None:
        payload = {}
        st.session_state["input_payload"] = payload
    st.markdown("### Sensitivity Analysis Configuration")
    _render_sensitivity_inputs(payload)
    st.session_state["input_payload"] = payload

    if not outputs.sensitivity_results:
        st.info("No sensitivity configurations provided in the assumptions file.")
        return

    supports_plotly = px is not None and pd is not None

    for variable, df in outputs.sensitivity_results.items():
        st.markdown(f"#### {variable}")
        frame = _with_year(df)
        st.dataframe(frame, use_container_width=True)

        if supports_plotly:
            if isinstance(frame, pd.DataFrame):
                plot_frame = frame
            else:
                plot_frame = pd.DataFrame(frame)
            numeric_columns = [
                column
                for column in plot_frame.columns
                if column not in {"Year", "Scenario"} and pd.api.types.is_numeric_dtype(plot_frame[column])
            ]
            if numeric_columns:
                if "Year" in plot_frame.columns:
                    long_frame = plot_frame.melt(id_vars=["Year"], value_vars=numeric_columns, var_name="Metric", value_name="Value")
                    fig = px.line(
                        long_frame,
                        x="Year",
                        y="Value",
                        color="Metric",
                        markers=True,
                        title=f"{variable} Sensitivity Trend",
                    )
                else:
                    long_frame = plot_frame.melt(value_vars=numeric_columns, var_name="Metric", value_name="Value")
                    fig = px.bar(
                        long_frame,
                        x="Metric",
                        y="Value",
                        title=f"{variable} Sensitivity Comparison",
                    )
                st.plotly_chart(fig, use_container_width=True)


def _render_scenarios(outputs: FinancialOutputs) -> None:
    st.subheader("Scenario / IFs Analysis")
    payload = st.session_state.get("input_payload")
    if payload is None:
        payload = {}
        st.session_state["input_payload"] = payload
    st.markdown("### Goal Seek Configuration")
    _render_goal_seek(payload)
    st.markdown("### Scenario / IFs Configuration")
    _render_scenario_inputs(payload)
    st.markdown("### Scenario Tool Configuration")
    _render_scenario_tool_inputs(payload)
    st.session_state["input_payload"] = payload

    supports_plotly = px is not None and pd is not None

    if not outputs.scenario_results:
        st.info("No scenario configurations provided in the assumptions file.")
    else:
        for name, df in outputs.scenario_results.items():
            st.markdown(f"#### {name}")
            frame = _with_year(df)
            st.dataframe(frame, use_container_width=True)

            if supports_plotly:
                if isinstance(frame, pd.DataFrame):
                    plot_frame = frame
                else:
                    plot_frame = pd.DataFrame(frame)
                numeric_columns = [
                    column
                    for column in plot_frame.columns
                    if column not in {"Year", "Scenario"} and pd.api.types.is_numeric_dtype(plot_frame[column])
                ]
                if numeric_columns:
                    if "Year" in plot_frame.columns:
                        long_frame = plot_frame.melt(id_vars=["Year"], value_vars=numeric_columns, var_name="Metric", value_name="Value")
                        fig = px.line(
                            long_frame,
                            x="Year",
                            y="Value",
                            color="Metric",
                            markers=True,
                            title=f"{name} Scenario Results",
                        )
                    else:
                        long_frame = plot_frame.melt(value_vars=numeric_columns, var_name="Metric", value_name="Value")
                        fig = px.bar(
                            long_frame,
                            x="Metric",
                            y="Value",
                            title=f"{name} Scenario Comparison",
                        )
                    st.plotly_chart(fig, use_container_width=True)

    st.markdown("### Scenario Tool Insights")
    if outputs.scenario_tool_results:
        for key, result in outputs.scenario_tool_results.items():
            label = SCENARIO_TOOL_LABELS.get(key, key.replace("_", " ").title())
            st.markdown(f"#### {label}")
            st.dataframe(_ensure_dataframe(result.rows), use_container_width=True)
            st.caption(result.interpretation)
    else:
        st.caption("No scenario tools have been configured.")


def _render_monte_carlo(outputs: FinancialOutputs) -> None:
    st.subheader("Monte Carlo Simulation")
    payload = st.session_state.get("input_payload")
    if payload is None:
        payload = {}
        st.session_state["input_payload"] = payload
    st.markdown("### Monte Carlo Simulation Configuration")
    _render_monte_carlo_inputs(payload)
    st.session_state["input_payload"] = payload

    monte_carlo_df = _ensure_dataframe(outputs.monte_carlo)
    if px is None or pd is None:
        st.warning("Plotly unavailable. Displaying Monte Carlo results in tabular form.")
        st.dataframe(monte_carlo_df, use_container_width=True)
    else:
        fig = px.histogram(monte_carlo_df, x="NPV", nbins=40, title="NPV Distribution")
        st.plotly_chart(fig, use_container_width=True)
        st.dataframe(monte_carlo_df.describe().T, use_container_width=True)


def _render_break_even(outputs: FinancialOutputs) -> None:
    st.subheader("Break-even Analysis")
    payload = st.session_state.get("input_payload")
    if payload is None:
        payload = {}
        st.session_state["input_payload"] = payload

    st.markdown("### Break-even Analysis Inputs")
    _render_break_even_inputs(payload)
    st.session_state["input_payload"] = payload

    break_even_df = _ensure_dataframe(outputs.break_even)
    if pd is not None and isinstance(break_even_df, pd.DataFrame):
        break_even_frame = break_even_df.reset_index().rename(columns={"index": "Product"})
    else:
        break_even_frame = pd.DataFrame(break_even_df)
    st.dataframe(break_even_frame, use_container_width=True)

    if px is not None and pd is not None and not break_even_frame.empty:
        y_column = (
            "Break-even Units"
            if "Break-even Units" in break_even_frame.columns
            else break_even_frame.columns[-1]
        )
        if y_column in break_even_frame.columns and "Product" in break_even_frame.columns:
            fig_break_even = px.bar(
                break_even_frame,
                x="Product",
                y=y_column,
                title="Break-even Units by Product",
            )
            st.plotly_chart(
                fig_break_even,
                use_container_width=True,
                key="break_even_units_chart",
            )

    st.markdown("### Payback Schedule")
    payback_df = _with_year(outputs.payback)
    st.dataframe(payback_df, use_container_width=True)

    if px is not None and pd is not None:
        if isinstance(payback_df, pd.DataFrame):
            payback_frame = payback_df
        else:
            payback_frame = pd.DataFrame(payback_df)
        if {"Year", "Cumulative"}.issubset(payback_frame.columns):
            fig_payback = px.line(
                payback_frame,
                x="Year",
                y="Cumulative",
                markers=True,
                title="Cumulative Payback",
            )
            st.plotly_chart(
                fig_payback,
                use_container_width=True,
                key="cumulative_payback_chart",
            )

    st.markdown("### Discounted Payback Schedule")
    discounted_df = _with_year(outputs.discounted_payback)
    st.dataframe(discounted_df, use_container_width=True)

    if px is not None and pd is not None:
        if isinstance(discounted_df, pd.DataFrame):
            discounted_frame = discounted_df
        else:
            discounted_frame = pd.DataFrame(discounted_df)
        if {"Year", "Cumulative"}.issubset(discounted_frame.columns):
            fig_discounted = px.line(
                discounted_frame,
                x="Year",
                y="Cumulative",
                markers=True,
                title="Discounted Cumulative Payback",
            )
            st.plotly_chart(
                fig_discounted,
                use_container_width=True,
                key="discounted_payback_chart",
            )



def _render_break_even_inputs(payload: dict) -> None:
    defaults = _payload_to_break_even_rows(payload)
    if not defaults:
        defaults = _default_break_even_rows(payload)

    default_map = {
        str(row.get("Product", "")): dict(row)
        for row in defaults
        if isinstance(row, Mapping) and row.get("Product")
    }

    overrides = set(st.session_state.get("break_even_overrides", []) or [])
    rows: list[dict] = list(st.session_state.get("break_even_rows", []) or [])

    if default_map:
        if not rows:
            rows = [dict(value) for value in default_map.values()]
        else:
            row_map = {
                str(row.get("Product", "")): dict(row)
                for row in rows
                if isinstance(row, Mapping) and row.get("Product")
            }
            aligned_rows: list[dict] = []
            for product, default_row in default_map.items():
                if product in overrides and product in row_map:
                    aligned_rows.append(row_map[product])
                else:
                    aligned_rows.append(dict(default_row))
                    if product in overrides and product not in row_map:
                        overrides.discard(product)
            for product, row in row_map.items():
                if product not in default_map:
                    aligned_rows.append(row)
                    overrides.add(product)
            rows = aligned_rows

        st.session_state["break_even_rows"] = rows

    if not rows:
        st.info("No break-even assumptions configured. Use the form below to add entries.")

    product_catalog = sorted(
        value
        for value in {
            *(payload.get("unit_costs", {}) or {}).keys(),
            *(row.get("Product", "") or "" for row in rows),
        }
        if value
    )

    cost_source = st.session_state.get("fixed_variable_rows", []) or _payload_to_fixed_variable_rows(
        payload
    )
    cost_lookup: dict[str, dict[str, float]] = {}
    for entry in cost_source:
        if not isinstance(entry, Mapping):
            continue
        name = str(entry.get("Product", "") or "").strip()
        if not name:
            continue
        cost_lookup[name] = {
            "fixed": float(entry.get("Fixed Cost", 0.0) or 0.0),
            "variable": float(entry.get("Variable Cost", 0.0) or 0.0),
        }

    unit_costs = payload.get("unit_costs", {}) if isinstance(payload, Mapping) else {}
    price_lookup: dict[str, float] = {}
    if isinstance(unit_costs, Mapping):
        for name, values in unit_costs.items():
            if isinstance(values, Mapping):
                price_lookup[str(name)] = float(values.get("price", 0.0) or 0.0)

    updated_rows: list[dict] = []
    overrides_updated = set(overrides)
    for index, row in enumerate(rows):
        container = st.container()
        with container:
            cols = st.columns([2.0, 1.4, 1.4, 1.4, 1.4, 1.4, 0.8])
            product_value = _select_or_create_option(
                cols[0],
                "Product",
                product_catalog,
                f"break_even_product_{index}",
                current_value=str(row.get("Product", "")),
            )
            clean_product = (product_value or "").strip()
            cost_info = cost_lookup.get(clean_product)
            fixed_default = (
                cost_info["fixed"]
                if cost_info is not None
                else float(row.get("Fixed Cost", 0.0))
            )
            variable_default = (
                cost_info["variable"]
                if cost_info is not None
                else float(row.get("Variable Cost", 0.0))
            )
            fixed_key = f"break_even_fixed_{index}"
            _set_widget_value(fixed_key, float(fixed_default))
            cols[1].number_input(
                "Fixed Cost",
                value=float(fixed_default),
                key=fixed_key,
                step=1000.0,
                format="%.2f",
                min_value=0.0,
                disabled=True,
                help="Managed via the Fixed & Variable Costs table.",
            )
            price_default = price_lookup.get(clean_product, float(row.get("Selling Price", 0.0)))
            price_key = f"break_even_price_{index}"
            _set_widget_value(price_key, float(price_default))
            selling_price = cols[2].number_input(
                "Selling Price",
                value=float(price_default),
                key=price_key,
                step=0.001,
                format="%.4f",
                min_value=0.0,
                disabled=True,
                help="Managed via the Core Assumptions table.",
            )
            variable_key = f"break_even_variable_{index}"
            _set_widget_value(variable_key, float(variable_default))
            cols[3].number_input(
                "Variable Cost",
                value=float(variable_default),
                key=variable_key,
                step=0.001,
                format="%.4f",
                min_value=0.0,
                disabled=True,
                help="Managed via the Fixed & Variable Costs table.",
            )
            target_profit = cols[4].number_input(
                "Target Profit",
                value=float(row.get("Target Profit", 0.0)),
                key=f"break_even_target_{index}",
                step=1000.0,
                format="%.2f",
                min_value=0.0,
            )
            expected_volume = cols[5].number_input(
                "Expected Volume",
                value=float(row.get("Expected Volume", 0.0)),
                key=f"break_even_volume_{index}",
                step=1.0,
                format="%.4f",
                min_value=0.0,
            )
            remove = cols[6].button("Remove", key=f"break_even_remove_{index}")

            if cost_info is None:
                cols[1].warning(
                    "Configure this product in the Fixed & Variable Costs table to set its costs."
                )
            if clean_product and clean_product not in price_lookup:
                cols[2].warning(
                    "Set this product's selling price in the Core Assumptions table."
                )

            fixed_cost = fixed_default
            variable_cost = variable_default
            metrics = _calculate_break_even_metrics(
                {
                    "Fixed Cost": fixed_cost,
                    "Selling Price": selling_price,
                    "Variable Cost": variable_cost,
                    "Target Profit": target_profit,
                    "Expected Volume": expected_volume,
                }
            )

            metric_cols = st.columns([1.6, 1.6, 1.6, 1.6, 1.6])

            metric_cols[0].metric(
                "Contribution Margin",
                _format_display(metrics["Contribution Margin"], 4),
            )
            metric_cols[1].metric(
                "Margin Ratio",
                _format_percentage(metrics["Contribution Margin Ratio"]),
            )
            metric_cols[2].metric(
                "Break-even Units",
                _format_display(metrics["Break-even Units"], 2),
            )
            metric_cols[3].metric(
                "Break-even Revenue",
                _format_display(metrics["Break-even Revenue"], 2),
            )
            metric_cols[4].metric(
                "Margin of Safety",
                _format_display(metrics["Margin of Safety (Units)"], 2),
            )

            if metrics["Contribution Margin"] <= 0:
                cols[3].error("Contribution margin non-positive")
            if (
                expected_volume > 0
                and metrics["Break-even Units"] == metrics["Break-even Units"]
                and metrics["Break-even Units"] > expected_volume
            ):
                cols[5].warning("Break-even exceeds expected volume")

            if remove:
                overrides_updated.discard(clean_product)
                continue

            current_row = {
                "Product": clean_product,
                "Fixed Cost": fixed_cost,
                "Selling Price": selling_price,
                "Variable Cost": variable_cost,
                "Target Profit": target_profit,
                "Expected Volume": expected_volume,
            }
            updated_rows.append(current_row)

            default_row = default_map.get(clean_product)
            if default_row is None:
                overrides_updated.add(clean_product)
            elif _row_matches_default(current_row, default_row):
                overrides_updated.discard(clean_product)
            else:
                overrides_updated.add(clean_product)

    st.session_state["break_even_rows"] = updated_rows
    st.session_state["break_even_overrides"] = sorted(overrides_updated)
    _break_even_rows_to_payload(updated_rows, payload)

    totals = _aggregate_break_even_metrics(updated_rows)
    if totals:
        summary_cols = st.columns(3)
        summary_cols[0].metric("Total Fixed Cost", _format_display(totals["total_fixed"], 2))
        summary_cols[1].metric(
            "Weighted Margin Ratio",
            _format_percentage(totals["weighted_margin_ratio"]),
        )
        summary_cols[2].metric(
            "Aggregate Break-even Revenue",
            _format_display(totals["aggregate_break_even_revenue"], 2),
        )

    st.markdown("#### Add Break-even Input")
    with st.form("break_even_add_form", clear_on_submit=True):
        new_product = _select_or_create_option(
            st,
            "Product",
            product_catalog,
            "break_even_new_product",
        )
        new_price_lookup = price_lookup.get(str(new_product or "").strip(), 0.0)
        st.number_input(
            "Selling Price",
            value=float(new_price_lookup),
            step=0.001,
            format="%.4f",
            key="break_even_new_price",
            min_value=0.0,
            disabled=True,
            help="Managed via the Core Assumptions table.",
        )
        new_target = st.number_input(
            "Target Profit",
            value=0.0,
            step=1000.0,
            format="%.2f",
            key="break_even_new_target",
            min_value=0.0,
        )
        new_volume = st.number_input(
            "Expected Volume",
            value=0.0,
            step=1.0,
            format="%.4f",
            key="break_even_new_volume",
            min_value=0.0,
        )
        submitted = st.form_submit_button("Add")

    if submitted:
        cleaned_product = (new_product or "").strip()
        if not cleaned_product:
            st.warning("Product name is required to add a break-even row.")
        elif cleaned_product not in cost_lookup:
            st.warning(
                "Add fixed and variable cost details for this product in the Fixed & Variable Costs table first."
            )
        elif cleaned_product not in price_lookup:
            st.warning(
                "Assign a selling price for this product in the Core Assumptions table before adding it."
            )
        else:
            additions = list(st.session_state.get("break_even_rows", []) or [])
            additions.append(
                {
                    "Product": cleaned_product,
                    "Fixed Cost": float(cost_lookup[cleaned_product]["fixed"]),
                    "Selling Price": float(price_lookup[cleaned_product]),
                    "Variable Cost": float(cost_lookup[cleaned_product]["variable"]),
                    "Target Profit": float(new_target),
                    "Expected Volume": float(new_volume),
                }
            )
            st.session_state["break_even_rows"] = additions
            overrides_updated = set(st.session_state.get("break_even_overrides", []) or [])
            overrides_updated.add(cleaned_product)
            st.session_state["break_even_overrides"] = sorted(overrides_updated)
            _break_even_rows_to_payload(additions, payload)
            for key in (
                "break_even_new_product",
                "break_even_new_product_select",
                "break_even_new_product_custom",
                "break_even_new_price",
                "break_even_new_target",
                "break_even_new_volume",
            ):
                st.session_state.pop(key, None)
            _rerun()

def _dict_to_dataframe(data: Mapping[str, float], index_label: str, value_label: str):
    if pd is None:
        return [
            {index_label: key, value_label: value}
            for key, value in sorted(data.items(), key=lambda item: item[0])
        ]
    return (
        pd.DataFrame(list(data.items()), columns=[index_label, value_label])
        .sort_values(index_label)
        .reset_index(drop=True)
    )


def _with_year(table) -> "pd.DataFrame | Table | list":
    frame = _ensure_dataframe(table)
    if pd is None:
        return frame
    result = frame.copy()
    if "Year" not in result.columns and not isinstance(frame.index, pd.RangeIndex):
        result.insert(0, "Year", list(frame.index))
    return result.reset_index(drop=True)


def _clean_streamlit_cell(value: object) -> object:
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, default=_json_default)
    return value


def _sanitize_dataframe(frame: "pd.DataFrame") -> "pd.DataFrame":
    cleaned = frame.copy()
    cleaned.columns = [_clean_streamlit_cell(column) for column in cleaned.columns]
    cleaned.columns = [str(column) for column in cleaned.columns]
    if cleaned.index.name is not None:
        cleaned.index.name = str(_clean_streamlit_cell(cleaned.index.name))
    if isinstance(cleaned.index, pd.MultiIndex):
        cleaned.index = pd.MultiIndex.from_tuples(
            [
                tuple(_clean_streamlit_cell(level) for level in levels)
                for levels in cleaned.index.to_list()
            ]
        )
    elif cleaned.index.dtype == "object":
        cleaned.index = cleaned.index.map(_clean_streamlit_cell)
    elif pd.api.types.is_string_dtype(cleaned.index):
        cleaned.index = cleaned.index.astype(object).map(_clean_streamlit_cell)
    for column in cleaned.columns:
        if pd.api.types.is_string_dtype(cleaned[column]):
            cleaned[column] = cleaned[column].astype(object).map(_clean_streamlit_cell)
        elif cleaned[column].dtype == "object":
            cleaned[column] = cleaned[column].map(_clean_streamlit_cell)
    return cleaned


def _ensure_dataframe(table) -> "pd.DataFrame | list":
    if isinstance(table, list):
        if pd is None:
            return table
        return _sanitize_dataframe(pd.DataFrame(table))
    if isinstance(table, Table):
        if pd is None:
            rows = []
            data = table.as_dict()
            for idx, label in enumerate(table.index):
                row = {table.index_name: label}
                for column, values in data.items():
                    row[column] = values[idx]
                rows.append(row)
            return rows
        return _sanitize_dataframe(table.to_frame())
    if hasattr(table, "to_frame"):
        try:
            frame = table.to_frame()
            if pd is not None and isinstance(frame, pd.DataFrame):
                return _sanitize_dataframe(frame)
            return frame
        except Exception:
            pass
    return table


def _format_number(value: float) -> str:
    if abs(value) >= 1_000_000:
        return f"{value/1_000_000:,.2f}M"
    if abs(value) >= 1_000:
        return f"{value/1_000:,.2f}K"
    return f"{value:,.2f}"


def _format_display(value: float, decimals: int = 2) -> str:
    if value is None or value != value:
        return "N/A"
    return f"{value:,.{decimals}f}"


def _format_percentage(value: float, decimals: int = 2) -> str:
    if value is None or value != value:
        return "N/A"
    return f"{value * 100:,.{decimals}f}%"


def _mapping_to_rows(mapping: Mapping[str, float], key_label: str, value_label: str) -> list[dict]:
    return [
        {key_label: str(name), value_label: float(cost)}
        for name, cost in mapping.items()
    ]


def _payload_to_sensitivity_rows(payload: Mapping) -> list[dict]:
    variables = (
        payload.get("sensitivity", {}).get("variables", {})
        if isinstance(payload.get("sensitivity"), Mapping)
        else {}
    )
    rows: list[dict] = []
    for name, values in variables.items():
        numeric = [float(value) for value in values]
        rows.append({"Variable": str(name), "Values": numeric})
    return rows


def _ensure_schedule_length(values: Iterable[float], length: int, fill: float = 0.0) -> List[float]:
    sequence = [float(value) for value in values]
    if length <= 0:
        return sequence
    if len(sequence) < length:
        sequence += [fill for _ in range(length - len(sequence))]
    return sequence[:length]


def _parse_float_list(text: str) -> List[float]:
    values: List[float] = []
    if not text:
        return values
    for token in text.replace("\n", ",").split(","):
        stripped = token.strip()
        if not stripped:
            continue
        values.append(float(stripped))
    return values


def _format_float_list(values: Iterable[float]) -> str:
    return ", ".join(f"{float(value):.4f}" for value in values)


def _render_labor_section(section: str, state_key: str, payload: dict) -> None:
    rows: list[dict] = st.session_state.get(state_key, [])
    updated: list[dict] = []
    for index, row in enumerate(rows):
        cols = st.columns([3, 2, 1])
        role = cols[0].text_input(
            "Role",
            value=row.get("Role", ""),
            key=f"{state_key}_role_{index}",
        )
        cost = cols[1].number_input(
            "Annual Cost",
            value=float(row.get("Annual Cost", 0.0)),
            key=f"{state_key}_cost_{index}",
            step=0.001,
            format="%.4f",
        )
        if cols[2].button("Remove", key=f"{state_key}_remove_{index}"):
            del rows[index]
            st.session_state[state_key] = rows
            _rerun()
        updated.append({"Role": role.strip(), "Annual Cost": cost})

    if updated != rows:
        st.session_state[state_key] = updated

    with st.form(f"add_{state_key}"):
        new_role = st.text_input("Role", key=f"{state_key}_new_role")
        new_cost = st.number_input(
            "Annual Cost",
            value=0.0,
            step=0.001,
            format="%.4f",
            key=f"{state_key}_new_cost",
        )
        submitted = st.form_submit_button("Add")

    if submitted:
        if not new_role.strip():
            st.warning("Role is required to add a labour cost entry.")
        else:
            rows.append({"Role": new_role.strip(), "Annual Cost": new_cost})
            st.session_state[state_key] = rows
            for key in (f"{state_key}_new_role", f"{state_key}_new_cost"):
                st.session_state.pop(key, None)
            _rerun()

    labor = payload.setdefault("labor", {})
    labor[section] = {
        row["Role"]: row["Annual Cost"]
        for row in st.session_state.get(state_key, [])
        if row.get("Role")
    }


def _render_fixed_variable_costs(payload: dict) -> None:
    rows: list[dict] = st.session_state.get("fixed_variable_rows", [])

    if not rows:
        defaults = _payload_to_fixed_variable_rows(payload)
        if defaults:
            rows = defaults
            st.session_state["fixed_variable_rows"] = rows

    if not rows:
        st.info(
            "No fixed or variable cost assumptions configured. Use the form below to add entries."
        )

    product_catalog = sorted(
        value
        for value in {
            *(payload.get("unit_costs", {}) or {}).keys(),
            *(row.get("Product", "") or "" for row in rows),
        }
        if value
    )

    visible_rows = rows[:MAX_VISIBLE_COST_ROWS] if rows else []
    if rows and len(rows) > MAX_VISIBLE_COST_ROWS:
        st.caption(
            "Showing the first few fixed and variable cost entries. Use the add form to access other products."
        )

    updated_rows: list[dict] = []
    for index, row in enumerate(visible_rows):
        container = st.container()
        with container:
            cols = st.columns([2.0, 1.4, 1.4, 0.8])
            product_value = _select_or_create_option(
                cols[0],
                "Product",
                product_catalog,
                f"fixed_variable_product_{index}",
                current_value=str(row.get("Product", "")),
            )
            previous_fixed = float(row.get("Fixed Cost", 0.0))
            previous_flag = bool(row.get("__has_fixed__", False))
            fixed_cost = cols[1].number_input(
                "Fixed Cost",
                value=previous_fixed,
                key=f"fixed_variable_fixed_{index}",
                step=1000.0,
                format="%.2f",
                min_value=0.0,
            )
            variable_cost = cols[2].number_input(
                "Variable Cost",
                value=float(row.get("Variable Cost", 0.0)),
                key=f"fixed_variable_variable_{index}",
                step=0.001,
                format="%.4f",
                min_value=0.0,
            )
            remove = cols[3].button("Remove", key=f"fixed_variable_remove_{index}")

            if remove:
                continue

            has_fixed = previous_flag or abs(float(fixed_cost) - previous_fixed) > 1e-9

            updated_rows.append(
                {
                    "Product": product_value.strip(),
                    "Fixed Cost": float(fixed_cost),
                    "Variable Cost": float(variable_cost),
                    "__has_fixed__": has_fixed,
                }
            )

    for hidden_row in rows[MAX_VISIBLE_COST_ROWS:]:
        updated_rows.append(dict(hidden_row))

    st.session_state["fixed_variable_rows"] = updated_rows
    _fixed_variable_rows_to_payload(updated_rows, payload)

    st.markdown("#### Add Fixed & Variable Cost")
    with st.form("fixed_variable_add_form", clear_on_submit=True):
        new_product = _select_or_create_option(
            st,
            "Product",
            product_catalog,
            "fixed_variable_new_product",
        )
        new_fixed = st.number_input(
            "Fixed Cost",
            value=0.0,
            step=1000.0,
            format="%.2f",
            key="fixed_variable_new_fixed",
            min_value=0.0,
        )
        new_variable = st.number_input(
            "Variable Cost",
            value=0.0,
            step=0.001,
            format="%.4f",
            key="fixed_variable_new_variable",
            min_value=0.0,
        )
        submitted = st.form_submit_button("Add")

    if submitted:
        cleaned_product = (new_product or "").strip()
        if not cleaned_product:
            st.warning("Product name is required to add a cost entry.")
        else:
            additions = st.session_state.get("fixed_variable_rows", [])
            additions.append(
                {
                    "Product": cleaned_product,
                    "Fixed Cost": float(new_fixed),
                    "Variable Cost": float(new_variable),
                    "__has_fixed__": True,
                }
            )
            st.session_state["fixed_variable_rows"] = additions
            for key in (
                "fixed_variable_new_product",
                "fixed_variable_new_fixed",
                "fixed_variable_new_variable",
            ):
                st.session_state.pop(key, None)
            _fixed_variable_rows_to_payload(additions, payload)
            _rerun()



def _render_utility_schedule(payload: dict) -> None:
    st.markdown("### Utility Schedule")
    st.caption(
        "Adjust electricity, water, and steam usage assumptions for each projection "
        "year. Use the plus/minus controls to tweak values or add a new year below."
    )

    rows: list[dict] = st.session_state.get("utility_entries") or []
    if not rows:
        rows = _payload_to_utility_entries(payload)
        st.session_state["utility_entries"] = rows

    if not rows:
        rows = [_default_utility_entry(0)]
        st.session_state["utility_entries"] = rows

    updated_rows: list[dict] = []
    payload_years: Sequence | None = payload.get("years")

    year_catalog = [str(year) for year in payload_years if year is not None]
    existing_labels = [
        str(row.get("label"))
        for row in rows
        if isinstance(row, Mapping) and row.get("label")
    ]
    for label in existing_labels:
        if label not in year_catalog:
            year_catalog.append(label)
    if not year_catalog:
        year_catalog = [f"Year {idx + 1}" for idx in range(len(rows) or 1)]

    st.markdown("#### Focused utility editor")
    selected_label = st.selectbox(
        "Utility year",
        options=year_catalog,
        index=0,
        key="utility_row_selector",
    )
    st.caption("Editing a single utility year entry. Change the selector to view another year.")

    filtered_rows: list[tuple[int, dict]] = []
    for idx, row in enumerate(rows):
        label = str(row.get("label") or row.get("Year") or row.get("year") or "")
        if label == selected_label:
            filtered_rows.append((idx, row))

    if not filtered_rows:
        st.caption(
            "No matching utility entry found for the selected year. "
            "Use the form below to add a new entry."
        )

    for slot in range(min(len(filtered_rows), 1)):
        row_index, row = filtered_rows[slot]
        entry = _normalise_utility_entry(row, row_index)
        container = st.container()
        with container:
            cols = st.columns([2.0, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5, 0.8])

            default_label = str(entry.get("label") or selected_label or f"Year {row_index + 1}")
            label_value = _select_or_create_option(
                cols[0],
                "Year",
                year_catalog,
                f"utility_label_{row_index}",
                current_value=default_label,
            )
            if label_value and label_value not in year_catalog:
                year_catalog.append(label_value)
            if not label_value:
                label_value = default_label

            parsed_year = _parse_year_value(
                label_value,
                entry.get("year")
                if isinstance(entry.get("year"), int)
                else row_index + 1,
            )

            electricity_per_day = cols[1].number_input(
                "Electricity per day",
                value=float(entry.get("electricity_per_day", 0.0)),
                key=f"utility_elec_per_day_{row_index}",
                min_value=0.0,
                step=0.1,
                format="%.4f",
            )
            electricity_rate = cols[2].number_input(
                "Price per kWh",
                value=float(entry.get("electricity_rate", 0.0)),
                key=f"utility_elec_rate_{row_index}",
                min_value=0.0,
                step=0.01,
                format="%.4f",
            )
            electricity_days = cols[3].number_input(
                "Electricity operating days",
                value=int(entry.get("electricity_days", 0)),
                key=f"utility_elec_days_{row_index}",
                min_value=0,
                step=1,
            )

            water_per_day = cols[4].number_input(
                "Water per day",
                value=float(entry.get("water_per_day", 0.0)),
                key=f"utility_water_per_day_{row_index}",
                min_value=0.0,
                step=0.1,
                format="%.4f",
            )
            water_rate = cols[5].number_input(
                "Price per cubic meter",
                value=float(entry.get("water_rate", 0.0)),
                key=f"utility_water_rate_{row_index}",
                min_value=0.0,
                step=0.01,
                format="%.4f",
            )
            water_days = cols[6].number_input(
                "Water operating days",
                value=int(entry.get("water_days", 0)),
                key=f"utility_water_days_{row_index}",
                min_value=0,
                step=1,
            )

            steam_per_hour = cols[7].number_input(
                "Steam per hour",
                value=float(entry.get("steam_per_hour", 0.0)),
                key=f"utility_steam_per_hour_{row_index}",
                min_value=0.0,
                step=0.1,
                format="%.4f",
            )
            steam_rate = cols[8].number_input(
                "Price per steam hour",
                value=float(entry.get("steam_rate", 0.0)),
                key=f"utility_steam_rate_{row_index}",
                min_value=0.0,
                step=0.01,
                format="%.4f",
            )
            steam_days = cols[9].number_input(
                "Steam operating days",
                value=int(entry.get("steam_days", 0)),
                key=f"utility_steam_days_{row_index}",
                min_value=0,
                step=1,
            )

            remove_clicked = cols[10].button(
                "Remove", key=f"utility_remove_{row_index}", help="Delete this utility row"
            )

            steam_hours = st.number_input(
                "Steam operating hours",
                value=int(entry.get("steam_hours", 0)),
                key=f"utility_steam_hours_{row_index}",
                min_value=0,
                step=1,
            )

        if remove_clicked and len(rows) > 1:
            del rows[row_index]
            st.session_state["utility_entries"] = rows
            _utility_entries_to_payload(rows, payload)
            _rerun()

        updated_rows.append(
            (
                row_index,
                _normalise_utility_entry(
                    {
                        "label": label_value,
                        "year": parsed_year,
                        "electricity_per_day": electricity_per_day,
                        "electricity_rate": electricity_rate,
                        "electricity_days": electricity_days,
                        "water_per_day": water_per_day,
                        "water_rate": water_rate,
                        "water_days": water_days,
                        "steam_per_hour": steam_per_hour,
                        "steam_rate": steam_rate,
                        "steam_days": steam_days,
                        "steam_hours": steam_hours,
                    },
                    row_index,
                ),
            )
        )

    updated_map = {idx: data for idx, data in updated_rows}
    merged_rows: list[dict] = []
    for idx, row in enumerate(rows):
        merged_rows.append(updated_map.get(idx, dict(row)))

    if merged_rows != rows:
        st.session_state["utility_entries"] = merged_rows
        rows = merged_rows

    _utility_entries_to_payload(rows, payload)

    st.markdown("#### Add utility assumption")
    default_entry = _next_utility_entry(rows, payload_years)
    default_label = str(default_entry.get("label", f"Year {len(rows) + 1}"))
    if default_label and default_label not in year_catalog:
        year_catalog.append(default_label)
    with st.form("utility_add_row"):
        new_label = _select_or_create_option(
            st,
            "Year",
            year_catalog,
            "utility_new_label",
            current_value=default_label,
        )
        if new_label and new_label not in year_catalog:
            year_catalog.append(new_label)
        new_electricity_per_day = st.number_input(
            "Electricity per day",
            value=float(default_entry.get("electricity_per_day", 0.0)),
            key="utility_new_elec_per_day",
            min_value=0.0,
            step=0.1,
            format="%.4f",
        )
        new_electricity_rate = st.number_input(
            "Price per kWh",
            value=float(default_entry.get("electricity_rate", 0.0)),
            key="utility_new_elec_rate",
            min_value=0.0,
            step=0.01,
            format="%.4f",
        )
        new_electricity_days = st.number_input(
            "Electricity operating days",
            value=int(default_entry.get("electricity_days", 0)),
            key="utility_new_elec_days",
            min_value=0,
            step=1,
        )
        new_water_per_day = st.number_input(
            "Water per day",
            value=float(default_entry.get("water_per_day", 0.0)),
            key="utility_new_water_per_day",
            min_value=0.0,
            step=0.1,
            format="%.4f",
        )
        new_water_rate = st.number_input(
            "Price per cubic meter",
            value=float(default_entry.get("water_rate", 0.0)),
            key="utility_new_water_rate",
            min_value=0.0,
            step=0.01,
            format="%.4f",
        )
        new_water_days = st.number_input(
            "Water operating days",
            value=int(default_entry.get("water_days", 0)),
            key="utility_new_water_days",
            min_value=0,
            step=1,
        )
        new_steam_per_hour = st.number_input(
            "Steam per hour",
            value=float(default_entry.get("steam_per_hour", 0.0)),
            key="utility_new_steam_per_hour",
            min_value=0.0,
            step=0.1,
            format="%.4f",
        )
        new_steam_rate = st.number_input(
            "Price per steam hour",
            value=float(default_entry.get("steam_rate", 0.0)),
            key="utility_new_steam_rate",
            min_value=0.0,
            step=0.01,
            format="%.4f",
        )
        new_steam_days = st.number_input(
            "Steam operating days",
            value=int(default_entry.get("steam_days", 0)),
            key="utility_new_steam_days",
            min_value=0,
            step=1,
        )
        new_steam_hours = st.number_input(
            "Steam operating hours",
            value=int(default_entry.get("steam_hours", 0)),
            key="utility_new_steam_hours",
            min_value=0,
            step=1,
        )

        submitted = st.form_submit_button("Add utility year")

    if submitted:
        new_entry = _normalise_utility_entry(
            {
                "label": new_label or f"Year {len(rows) + 1}",
                "year": _parse_year_value(new_label, len(rows) + 1),
                "electricity_per_day": new_electricity_per_day,
                "electricity_rate": new_electricity_rate,
                "electricity_days": new_electricity_days,
                "water_per_day": new_water_per_day,
                "water_rate": new_water_rate,
                "water_days": new_water_days,
                "steam_per_hour": new_steam_per_hour,
                "steam_rate": new_steam_rate,
                "steam_days": new_steam_days,
                "steam_hours": new_steam_hours,
            },
            len(rows),
        )
        updated = rows + [new_entry]
        st.session_state["utility_entries"] = updated
        _utility_entries_to_payload(updated, payload)
        for key in (
            "utility_new_label",
            "utility_new_elec_per_day",
            "utility_new_elec_rate",
            "utility_new_elec_days",
            "utility_new_water_per_day",
            "utility_new_water_rate",
            "utility_new_water_days",
            "utility_new_steam_per_hour",
            "utility_new_steam_rate",
            "utility_new_steam_days",
            "utility_new_steam_hours",
        ):
            st.session_state.pop(key, None)
        _rerun()


def _render_receivable_inputs(payload: dict) -> None:
    rows: list[dict] = st.session_state.get("receivable_rows", [])
    payload_years = payload.get("years", [])

    if not rows:
        st.info(
            "No accounts receivable assumptions configured. Use the form below to add entries."
        )

    updated_rows: list[dict] = list(rows)

    def _build_year_catalog() -> list[str]:
        base_year_labels = [str(year) for year in payload_years if year is not None]
        existing_labels = [
            str(row.get("label", "")) for row in updated_rows if row.get("label")
        ]
        catalog = list(dict.fromkeys([*base_year_labels, *existing_labels]))
        if not catalog:
            max_length = max(len(updated_rows), len(payload_years), 1)
            catalog = [f"Year {index + 1}" for index in range(max_length)]
        return catalog

    visible_count = min(len(updated_rows), MAX_VISIBLE_RECEIVABLE_ROWS)

    if visible_count and len(updated_rows) > MAX_VISIBLE_RECEIVABLE_ROWS:
        st.caption(
            "Select which receivable year to edit using the dropdowns below. Additional "
            "years remain available in the model and can be chosen from the selectors."
        )

    for slot in range(visible_count):
        year_catalog = _build_year_catalog()
        option_indices = list(range(len(updated_rows)))
        if not option_indices:
            break

        default_index = option_indices[min(slot, len(option_indices) - 1)]
        container = st.container()
        with container:
            selected_index = container.selectbox(
                "Receivable year",
                option_indices,
                index=option_indices.index(default_index),
                format_func=lambda idx: str(
                    updated_rows[idx].get("label")
                    or updated_rows[idx].get("Year")
                    or f"Year {idx + 1}"
                ),
                key=f"receivable_row_selector_{slot}",
            )

            row = updated_rows[selected_index]

            cols = st.columns([2.0, 1.2, 1.2, 1.2, 1.2, 0.7])

            current_label = str(
                row.get(
                    "label",
                    year_catalog[selected_index]
                    if selected_index < len(year_catalog)
                    else f"Year {selected_index + 1}",
                )
            )
            selected_label = _select_or_create_option(
                cols[0],
                "Year",
                year_catalog,
                f"receivable_label_{slot}_{selected_index}",
                current_value=current_label,
            )
            if selected_label and selected_label not in year_catalog:
                year_catalog.append(selected_label)
            label = selected_label or f"Year {selected_index + 1}"

            fallback_year = row.get("year")
            if not isinstance(fallback_year, (int, float)):
                if selected_index < len(payload_years):
                    fallback_year = payload_years[selected_index]
                else:
                    fallback_year = selected_index + 1
            parsed_year = _parse_year_value(
                label, int(fallback_year) if fallback_year else selected_index + 1
            )

            days_in_year = cols[1].number_input(
                "Days in Year",
                value=int(row.get("days_in_year", 365)),
                key=f"receivable_days_in_year_{slot}_{selected_index}",
                min_value=0,
                step=1,
            )

            receivable_days = cols[2].number_input(
                "Accounts Receivable Days",
                value=int(row.get("accounts_receivable_days", 0)),
                key=f"receivable_accounts_receivable_days_{slot}_{selected_index}",
                min_value=0,
                step=1,
            )

            prepaid_days = cols[3].number_input(
                "Prepaid Expense Days",
                value=int(row.get("prepaid_expense_days", 0)),
                key=f"receivable_prepaid_days_{slot}_{selected_index}",
                min_value=0,
                step=1,
            )

            other_asset_days = cols[4].number_input(
                "Other Asset Days",
                value=int(row.get("other_asset_days", 0)),
                key=f"receivable_other_asset_days_{slot}_{selected_index}",
                min_value=0,
                step=1,
            )

            if cols[5].button("Remove", key=f"receivable_remove_{slot}_{selected_index}"):
                del updated_rows[selected_index]
                st.session_state["receivable_rows"] = updated_rows
                _rerun()

            updated_rows[selected_index] = {
                "label": label,
                "year": parsed_year,
                "days_in_year": int(days_in_year),
                "accounts_receivable_days": int(receivable_days),
                "prepaid_expense_days": int(prepaid_days),
                "other_asset_days": int(other_asset_days),
            }

    if updated_rows != rows:
        st.session_state["receivable_rows"] = updated_rows
        rows = updated_rows

    year_catalog = _build_year_catalog()

    reference = rows[-1] if rows else {"label": year_catalog[0] if year_catalog else "Year 1"}

    st.markdown("#### Add accounts receivable assumption")
    with st.form("receivable_add_row"):
        default_label = str(
            reference.get(
                "label",
                year_catalog[len(rows)] if len(rows) < len(year_catalog) else f"Year {len(rows) + 1}",
            )
        )
        new_label = _select_or_create_option(
            st,
            "Year",
            year_catalog,
            "receivable_new_label",
            current_value=default_label,
        )
        if new_label and new_label not in year_catalog:
            year_catalog.append(new_label)

        new_days = st.number_input(
            "Days in Year (new)",
            value=int(reference.get("days_in_year", 365)),
            key="receivable_new_days",
            min_value=0,
            step=1,
        )
        new_receivable_days = st.number_input(
            "Accounts Receivable Days (new)",
            value=int(reference.get("accounts_receivable_days", 0)),
            key="receivable_new_receivable",
            min_value=0,
            step=1,
        )
        new_prepaid_days = st.number_input(
            "Prepaid Expense Days (new)",
            value=int(reference.get("prepaid_expense_days", 0)),
            key="receivable_new_prepaid",
            min_value=0,
            step=1,
        )
        new_other_asset_days = st.number_input(
            "Other Asset Days (new)",
            value=int(reference.get("other_asset_days", 0)),
            key="receivable_new_other",
            min_value=0,
            step=1,
        )
        submitted = st.form_submit_button("Add Year")

    if submitted:
        cleaned_label = (new_label or "").strip()
        if not cleaned_label:
            st.warning("Year label is required to add an accounts receivable assumption.")
        else:
            parsed_year = _parse_year_value(cleaned_label, len(rows) + 1)
            rows.append(
                {
                    "label": cleaned_label,
                    "year": parsed_year,
                    "days_in_year": int(new_days),
                    "accounts_receivable_days": int(new_receivable_days),
                    "prepaid_expense_days": int(new_prepaid_days),
                    "other_asset_days": int(new_other_asset_days),
                }
            )
            st.session_state["receivable_rows"] = rows
            for key in (
                "receivable_new_label",
                "receivable_new_label_select",
                "receivable_new_label_custom",
                "receivable_new_days",
                "receivable_new_receivable",
                "receivable_new_prepaid",
                "receivable_new_other",
            ):
                st.session_state.pop(key, None)
            _rerun()
def _render_inventory_inputs(payload: dict) -> None:
    rows: list[dict] = st.session_state.get("inventory_rows", [])
    payload_years = payload.get("years", [])

    updated_rows: list[dict] = list(rows)

    if not rows:
        st.info("No inventory assumptions configured. Use the form below to add entries.")

    def _build_year_catalog() -> list[str]:
        base_year_labels = [str(year) for year in payload_years if year is not None]
        existing_labels = [
            str(row.get("label", "")) for row in updated_rows if row.get("label")
        ]
        catalog = list(dict.fromkeys([*base_year_labels, *existing_labels]))
        if not catalog:
            max_length = max(len(updated_rows), len(payload_years), 1)
            catalog = [f"Year {index + 1}" for index in range(max_length)]
        return catalog

    visible_count = min(len(updated_rows), MAX_VISIBLE_INVENTORY_ROWS)

    if visible_count and len(updated_rows) > MAX_VISIBLE_INVENTORY_ROWS:
        st.caption(
            "Select which inventory year to edit using the dropdowns below. Additional "
            "years remain available in the model and can be chosen from the selectors."
        )

    for slot in range(visible_count):
        year_catalog = _build_year_catalog()
        option_indices = list(range(len(updated_rows)))
        if not option_indices:
            break

        default_index = option_indices[min(slot, len(option_indices) - 1)]
        container = st.container()
        with container:
            selected_index = container.selectbox(
                "Inventory year",
                option_indices,
                index=option_indices.index(default_index),
                format_func=lambda idx: str(
                    updated_rows[idx].get("label")
                    or updated_rows[idx].get("Year")
                    or f"Year {idx + 1}"
                ),
                key=f"inventory_row_selector_{slot}",
            )

            row = updated_rows[selected_index]

            cols = st.columns([2.0, 1.2, 1.2, 1.2, 0.7])

            current_label = str(
                row.get(
                    "label",
                    year_catalog[selected_index]
                    if selected_index < len(year_catalog)
                    else f"Year {selected_index + 1}",
                )
            )
            selected_label = _select_or_create_option(
                cols[0],
                "Year",
                year_catalog,
                f"inventory_label_{slot}_{selected_index}",
                current_value=current_label,
            )
            if selected_label and selected_label not in year_catalog:
                year_catalog.append(selected_label)
            label = selected_label or f"Year {selected_index + 1}"

            fallback_year = row.get("year")
            if not isinstance(fallback_year, (int, float)):
                if selected_index < len(payload_years):
                    fallback_year = payload_years[selected_index]
                else:
                    fallback_year = selected_index + 1
            parsed_year = _parse_year_value(
                label, int(fallback_year) if fallback_year else selected_index + 1
            )

            days_in_year = cols[1].number_input(
                "Days in Year",
                value=int(row.get("days_in_year", 365)),
                key=f"inventory_days_in_year_{slot}_{selected_index}",
                min_value=0,
                step=1,
            )

            inventory_days = cols[2].number_input(
                "Inventory Days",
                value=int(row.get("inventory_days", 0)),
                key=f"inventory_inventory_days_{slot}_{selected_index}",
                min_value=0,
                step=1,
            )

            payable_days = cols[3].number_input(
                "Accounts Payable Days",
                value=int(row.get("accounts_payable_days", 0)),
                key=f"inventory_accounts_payable_days_{slot}_{selected_index}",
                min_value=0,
                step=1,
            )

            if cols[4].button("Remove", key=f"inventory_remove_{slot}_{selected_index}"):
                del updated_rows[selected_index]
                st.session_state["inventory_rows"] = updated_rows
                _rerun()

            updated_rows[selected_index] = {
                "label": label,
                "year": parsed_year,
                "days_in_year": int(days_in_year),
                "inventory_days": int(inventory_days),
                "accounts_payable_days": int(payable_days),
            }

    if updated_rows != rows:
        st.session_state["inventory_rows"] = updated_rows
        rows = updated_rows

    year_catalog = _build_year_catalog()

    reference = rows[-1] if rows else {"label": year_catalog[0] if year_catalog else "Year 1"}

    st.markdown("#### Add inventory assumption")
    with st.form("inventory_add_row"):
        default_label = str(
            reference.get(
                "label",
                year_catalog[len(rows)] if len(rows) < len(year_catalog) else f"Year {len(rows) + 1}",
            )
        )
        new_label = _select_or_create_option(
            st,
            "Year",
            year_catalog,
            "inventory_new_label",
            current_value=default_label,
        )
        if new_label and new_label not in year_catalog:
            year_catalog.append(new_label)

        new_days = st.number_input(
            "Days in Year (new)",
            value=int(reference.get("days_in_year", 365)),
            key="inventory_new_days",
            min_value=0,
            step=1,
        )
        new_inventory_days = st.number_input(
            "Inventory Days (new)",
            value=int(reference.get("inventory_days", 0)),
            key="inventory_new_inventory",
            min_value=0,
            step=1,
        )
        new_payable_days = st.number_input(
            "Accounts Payable Days (new)",
            value=int(reference.get("accounts_payable_days", 0)),
            key="inventory_new_payable",
            min_value=0,
            step=1,
        )
        submitted = st.form_submit_button("Add Year")

    if submitted:
        cleaned_label = (new_label or "").strip()
        if not cleaned_label:
            st.warning("Year label is required to add an inventory assumption.")
        else:
            parsed_year = _parse_year_value(cleaned_label, len(rows) + 1)
            rows.append(
                {
                    "label": cleaned_label,
                    "year": parsed_year,
                    "days_in_year": int(new_days),
                    "inventory_days": int(new_inventory_days),
                    "accounts_payable_days": int(new_payable_days),
                }
            )
            st.session_state["inventory_rows"] = rows
            for key in (
                "inventory_new_label",
                "inventory_new_label_select",
                "inventory_new_label_custom",
                "inventory_new_days",
                "inventory_new_inventory",
                "inventory_new_payable",
            ):
                st.session_state.pop(key, None)
            _rerun()


def _commission_revenue_estimate(payload: Mapping, year_value: int, product: str) -> float:
    years = [int(year) for year in payload.get("years", [])] if isinstance(payload, Mapping) else []
    unit_costs = payload.get("unit_costs", {}) if isinstance(payload, Mapping) else {}
    factors = _inflation_factors_from_payload(payload)
    production = payload.get("production_estimate", {}) if isinstance(payload, Mapping) else {}

    try:
        year_position = years.index(int(year_value)) if years else 0
    except ValueError:
        year_position = 0

    units_series = production.get(product, []) if isinstance(production, Mapping) else []
    units = 0.0
    if isinstance(units_series, Sequence) and year_position < len(units_series):
        try:
            units = float(units_series[year_position])
        except (TypeError, ValueError):
            units = 0.0

    price_mapping = unit_costs.get(product, {}) if isinstance(unit_costs, Mapping) else {}
    price = float(price_mapping.get("price", 0.0) or 0.0)
    factor = float(factors[year_position]) if year_position < len(factors) and factors[year_position] else 1.0
    return units * price * factor


def _commission_base_rates(payload: Mapping) -> dict[str, float]:
    section = payload.get("distributor_commission") if isinstance(payload, Mapping) else None
    if isinstance(section, Mapping):
        raw_rows = section.get("rows", section)
    else:
        raw_rows = section

    base_rates: dict[str, tuple[int, float]] = {}
    if isinstance(raw_rows, Iterable) and not isinstance(raw_rows, (str, bytes)):
        for item in raw_rows:
            if not isinstance(item, Mapping):
                continue
            product = str(item.get("product", "")).strip()
            if not product:
                continue
            year_value = item.get("year")
            try:
                year = int(year_value)
            except (TypeError, ValueError):
                continue
            try:
                rate = float(item.get("rate", 0.0) or 0.0)
            except (TypeError, ValueError):
                rate = 0.0
            existing = base_rates.get(product)
            if existing is None or year < existing[0]:
                base_rates[product] = (year, rate)

    return {product: rate for product, (_, rate) in base_rates.items()}


def _commission_effective_rate(
    rows: Sequence[Mapping],
    payload: Mapping,
    product: str,
    year_value: int,
) -> float:
    base_rates = _commission_base_rates(payload)
    base_rate = base_rates.get(product, 0.05)
    product_rows = [
        row for row in rows if str(row.get("Product", "")).strip() == product
    ]
    if not product_rows:
        return base_rate
    try:
        target_year = int(year_value)
    except (TypeError, ValueError):
        return base_rate

    product_rows.sort(key=lambda row: int(row.get("Year", 0) or 0))
    rate = base_rate
    for idx, row in enumerate(product_rows):
        year = int(row.get("Year", 0) or 0)
        increment_value = row.get("Yearly Commission %", 0.0)
        try:
            increment = float(increment_value)
        except (TypeError, ValueError):
            increment = 0.0
        rate = rate * (1 + increment / 100.0)
        if year >= target_year:
            break
    return rate


def _render_distributor_commission(payload: Mapping) -> None:
    st.caption(
        "Configure distributor commission assumptions by year and product. "
        "Yearly Commission % is treated as an incremental change that compounds over time."
    )
    st.info(
        "Yearly Commission % (Increment) applies to the prior year's commission rate. "
        "The effective rate shown below reflects the compounded result for each year."
    )

    rows: list[dict] = st.session_state.get("commission_rows") or []
    if not rows:
        rows = _payload_to_commission_rows(payload)
        st.session_state["commission_rows"] = rows

    years = [int(year) for year in payload.get("years", [])] if isinstance(payload, Mapping) else []
    row_years = [int(row.get("Year", 0)) for row in rows if row.get("Year") is not None]
    year_catalog = sorted({*years, *row_years}) or [0]
    year_options = [str(year) for year in year_catalog]
    unit_costs = payload.get("unit_costs", {}) if isinstance(payload, Mapping) else {}
    base_products = (
        {str(name) for name in unit_costs.keys()}
        if isinstance(unit_costs, Mapping)
        else set()
    )
    product_options = sorted(
        base_products | {str(row.get("Product", "")) for row in rows if row.get("Product")}
    )

    if not rows:
        st.info(
            "No distributor commission assumptions configured. Use the form below to add entries."
        )

    st.markdown("#### Focused row editor")
    selected_product = st.selectbox(
        "Product filter",
        options=product_options if product_options else [""],
        index=0,
        key="commission_filter_product",
    )
    selected_year = st.selectbox(
        "Year filter",
        options=year_options,
        index=0,
        key="commission_filter_year",
    )
    if selected_product:
        st.caption(
            "Editing a single product/year entry. Change the filters to view another row."
        )

    filtered_rows: list[tuple[int, dict]] = []
    for idx, row in enumerate(rows):
        if selected_product and str(row.get("Product", "")).strip() != selected_product:
            continue
        if selected_year and str(row.get("Year", "")) != selected_year:
            continue
        filtered_rows.append((idx, row))

    if not filtered_rows:
        st.caption(
            "No matching commission row found for the selected filters. "
            "Use the form below to add a new entry."
        )
    visible_count = min(len(filtered_rows), 1)

    updated_rows: list[dict] = []

    base_rates = _commission_base_rates(payload)

    for view_index in range(visible_count):
        row_index, row = filtered_rows[view_index]
        container = st.container()
        with container:
            cols = st.columns([1.0, 1.8, 1.3, 1.4, 1.4, 1.6, 0.7])
            year_default = int(row.get("Year", years[row_index] if row_index < len(years) else 0))
            year_label = str(year_default) if str(year_default) in year_options else year_options[0]
            year_value = int(
                cols[0].selectbox(
                    "Year",
                    options=year_options,
                    index=year_options.index(year_label),
                    key=f"commission_year_{row_index}",
                )
            )
            product_value = _select_or_create_option(
                cols[1],
                "Product",
                product_options,
                f"commission_product_{row_index}",
                current_value=str(row.get("Product", "")),
            )
            base_rate = base_rates.get(product_value, 0.05)
            cols[2].number_input(
                "Base Rate (%)",
                value=float(base_rate * 100.0),
                step=0.1,
                min_value=0.0,
                format="%.2f",
                disabled=True,
                key=f"commission_base_{row_index}",
            )
            rate_value = cols[3].number_input(
                "Yearly Commission % (Increment)",
                value=float(row.get("Yearly Commission %", 0.0)),
                step=0.05,
                min_value=0.0,
                format="%.4f",
                key=f"commission_increment_{row_index}",
            )
            revenue_estimate = _commission_revenue_estimate(payload, year_value, product_value)
            revenue_default = float(row.get("Revenue", revenue_estimate))
            preview_row = {
                "Year": int(year_value),
                "Product": product_value.strip(),
                "Yearly Commission %": float(rate_value),
            }
            preview_rows = [dict(item) for item in rows]
            if row_index < len(preview_rows):
                preview_rows[row_index] = preview_row
            effective_rate = _commission_effective_rate(
                preview_rows, payload, product_value, year_value
            )
            cols[4].number_input(
                "Effective Rate (%)",
                value=float(effective_rate * 100.0),
                step=0.1,
                min_value=0.0,
                format="%.2f",
                disabled=True,
                key=f"commission_effective_{row_index}",
            )
            revenue_value = cols[5].number_input(
                "Revenue",
                value=float(revenue_default),
                step=1000.0,
                min_value=0.0,
                format="%.2f",
                key=f"commission_revenue_{row_index}",
            )
            cols[6].markdown("&nbsp;")
            if cols[6].button("Remove", key=f"commission_remove_{row_index}"):
                rows.pop(row_index)
                st.session_state["commission_rows"] = rows
                _commission_rows_to_payload(rows, payload)
                _rerun()

        commission_estimate = revenue_value * effective_rate
        st.caption(
            f"Estimated revenue for {product_value or 'product'} in {int(year_value)}: "
            f"{revenue_estimate:,.2f}. Effective commission rate: {effective_rate * 100:.2f}%. "
            f"Commission impact: {commission_estimate:,.2f}."
        )

        updated_rows.append(
            (
                row_index,
                {
                    "Year": int(year_value),
                    "Product": product_value.strip(),
                    "Yearly Commission %": float(rate_value),
                    "Revenue": float(revenue_value),
                    "Payment Days": int(row.get("Payment Days", 30)),
                },
            )
        )

    updated_map = {idx: data for idx, data in updated_rows}
    merged_rows: list[dict] = []
    for idx, row in enumerate(rows):
        merged_rows.append(updated_map.get(idx, dict(row)))

    if merged_rows != rows:
        st.session_state["commission_rows"] = merged_rows
        _commission_rows_to_payload(merged_rows, payload)
        rows = merged_rows
    else:
        rows = st.session_state.get("commission_rows", rows)

    st.markdown("#### Add distributor commission assumption")
    with st.form("commission_add_form", clear_on_submit=True):
        existing_years = [int(row.get("Year", 0)) for row in rows if row.get("Year") is not None]
        if years:
            remaining = [year for year in years if year not in existing_years]
            default_year = remaining[0] if remaining else years[-1]
        elif existing_years:
            default_year = existing_years[-1] + 1
        else:
            default_year = 0

        add_year = int(
            st.selectbox(
                "Year",
                options=year_options,
                index=year_options.index(str(default_year))
                if str(default_year) in year_options
                else 0,
                key="commission_new_year",
            )
        )
        add_product = _select_or_create_option(
            st,
            "Product",
            product_options,
            "commission_new_product",
        )
        add_rate = st.number_input(
            "Yearly Commission % (Increment)",
            value=0.0,
            step=0.05,
            min_value=0.0,
            format="%.4f",
            key="commission_new_increment",
        )
        add_revenue = st.number_input(
            "Revenue",
            value=float(
                _commission_revenue_estimate(payload, int(default_year), add_product or "")
            ),
            step=1000.0,
            min_value=0.0,
            format="%.2f",
            key="commission_new_revenue",
        )
        submitted = st.form_submit_button("Add distributor commission")

    if submitted:
        product_clean = (add_product or "").strip()
        if not product_clean:
            st.warning("Product is required to add a distributor commission entry.")
        else:
            new_rows = st.session_state.get("commission_rows", []) or []
            new_rows.append(
                {
                    "Year": int(add_year),
                    "Product": product_clean,
                    "Yearly Commission %": float(add_rate),
                    "Revenue": float(add_revenue),
                    "Payment Days": 30,
                }
            )
            st.session_state["commission_rows"] = new_rows
            _commission_rows_to_payload(new_rows, payload)
            for key in (
                "commission_new_year",
                "commission_new_product_select",
                "commission_new_product_custom",
                "commission_new_increment",
                "commission_new_revenue",
            ):
                st.session_state.pop(key, None)
            _rerun()

    if product_options:
        st.markdown("#### Commission horizon preview")
        preview_product = selected_product or product_options[0]
        if not selected_product:
            preview_product = st.selectbox(
                "Preview Product",
                options=product_options,
                index=0,
                key="commission_preview_product",
            )
        else:
            st.caption(f"Previewing horizon for: {preview_product}")
        increments = {
            int(row.get("Year", 0)): float(row.get("Yearly Commission %", 0.0) or 0.0)
            for row in rows
            if str(row.get("Product", "")).strip() == preview_product
        }
        preview_rows = []
        rate = base_rates.get(preview_product, 0.05)
        for year in year_catalog:
            increment = increments.get(int(year), 0.0)
            rate = rate * (1 + increment / 100.0)
            preview_rows.append(
                {
                    "Year": int(year),
                    "Yearly Commission % (Increment)": float(increment),
                    "Effective Rate (%)": float(rate * 100.0),
                }
            )
        if pd is None:
            st.table(preview_rows)
        else:
            preview_frame = pd.DataFrame(preview_rows)
            preview_frame = preview_frame.astype(str)
            st.table(preview_frame)




def _calculate_depreciation_preview(rows: Sequence[Mapping[str, object]]) -> list[dict]:
    """Return derived depreciation metrics for each configured asset row.

    The Streamlit editor stores user-provided depreciation assumptions in
    ``session_state`` as a list of dictionaries.  Rendering the schedule
    requires deriving roll-forward values—opening balances, total depreciation,
    cumulative depreciation, and net book value—for each entry so that the UI
    can surface read-only previews while keeping editable fields uncluttered.

    This helper mirrors the logic used by :meth:`FinancialModel` when it builds
    the depreciation roll-forward: rows are grouped by asset type, ordered by
    year, and then processed sequentially while honouring manual overrides for
    opening balances.  The resulting dictionaries are aligned to the original
    row order so callers can safely index into the list during widget
    rendering.
    """

    if not rows:
        return []

    grouped: dict[str, list[tuple[int, Mapping[str, object]]]] = {}
    for index, row in enumerate(rows):
        asset_key = str(row.get("asset_type", "") or "").strip().lower()
        grouped.setdefault(asset_key, []).append((index, row))

    derived_by_index: dict[int, dict] = {}

    for group in grouped.values():
        group.sort(key=lambda item: (int(item[1].get("year", 0) or 0), item[0]))

        previous_net: Optional[float] = None
        previous_cumulative: Optional[float] = None

        for life_index, (original_index, row) in enumerate(group):
            method = str(row.get("method", "straight_line") or "straight_line").strip().lower()
            if method not in DEPRECIATION_METHOD_LABELS:
                method = "straight_line"

            acquisition = float(row.get("acquisition", 0.0) or 0.0)
            depreciation_rate = float(row.get("depreciation_rate", 0.0) or 0.0)
            asset_life = int(row.get("asset_life") or 0)
            opening_net = float(row.get("opening_net_book", 0.0) or 0.0)
            opening_cumulative = float(row.get("opening_cumulative", 0.0) or 0.0)

            override_net = bool(row.get("override_net_book"))
            override_cumulative = bool(row.get("override_cumulative"))

            if previous_net is None or override_net:
                prior_net = opening_net
            else:
                prior_net = previous_net

            if previous_cumulative is None or override_cumulative:
                prior_cumulative = opening_cumulative
            else:
                prior_cumulative = previous_cumulative

            total_asset_cost = prior_net + prior_cumulative + acquisition
            allowable = max(total_asset_cost - prior_cumulative, 0.0)

            if method == "straight_line" and asset_life > 0:
                remaining_periods = max(asset_life - life_index, 1)
                total_depreciation = allowable / remaining_periods if remaining_periods else allowable
            else:
                if method == "reducing_balance":
                    depreciation_base = prior_net + (acquisition * 0.5)
                else:
                    depreciation_base = total_asset_cost

                total_depreciation = depreciation_base * depreciation_rate

                if asset_life > 0 and life_index >= asset_life - 1:
                    total_depreciation = allowable
                elif total_depreciation > allowable:
                    total_depreciation = allowable

            cumulative_depreciation = prior_cumulative + total_depreciation
            net_book_value = max(total_asset_cost - cumulative_depreciation, 0.0)

            derived_by_index[original_index] = {
                "asset_type": row.get("asset_type"),
                "year": int(row.get("year", 0) or 0),
                "prior_net_book": prior_net,
                "prior_cumulative": prior_cumulative,
                "total_asset_cost": total_asset_cost,
                "total_depreciation": total_depreciation,
                "cumulative_depreciation": cumulative_depreciation,
                "net_book_value": net_book_value,
                "method": method,
                "life_year_index": life_index,
                "asset_life": asset_life,
                "is_first": previous_net is None,
            }

            previous_net = net_book_value
            previous_cumulative = cumulative_depreciation

    return [derived_by_index.get(index, {}) for index in range(len(rows))]


def _render_depreciation_schedule(payload: dict) -> None:
    rows: list[dict] = st.session_state.get("depreciation_rows", [])
    years = payload.get("years", [])
    asset_catalog = _collect_asset_type_options(payload, rows)
    asset_options = list(asset_catalog)

    if not rows:
        st.info("No fixed asset schedule configured. Use the form below to add entries.")

    for index in range(len(rows)):
        derived = _calculate_depreciation_preview(rows)
        row = rows[index]
        data = derived[index] if index < len(derived) else {}

        cols = st.columns([1.9, 1.2, 1.0, 1.2, 1.0, 1.0, 1.2, 1.2, 1.2, 1.2, 1.2, 0.7])

        asset_input = _select_or_create_option(
            cols[0],
            "Asset Type",
            asset_options,
            f"dep_asset_{index}",
            current_value=str(row.get("asset_type", "")),
        )
        if asset_input and asset_input not in asset_options:
            asset_options.append(asset_input)

        method_value = str(row.get("method", "straight_line") or "straight_line").strip().lower()
        if method_value not in DEPRECIATION_METHOD_LABELS:
            method_value = "straight_line"
        method_label = DEPRECIATION_METHOD_LABELS[method_value]
        try:
            default_method_index = DEPRECIATION_METHOD_OPTIONS.index(method_label)
        except ValueError:
            default_method_index = 0
        method_selection_label = cols[1].selectbox(
            "Method",
            options=DEPRECIATION_METHOD_OPTIONS,
            index=default_method_index,
            key=f"dep_method_{index}",
        )
        method_selection = DEPRECIATION_LABEL_TO_VALUE.get(
            method_selection_label,
            "straight_line",
        )

        default_year = int(row.get("year", years[0] if years else 0))
        year_options: list[int] = list(dict.fromkeys([*years, default_year])) if years else [default_year]
        try:
            default_index = year_options.index(default_year)
        except ValueError:
            default_index = 0
        year_input = cols[2].selectbox(
            "Year",
            options=year_options,
            index=default_index,
            key=f"dep_year_{index}",
        )

        acquisition_input = cols[3].number_input(
            "Acquisition",
            value=float(row.get("acquisition", 0.0)),
            key=f"dep_acq_{index}",
            step=0.001,
            format="%.4f",
        )

        asset_life_value = int(row.get("asset_life", 0) or 0)
        asset_life_input = cols[4].number_input(
            "Asset Life",
            value=asset_life_value,
            key=f"dep_life_{index}",
            min_value=0,
            step=1,
        )

        prior_net_book = float(
            data.get("prior_net_book", row.get("opening_net_book", 0.0) or 0.0)
        )
        prior_cumulative = float(
            data.get("prior_cumulative", row.get("opening_cumulative", 0.0) or 0.0)
        )
        _set_widget_value(f"dep_open_nb_{index}", prior_net_book)
        cols[5].number_input(
            "Net Book Value (prev year)",
            key=f"dep_open_nb_{index}",
            step=0.001,
            format="%.4f",
            disabled=True,
        )

        depreciation_rate_input = cols[6].number_input(
            "Depreciation Rate",
            value=float(row.get("depreciation_rate", 0.0)),
            key=f"dep_rate_{index}",
            step=0.0001,
            format="%.5f",
        )

        total_asset_cost = float(
            data.get(
                "total_asset_cost",
                acquisition_input + prior_net_book + prior_cumulative,
            )
        )
        total_depreciation = float(data.get("total_depreciation", 0.0))
        if total_depreciation <= 0.0 and total_asset_cost > 0.0:
            method_used = str(data.get("method", method_selection) or method_selection)
            if method_used not in DEPRECIATION_METHOD_LABELS:
                method_used = "straight_line"
            allowable = max(total_asset_cost - prior_cumulative, 0.0)
            if method_used == "straight_line" and asset_life_input > 0:
                remaining = max(asset_life_input - int(data.get("life_year_index", 0)), 1)
                total_depreciation = allowable / remaining if remaining else allowable
            else:
                depreciation_base = (
                    prior_net_book + (acquisition_input * 0.5)
                    if method_used == "reducing_balance"
                    else total_asset_cost
                )
                total_depreciation = min(depreciation_base * depreciation_rate_input, allowable)
        cumulative_depreciation = float(
            data.get("cumulative_depreciation", prior_cumulative + total_depreciation)
        )
        net_book_value = float(data.get("net_book_value", max(total_asset_cost - cumulative_depreciation, 0.0)))

        _set_widget_value(f"dep_total_cost_{index}", total_asset_cost)
        cols[7].number_input(
            "Total Asset cost",
            value=total_asset_cost,
            key=f"dep_total_cost_{index}",
            format="%.4f",
            disabled=True,
        )

        _set_widget_value(f"dep_total_dep_{index}", total_depreciation)
        cols[8].number_input(
            "Total Depreciation",
            value=total_depreciation,
            key=f"dep_total_dep_{index}",
            format="%.4f",
            disabled=True,
        )

        _set_widget_value(f"dep_cum_dep_{index}", cumulative_depreciation)
        cols[9].number_input(
            "Cumulative Depreciation",
            value=cumulative_depreciation,
            key=f"dep_cum_dep_{index}",
            format="%.4f",
            disabled=True,
        )

        _set_widget_value(f"dep_net_book_{index}", net_book_value)
        cols[10].number_input(
            "Net Book Value",
            value=net_book_value,
            key=f"dep_net_book_{index}",
            format="%.4f",
            disabled=True,
        )

        if cols[11].button("Remove", key=f"dep_remove_{index}"):
            del rows[index]
            st.session_state["depreciation_rows"] = rows
            _rerun()

        override_net_book = bool(row.get("override_net_book", False)) or bool(
            data.get("is_first", False)
        )
        override_cumulative = bool(data.get("is_first", False))

        new_row = {
            "asset_type": asset_input.strip(),
            "year": int(year_input),
            "acquisition": float(acquisition_input),
            "depreciation_rate": float(depreciation_rate_input),
            "asset_life": int(asset_life_input),
            "method": method_selection,
            "opening_net_book": prior_net_book,
            "opening_cumulative": prior_cumulative,
            "override_net_book": override_net_book,
            "override_cumulative": override_cumulative,
        }
        rows[index] = new_row

    st.session_state["depreciation_rows"] = rows

    with st.form("add_depreciation_row"):
        new_asset = _select_or_create_option(
            st,
            "Asset Type",
            asset_options,
            "dep_new_asset",
        )
        method_default_label = DEPRECIATION_METHOD_LABELS["straight_line"]
        try:
            method_default_index = DEPRECIATION_METHOD_OPTIONS.index(method_default_label)
        except ValueError:
            method_default_index = 0
        new_method_label = st.selectbox(
            "Depreciation Method",
            options=DEPRECIATION_METHOD_OPTIONS,
            index=method_default_index,
            key="dep_new_method",
        )
        new_method = DEPRECIATION_LABEL_TO_VALUE.get(new_method_label, "straight_line")
        default_year = int(years[0]) if years else 0
        year_options: list[int] = list(dict.fromkeys([*years, default_year])) if years else [default_year]
        try:
            default_index = year_options.index(default_year)
        except ValueError:
            default_index = 0
        new_year = st.selectbox(
            "Year",
            options=year_options,
            index=default_index,
            key="dep_new_year",
        )
        new_acquisition = st.number_input(
            "Acquisition",
            value=0.0,
            step=0.001,
            format="%.4f",
            key="dep_new_acquisition",
        )
        new_opening_nb = st.number_input(
            "Opening Net Book Value",
            value=0.0,
            step=0.001,
            format="%.4f",
            key="dep_new_opening_nb",
        )
        new_asset_life = st.number_input(
            "Asset Life (years)",
            value=0,
            min_value=0,
            step=1,
            key="dep_new_life",
        )
        new_rate = st.number_input(
            "Depreciation Rate",
            value=0.0,
            step=0.0001,
            format="%.5f",
            key="dep_new_rate",
        )
        submitted = st.form_submit_button("Add")

    if submitted:
        if not new_asset.strip():
            st.warning("Asset type is required to add a depreciation entry.")
        else:
            rows.append(
                {
                    "asset_type": new_asset.strip(),
                    "year": int(new_year),
                    "acquisition": float(new_acquisition),
                    "depreciation_rate": float(new_rate),
                    "asset_life": int(new_asset_life),
                    "method": new_method,
                    "opening_net_book": float(new_opening_nb),
                    "opening_cumulative": 0.0,
                    "override_net_book": abs(new_opening_nb) > 1e-6,
                    "override_cumulative": True,
                }
            )
            st.session_state["depreciation_rows"] = rows
            for key in (
                "dep_new_asset",
                "dep_new_asset_select",
                "dep_new_asset_custom",
                "dep_new_method",
                "dep_new_year",
                "dep_new_acquisition",
                "dep_new_opening_nb",
                "dep_new_life",
                "dep_new_rate",
            ):
                st.session_state.pop(key, None)
            _rerun()


def _render_cost_and_financing(payload: dict) -> None:
    raw = payload.setdefault("raw_material_cost", {})
    raw["per_unit"] = st.number_input(
        "Raw material cost per unit",
        value=float(raw.get("per_unit", 0.0)),
        step=0.0001,
        format="%.4f",
        key="raw_material_per_unit",
    )
    annual_text = st.text_area(
        "Annual raw material spend (comma separated, optional)",
        value=_format_float_list(raw.get("annual", [])),
        key="raw_material_annual",
    )
    try:
        raw["annual"] = _parse_float_list(annual_text)
    except ValueError as exc:
        st.warning(f"Raw material schedule ignored: {exc}")

    financing = payload.setdefault("financing", {})
    finance_cols = st.columns(3)
    financing["initial_investment"] = finance_cols[0].number_input(
        "Initial investment",
        value=float(financing.get("initial_investment", 0.0)),
        step=0.1,
        format="%.4f",
        key="finance_initial",
    )
    financing["discount_rate"] = finance_cols[1].number_input(
        "Discount rate",
        value=float(financing.get("discount_rate", 0.0)),
        step=0.001,
        format="%.4f",
        key="finance_discount",
    )
    financing["share_capital"] = finance_cols[2].number_input(
        "Share capital",
        value=float(financing.get("share_capital", 0.0)),
        step=0.1,
        format="%.4f",
        key="finance_share_capital",
    )

    finance_cols = st.columns(3)
    financing["senior_debt_interest"] = finance_cols[0].number_input(
        "Senior debt interest",
        value=float(financing.get("senior_debt_interest", 0.0)),
        step=0.001,
        format="%.4f",
        key="finance_senior_interest",
    )
    financing["revolver_interest"] = finance_cols[1].number_input(
        "Revolver interest",
        value=float(financing.get("revolver_interest", 0.0)),
        step=0.001,
        format="%.4f",
        key="finance_revolver_interest",
    )
    financing["cash_interest"] = finance_cols[2].number_input(
        "Cash interest",
        value=float(financing.get("cash_interest", 0.0)),
        step=0.001,
        format="%.4f",
        key="finance_cash_interest",
    )

    years: Sequence[int] = payload.get("years", [])

    _render_debt_section(
        title="Senior Debt",
        session_key="senior_debt_rows",
        amount_label="Senior Debt Amount",
        outstanding_label="Remaining Senior Debt",
        interest_label="Interest Payment",
        principal_label="Principal Payment",
        interest_rate=float(financing.get("senior_debt_interest", 0.0)),
        years=years,
        include_duration=True,
        show_amortisation=True,
    )

    _render_debt_section(
        title="Revolver Loan",
        session_key="revolver_rows",
        amount_label="Revolver Balance",
        outstanding_label="Remaining Revolver Balance",
        interest_label="Interest Payment",
        principal_label="Principal Payment",
        interest_rate=float(financing.get("revolver_interest", 0.0)),
        years=years,
        include_duration=True,
        show_amortisation=True,
    )

    _render_debt_section(
        title="Overdraft",
        session_key="overdraft_rows",
        amount_label="Overdraft Balance",
        outstanding_label="Remaining Overdraft Balance",
        interest_label="Interest Payment",
        principal_label="Principal Payment",
        interest_rate=float(financing.get("cash_interest", 0.0)),
        years=years,
        include_duration=True,
        show_amortisation=True,
    )

    financing["dividend_payout"] = st.number_input(
        "Dividend payout ratio",
        value=float(financing.get("dividend_payout", 0.0)),
        step=0.01,
        format="%.4f",
        key="finance_dividend",
    )



def _render_debt_section(
    *,
    title: str,
    session_key: str,
    amount_label: str,
    outstanding_label: str,
    interest_label: str,
    interest_rate: float,
    years: Sequence[int],
    include_duration: bool = False,
    show_amortisation: bool = False,
    principal_label: str = "Principal Payment",
    max_schedule_rows: Optional[int] = None,
) -> None:
    st.markdown(f"#### {title}")

    rows: List[dict] = st.session_state.get(session_key, [])

    if not rows:
        st.info("No entries configured. Use the form below to add debt details.")

    if include_duration:
        if show_amortisation:
            column_widths = [1.0, 1.0, 1.4, 0.7]
            header_titles = ["Year", "Duration", amount_label, ""]
        else:
            column_widths = [1.1, 1.1, 1.4, 1.4, 1.4, 0.7]
            header_titles = [
                "Year",
                "Duration",
                amount_label,
                outstanding_label,
                interest_label,
                "",
            ]
    else:
        column_widths = [1.2, 1.6, 1.6, 1.6, 0.8]
        header_titles = ["Year", amount_label, outstanding_label, interest_label, ""]

    header = st.columns(column_widths)
    for column, title_text in zip(header, header_titles):
        column.markdown(f"**{title_text}**" if title_text else " ")

    updated_rows: List[dict] = []
    for index, row in enumerate(rows):
        cols = st.columns(column_widths)
        default_year = int(
            row.get(
                "Year",
                years[index] if index < len(years) else (years[0] if years else 0),
            )
        )
        year_value = cols[0].number_input(
            "Year",
            value=default_year,
            key=f"{session_key}_year_{index}",
            step=1,
            format="%d",
        )

        if include_duration:
            default_duration = int(row.get("Duration", 1))
            duration_value = cols[1].number_input(
                "Duration",
                value=max(1, default_duration),
                key=f"{session_key}_duration_{index}",
                step=1,
                format="%d",
                min_value=1,
            )
            amount_column_index = 2
        else:
            duration_value = int(row.get("Duration", 1)) or 1
            amount_column_index = 1

        amount_value = cols[amount_column_index].number_input(
            amount_label,
            value=float(row.get("Amount", 0.0)),
            key=f"{session_key}_amount_{index}",
            min_value=0.0,
            step=0.1,
            format="%.4f",
        )

        outstanding_value = float(row.get("Outstanding", amount_value))

        if include_duration and show_amortisation:
            outstanding_value = float(amount_value)
            action_column_index = len(column_widths) - 1
        else:
            outstanding_column_index = amount_column_index + 1
            _set_widget_value(
                f"{session_key}_outstanding_{index}", float(outstanding_value)
            )
            cols[outstanding_column_index].number_input(
                outstanding_label,
                value=float(outstanding_value),
                key=f"{session_key}_outstanding_{index}",
                min_value=0.0,
                step=0.1,
                format="%.4f",
            )

            if include_duration:
                entry = DebtEntry(
                    year=int(year_value),
                    amount=float(amount_value),
                    outstanding=float(outstanding_value),
                    duration=max(1, int(duration_value)),
                )
                interest_value = entry.first_payment(float(interest_rate))
            else:
                interest_value = float(outstanding_value) * float(interest_rate)

            interest_column_index = outstanding_column_index + 1
            interest_key = f"{session_key}_interest_{index}"
            _set_widget_value(interest_key, interest_value)
            cols[interest_column_index].number_input(
                interest_label,
                value=interest_value,
                key=interest_key,
                format="%.4f",
                disabled=True,
            )
            action_column_index = interest_column_index + 1

        if cols[action_column_index].button("Remove", key=f"{session_key}_remove_{index}"):
            del rows[index]
            st.session_state[session_key] = rows
            _rerun()
            return

        updated_rows.append(
            {
                "Year": int(year_value),
                "Amount": float(amount_value),
                "Outstanding": float(outstanding_value),
                "Duration": int(max(1, int(duration_value))),
            }
        )

    if updated_rows != rows:
        st.session_state[session_key] = updated_rows

    if show_amortisation and include_duration and updated_rows:
        entries: List[DebtEntry] = []
        for index, row in enumerate(updated_rows):
            if years:
                fallback_year = years[index] if index < len(years) else years[-1]
            else:
                fallback_year = int(row.get("Year", 0))

            entries.append(
                DebtEntry(
                    year=int(row.get("Year", fallback_year)),
                    amount=float(row.get("Amount", 0.0)),
                    outstanding=float(row.get("Outstanding", row.get("Amount", 0.0))),
                    duration=int(max(1, int(row.get("Duration", 1)))),
                )
            )

        if years:
            schedule_years = list(years)
        else:
            derived_years = {
                entry.year + offset
                for entry in entries
                for offset in range(max(int(entry.duration or 0), 1))
            }
            schedule_years = sorted(derived_years)

        (
            interest_series,
            principal_series,
            outstanding_series,
            _,
        ) = amortise_entries(entries, float(interest_rate), schedule_years)

        schedule_rows: List[dict] = []
        for idx, year in enumerate(schedule_years):
            interest_value = interest_series[idx] if idx < len(interest_series) else 0.0
            principal_value = (
                principal_series[idx] if idx < len(principal_series) else 0.0
            )
            outstanding_value = (
                outstanding_series[idx] if idx < len(outstanding_series) else 0.0
            )

            if not (interest_value or principal_value or outstanding_value):
                continue

            schedule_rows.append(
                {
                    "Year": year,
                    interest_label: interest_value,
                    principal_label: principal_value,
                    outstanding_label: outstanding_value,
                }
            )

        if schedule_rows:
            if max_schedule_rows is not None:
                schedule_rows = schedule_rows[:max_schedule_rows]
            st.markdown(f"**{title} Amortisation Schedule**")
            st.dataframe(
                _ensure_dataframe(schedule_rows),
                use_container_width=True,
            )

    next_year = (
        years[len(rows)]
        if years and len(rows) < len(years)
        else (years[-1] + 1 if years else 0)
    )

    with st.form(f"add_{session_key}"):
        new_year = st.number_input(
            "Year",
            value=int(next_year),
            step=1,
            format="%d",
            key=f"{session_key}_new_year",
        )
        if include_duration:
            new_duration = st.number_input(
                "Duration",
                value=1,
                min_value=1,
                step=1,
                format="%d",
                key=f"{session_key}_new_duration",
            )
        else:
            new_duration = 1
        new_amount = st.number_input(
            amount_label,
            value=0.0,
            min_value=0.0,
            step=0.1,
            format="%.4f",
            key=f"{session_key}_new_amount",
        )
        if st.form_submit_button("Add"):
            rows.append(
                {
                    "Year": int(new_year),
                    "Amount": float(new_amount),
                    "Outstanding": float(new_amount),
                    "Duration": int(max(1, int(new_duration))),
                }
            )
            st.session_state[session_key] = rows
            for key in (
                f"{session_key}_new_year",
                f"{session_key}_new_duration",
                f"{session_key}_new_amount",
            ):
                st.session_state.pop(key, None)
            _rerun()


def _render_tax_schedule(payload: dict) -> None:
    tax = payload.setdefault("tax", {})
    years = payload.get("years", [])

    if not years:
        st.info(
            "No projection years available. Configure projection years to edit tax "
            "assumptions."
        )
        return

    base_rate = float(tax.get("rate", 0.0))
    timing_adjustment = float(tax.get("timing_adjustment", 0.0))

    col_base, col_timing = st.columns(2)
    with col_base:
        base_rate = col_base.number_input(
            "Base tax rate",
            value=base_rate,
            step=0.01,
            format="%.4f",
            key="tax_base_rate",
        )
    with col_timing:
        timing_adjustment = col_timing.number_input(
            "Timing adjustment",
            value=timing_adjustment,
            step=0.01,
            format="%.4f",
            key="tax_timing",
        )

    tax["rate"] = float(base_rate)
    tax["timing_adjustment"] = float(timing_adjustment)

    schedule = _ensure_schedule_length(tax.get("schedule", []), len(years), fill=base_rate)

    session_key = "tax_entries"
    default_entries = [
        _normalise_tax_entry(
            {
                "label": str(year) if year is not None else f"Year {index + 1}",
                "rate": float(schedule[index]),
            },
            index,
            base_rate,
        )
        for index, year in enumerate(years)
    ]
    if not default_entries:
        default_entries = [_normalise_tax_entry({}, 0, base_rate)]

    if session_key not in st.session_state or not st.session_state.get(session_key):
        st.session_state[session_key] = default_entries

    rows: list[dict] = list(st.session_state.get(session_key, default_entries))
    updated_rows: list[dict] = []

    for index, row in enumerate(rows):
        entry = _normalise_tax_entry(row, index, base_rate)
        container = st.container()
        with container:
            cols = st.columns([2.0, 1.2, 0.8])
            label_value = cols[0].text_input(
                "Year",
                value=entry["label"],
                key=f"tax_year_label_{index}",
            ).strip()
            if not label_value:
                label_value = f"Year {index + 1}"

            rate_value = cols[1].number_input(
                "Tax rate",
                value=float(entry["rate"]),
                min_value=0.0,
                step=0.01,
                format="%.4f",
                key=f"tax_rate_value_{index}",
            )

            remove_clicked = cols[2].button("Remove", key=f"tax_remove_{index}")

        if remove_clicked and len(rows) > 1:
            rows.pop(index)
            st.session_state[session_key] = rows
            _tax_entries_to_payload(rows, tax, years, base_rate)
            _rerun()
        else:
            updated_rows.append({"label": label_value, "rate": float(rate_value)})

    if updated_rows != rows and updated_rows:
        rows = updated_rows
        st.session_state[session_key] = rows

    _tax_entries_to_payload(rows, tax, years, base_rate)

    st.markdown("#### Add tax assumption")
    default_entry = _next_tax_entry(rows, years, base_rate)
    with st.form("tax_add_form", clear_on_submit=True):
        new_label = st.text_input(
            "Year",
            value=default_entry.get("label", f"Year {len(rows) + 1}"),
            key="tax_new_label",
        ).strip()
        new_rate = st.number_input(
            "Tax rate",
            value=float(default_entry.get("rate", base_rate)),
            min_value=0.0,
            step=0.01,
            format="%.4f",
            key="tax_new_rate",
        )
        submitted = st.form_submit_button("Add tax year")

    if submitted:
        label = new_label or default_entry.get("label", f"Year {len(rows) + 1}")
        new_entry = _normalise_tax_entry({"label": label, "rate": new_rate}, len(rows), base_rate)
        updated = rows + [new_entry]
        st.session_state[session_key] = updated
        _tax_entries_to_payload(updated, tax, years, base_rate)
        for key in ("tax_new_label", "tax_new_rate"):
            st.session_state.pop(key, None)
        _rerun()


def _render_inflation_schedule(payload: dict) -> None:
    rows: list[dict] = st.session_state.get("inflation_rows", [])
    payload_years = payload.get("years", [])

    if not rows:
        st.info("No inflation assumptions configured. Use the form below to add entries.")

    def _build_year_catalog(current_rows: Sequence[Mapping]) -> list[str]:
        base_years = [str(year) for year in payload_years if year is not None]
        existing = [
            str(row.get("Year", "")).strip() for row in current_rows if row.get("Year")
        ]
        catalog = list(dict.fromkeys([*base_years, *existing]))
        if not catalog:
            max_length = max(len(current_rows), len(payload_years), 1)
            catalog = [f"Year {index + 1}" for index in range(max_length)]
        return catalog

    updated_rows: list[dict] = []
    removal_index: int | None = None
    year_catalog = _build_year_catalog(rows)

    for index, row in enumerate(rows):
        cols = st.columns([2.0, 1.2, 0.7])
        current_label = str(
            row.get("Year")
            or (year_catalog[index] if index < len(year_catalog) else f"Year {index + 1}")
        )
        selected_label = _select_or_create_option(
            cols[0],
            "Year",
            year_catalog,
            f"inflation_label_{index}",
            current_value=current_label,
        )
        if selected_label and selected_label not in year_catalog:
            year_catalog.append(selected_label)

        label = (selected_label or current_label).strip() or f"Year {index + 1}"
        rate_value = cols[1].number_input(
            "Rate",
            value=float(row.get("Rate", 0.0)),
            min_value=0.0,
            step=0.001,
            format="%.4f",
            key=f"inflation_rate_{index}",
        )

        remove_clicked = cols[2].button(
            "Remove", key=f"inflation_remove_{index}", help="Delete this inflation row"
        )

        if remove_clicked and len(rows) > 0:
            removal_index = index
            continue

        updated_rows.append({"Year": label, "Rate": float(rate_value)})

    if removal_index is not None:
        del rows[removal_index]
        st.session_state["inflation_rows"] = rows
        _rerun()
        return

    if updated_rows != rows:
        st.session_state["inflation_rows"] = updated_rows
        rows = updated_rows

    year_catalog = _build_year_catalog(rows)
    reference = rows[-1] if rows else {"Year": year_catalog[0] if year_catalog else "Year 1"}

    st.markdown("#### Add inflation entry")
    with st.form("add_inflation_row"):
        if len(year_catalog) > len(rows):
            fallback_label = year_catalog[len(rows)]
        elif year_catalog:
            fallback_label = year_catalog[-1]
        else:
            fallback_label = f"Year {len(rows) + 1}"

        new_label = _select_or_create_option(
            st,
            "Year",
            year_catalog,
            "inflation_new_label",
            current_value=str(reference.get("Year", fallback_label)),
        )
        if new_label and new_label not in year_catalog:
            year_catalog.append(new_label)

        new_rate = st.number_input(
            "Rate",
            value=float(reference.get("Rate", 0.0)),
            min_value=0.0,
            step=0.001,
            format="%.4f",
            key="inflation_new_rate",
        )
        submitted = st.form_submit_button("Add")

    if submitted:
        clean_label = (new_label or "").strip() or f"Year {len(rows) + 1}"
        updated_rows = list(rows)
        updated_rows.append({"Year": clean_label, "Rate": float(new_rate)})
        st.session_state["inflation_rows"] = updated_rows
        for key in (
            "inflation_new_label",
            "inflation_new_label_select",
            "inflation_new_label_custom",
            "inflation_new_rate",
        ):
            st.session_state.pop(key, None)
        _rerun()


def _render_risk_schedule(payload: dict) -> None:
    rows: list[dict] = st.session_state.get("risk_rows", [])
    categories = _risk_categories(payload, rows)

    if not rows:
        st.info("No risk assumptions configured. Use the form below to add entries.")

    def _build_year_catalog(current_rows: Sequence[Mapping]) -> list[str]:
        payload_years = payload.get("years") or []
        base_years = [str(year) for year in payload_years if year is not None]
        existing = [
            str(row.get("Year", "")).strip() for row in current_rows if row.get("Year")
        ]
        catalog = list(dict.fromkeys([*base_years, *existing]))
        if not catalog:
            max_length = max(len(current_rows), len(payload_years), 1)
            catalog = [f"Year {index + 1}" for index in range(max_length)]
        return catalog

    updated_rows: list[dict] = []
    removal_index: int | None = None
    year_catalog = _build_year_catalog(rows)

    for index, row in enumerate(rows):
        column_widths = [2.0] + [1.0 for _ in categories] + [0.7]
        cols = st.columns(column_widths)

        current_label = str(
            row.get("Year")
            or (year_catalog[index] if index < len(year_catalog) else f"Year {index + 1}")
        )
        selected_label = _select_or_create_option(
            cols[0],
            "Year",
            year_catalog,
            f"risk_label_{index}",
            current_value=current_label,
        )
        if selected_label and selected_label not in year_catalog:
            year_catalog.append(selected_label)

        label = (selected_label or current_label).strip() or f"Year {index + 1}"
        cleaned_row = {"Year": label}
        for position, category in enumerate(categories, start=1):
            cleaned_row[category] = cols[position].number_input(
                f"{category.title()} Risk",
                value=float(row.get(category, 0.0)),
                min_value=0.0,
                max_value=1.0,
                step=0.01,
                format="%.4f",
                key=f"risk_{category}_{index}",
            )

        remove_clicked = cols[-1].button(
            "Remove", key=f"risk_remove_{index}", help="Delete this risk row"
        )

        if remove_clicked and len(rows) > 0:
            removal_index = index
            continue

        updated_rows.append(cleaned_row)

    if removal_index is not None:
        del rows[removal_index]
        st.session_state["risk_rows"] = rows
        _rerun()
        return

    if updated_rows != rows:
        st.session_state["risk_rows"] = updated_rows
        rows = updated_rows

    year_catalog = _build_year_catalog(rows)
    reference = rows[-1] if rows else {"Year": year_catalog[0] if year_catalog else "Year 1"}

    with st.form("add_risk_row"):
        st.markdown("#### Add risk entry")
        if len(year_catalog) > len(rows):
            fallback_label = year_catalog[len(rows)]
        elif year_catalog:
            fallback_label = year_catalog[-1]
        else:
            fallback_label = f"Year {len(rows) + 1}"

        new_label = _select_or_create_option(
            st,
            "Year",
            year_catalog,
            "risk_new_label",
            current_value=str(reference.get("Year", fallback_label)),
        )
        if new_label and new_label not in year_catalog:
            year_catalog.append(new_label)

        new_values: dict[str, float] = {}
        for category in categories:
            new_values[category] = st.number_input(
                f"{category.title()} Risk",
                value=float(reference.get(category, 0.0)),
                min_value=0.0,
                max_value=1.0,
                step=0.01,
                format="%.4f",
                key=f"risk_new_{category}",
            )
        submitted = st.form_submit_button("Add")

    if submitted:
        clean_label = (new_label or "").strip() or f"Year {len(rows) + 1}"
        updated_rows = list(rows)
        updated_rows.append({"Year": clean_label, **new_values})
        st.session_state["risk_rows"] = updated_rows
        for key in (
            "risk_new_label",
            "risk_new_label_select",
            "risk_new_label_custom",
        ):
            st.session_state.pop(key, None)
        for category in categories:
            st.session_state.pop(f"risk_new_{category}", None)
        _rerun()


def _render_goal_seek(payload: dict) -> None:
    goal = payload.get("goal_seek", {}) if isinstance(payload, Mapping) else {}
    source_options = ["income_statement", "cash_flow", "summary"]
    default_source = str(goal.get("source", "income_statement"))
    if default_source not in source_options:
        default_source = "income_statement"
    source = st.selectbox(
        "Metric Source",
        source_options,
        index=source_options.index(default_source),
        key="goal_source",
        help="Choose whether to evaluate income statement, cash flow, or summary metrics.",
    )

    metric_options = {
        "income_statement": [
            "Gross Revenue",
            "Distributors Commission",
            "Net Revenue",
            "Cost of Sales",
            "Gross Profit",
            "General & Admin",
            "EBITDA",
            "Total Depreciation Expense",
            "EBIT",
            "Interest",
            "EBT",
            "Taxes",
            "Net Income",
            "Gross Profit Margin",
            "EBITDA Margin",
            "EBIT Margin",
            "Return on Equity",
        ],
        "cash_flow": [
            "Cash Flow from Operations",
            "Net Cash Generated from Operating Activities",
            "Net Cash Used in Investing Activities",
            "Net Cash Used in Financing Activities",
            CASH_FLOW_NET_COLUMN,
            CASH_FLOW_BEGIN_COLUMN,
            CASH_FLOW_END_COLUMN,
            "Net Increase/Decrease in Cash",
        ],
        "summary": ["NPV", "IRR", "Payback Period", "Discounted Payback"],
    }

    available_metrics = metric_options.get(source, [])
    if not available_metrics:
        available_metrics = [str(goal.get("metric", "Net Income"))]

    default_metric = str(goal.get("metric", available_metrics[0]))
    if default_metric not in available_metrics:
        default_metric = available_metrics[0]

    metric = st.selectbox(
        "Metric",
        available_metrics,
        index=available_metrics.index(default_metric),
        key="goal_metric",
    )

    target_value = float(goal.get("target", 0.0))
    target = st.number_input(
        "Target Value",
        value=target_value,
        step=0.01,
        format="%.4f",
        key="goal_target",
    )

    selected_year: Optional[int] = None
    if source != "summary":
        years = payload.get("years", []) if isinstance(payload, Mapping) else []
        if years:
            try:
                default_year = int(goal.get("year", years[-1]))
            except (TypeError, ValueError):
                default_year = years[-1]
            if default_year not in years:
                default_year = years[-1]
            selected_year = st.selectbox(
                "Year",
                years,
                index=years.index(default_year),
                key="goal_year",
            )
    else:
        st.caption("Summary metrics apply across the full projection horizon.")

    goal_payload: dict[str, object] = {
        "metric": metric,
        "target": float(target),
        "source": source,
    }
    if selected_year is not None:
        goal_payload["year"] = int(selected_year)
    payload["goal_seek"] = goal_payload


def _render_sensitivity_inputs(payload: dict) -> None:
    rows = st.session_state.get("sensitivity_rows", [])
    updated: list[dict] = []
    for index, row in enumerate(rows):
        cols = st.columns([3, 5, 1])
        variable = cols[0].text_input(
            "Variable",
            value=row.get("Variable", ""),
            key=f"sensitivity_var_{index}",
        )
        values_text = cols[1].text_input(
            "Multipliers",
            value=_format_float_list(row.get("Values", [])),
            help="Comma-separated multipliers applied during sensitivity analysis.",
            key=f"sensitivity_vals_{index}",
        )
        if cols[2].button("Remove", key=f"sensitivity_remove_{index}"):
            del rows[index]
            st.session_state["sensitivity_rows"] = rows
            _rerun()
        try:
            values = _parse_float_list(values_text)
        except ValueError as exc:
            st.warning(f"Sensitivity entry ignored due to invalid number: {exc}")
            values = row.get("Values", [])
        updated.append({"Variable": variable.strip(), "Values": values})

    if updated != rows:
        st.session_state["sensitivity_rows"] = updated

    with st.form("add_sensitivity"):
        new_variable = st.text_input("Variable Name", key="sensitivity_new_variable")
        new_values_text = st.text_input(
            "Multipliers",
            key="sensitivity_new_values",
            help="Comma-separated list such as 0.9, 1.0, 1.1",
        )
        submitted = st.form_submit_button("Add Variable")

    if submitted:
        if not new_variable.strip():
            st.warning("Variable name is required for sensitivity analysis.")
        else:
            try:
                new_values = _parse_float_list(new_values_text)
            except ValueError as exc:
                st.warning(f"Unable to add sensitivity variable: {exc}")
                new_values = []
            if new_values:
                rows.append({"Variable": new_variable.strip(), "Values": new_values})
                st.session_state["sensitivity_rows"] = rows
                for key in ("sensitivity_new_variable", "sensitivity_new_values"):
                    st.session_state.pop(key, None)
                _rerun()
            else:
                st.warning("At least one multiplier is required.")

    variables = {
        row["Variable"]: row["Values"]
        for row in st.session_state.get("sensitivity_rows", [])
        if row.get("Variable") and row.get("Values")
    }
    payload.setdefault("sensitivity", {})["variables"] = variables


def _render_scenario_inputs(payload: dict) -> None:
    scenarios = payload.setdefault("scenarios", {})
    updated: dict[str, dict[str, List[float]]] = {}
    scenario_items = list(scenarios.items())
    for index, (name, values) in enumerate(scenario_items):
        with st.expander(f"Scenario: {name}", expanded=False):
            new_name = st.text_input(
                "Scenario Name",
                value=name,
                key=f"scenario_name_{index}",
            )
            inflation_text = st.text_area(
                "Inflation Series",
                value=_format_float_list(values.get("inflation", [])),
                key=f"scenario_inflation_{index}",
            )
            interest_text = st.text_area(
                "Interest Series",
                value=_format_float_list(values.get("interest", [])),
                key=f"scenario_interest_{index}",
            )
            remove = st.checkbox(
                "Remove scenario",
                key=f"scenario_remove_{index}",
                value=False,
            )

        if remove:
            continue

        try:
            inflation_values = _parse_float_list(inflation_text)
            interest_values = _parse_float_list(interest_text)
        except ValueError as exc:
            st.warning(f"Scenario '{name}' ignored due to invalid number: {exc}")
            inflation_values = values.get("inflation", [])
            interest_values = values.get("interest", [])

        key_name = new_name.strip() or name
        updated[key_name] = {
            "inflation": inflation_values,
            "interest": interest_values,
        }

    with st.form("add_scenario"):
        new_name = st.text_input("Scenario Name", key="scenario_new_name")
        new_inflation = st.text_area(
            "Inflation Series",
            key="scenario_new_inflation",
        )
        new_interest = st.text_area(
            "Interest Series",
            key="scenario_new_interest",
        )
        submitted = st.form_submit_button("Add Scenario")

    if submitted:
        if not new_name.strip():
            st.warning("Scenario name is required.")
        else:
            try:
                inflation_values = _parse_float_list(new_inflation)
                interest_values = _parse_float_list(new_interest)
            except ValueError as exc:
                st.warning(f"Unable to add scenario: {exc}")
            else:
                updated[new_name.strip()] = {
                    "inflation": inflation_values,
                    "interest": interest_values,
                }
                for key in (
                    "scenario_new_name",
                    "scenario_new_inflation",
                    "scenario_new_interest",
                ):
                    st.session_state.pop(key, None)
                _rerun()

    payload["scenarios"] = updated


def _render_scenario_tool_inputs(payload: dict) -> None:
    tools = payload.setdefault("scenario_tools", {})

    # normalise alias keys so uploaded payloads remain compatible
    for alias, canonical in SCENARIO_TOOL_ALIASES.items():
        if alias in tools:
            values = tools.pop(alias)
            if canonical in tools:
                existing = tools[canonical]
                merged = list(dict.fromkeys([*(existing or []), *(values or [])]))
                tools[canonical] = merged
            else:
                tools[canonical] = list(values or [])

    for key in SCENARIO_TOOL_LABELS:
        tools.setdefault(key, [])

    for key, label in SCENARIO_TOOL_LABELS.items():
        st.markdown(f"#### {label}")
        entries = list(tools.get(key) or [])
        updated_entries: List[str] = []
        if not entries:
            st.caption("No variables configured yet. Use the form below to add one.")

        for index, value in enumerate(entries):
            cols = st.columns([0.8, 0.2])
            cleaned = cols[0].text_input(
                "Variable",
                value=value,
                key=f"scenario_tool_{key}_{index}",
            )
            remove = cols[1].button("Remove", key=f"scenario_tool_remove_{key}_{index}")
            if remove:
                continue
            if cleaned.strip():
                updated_entries.append(cleaned.strip())
        tools[key] = updated_entries

        with st.form(f"scenario_tool_add_{key}"):
            new_value = st.text_input("Variable", key=f"scenario_tool_new_{key}")
            submitted = st.form_submit_button("Add Variable")

        if submitted:
            if not new_value.strip():
                st.warning(f"Provide a variable name before adding to {label}.")
            else:
                updated_entries.append(new_value.strip())
                tools[key] = updated_entries
                for suffix in ("", "_select", "_custom"):
                    st.session_state.pop(f"scenario_tool_new_{key}{suffix}", None)
                _rerun()

    payload["scenario_tools"] = tools


def _render_monte_carlo_inputs(payload: dict) -> None:
    monte = payload.setdefault("monte_carlo", {})
    iterations = st.number_input(
        "Iterations",
        min_value=1,
        value=int(monte.get("iterations", 1000)),
        step=10,
        key="monte_iterations",
    )
    growth_range = list(monte.get("revenue_growth_range", [0.05, 0.15]))
    if len(growth_range) < 2:
        growth_range = [0.0, 0.0]
    min_growth = st.number_input(
        "Minimum revenue growth",
        value=float(growth_range[0]),
        format="%.4f",
        key="monte_growth_min",
    )
    max_growth = st.number_input(
        "Maximum revenue growth",
        value=float(growth_range[1]),
        format="%.4f",
        key="monte_growth_max",
    )
    if max_growth < min_growth:
        st.warning("Maximum growth cannot be less than minimum growth. Adjusted automatically.")
        max_growth = min_growth

    metric_options = [
        "NPV",
        "Average Net Income",
        "Average EBITDA",
        "Average Cash Flow",
    ]
    metrics = st.multiselect(
        "Metrics to capture",
        options=metric_options,
        default=[m for m in monte.get("metrics", ["NPV"]) if m in metric_options],
        key="monte_metrics",
    )
    if not metrics:
        metrics = ["NPV"]

    monte["iterations"] = int(iterations)
    monte["revenue_growth_range"] = [float(min_growth), float(max_growth)]
    monte["metrics"] = metrics

    variable_options = [
        ("revenue_growth", "Revenue Growth"),
        ("raw_material_cost", "Cost of Materials"),
        ("labor_cost", "Labour"),
        ("tax_rate", "Tax Rate"),
        ("utility_cost", "Utility"),
        ("senior_debt", "Senior Debt"),
        ("other", "Other"),
    ]
    variable_map = {code: label for code, label in variable_options}
    stored_variables = [
        code for code in monte.get("variables", ["revenue_growth"]) if code in variable_map
    ]
    default_labels = [variable_map[code] for code in stored_variables]
    selected_labels = st.multiselect(
        "Variables to randomise",
        options=[label for _, label in variable_options],
        default=default_labels if default_labels else [variable_map["revenue_growth"]],
        key="monte_variables",
    )
    selected_codes = [
        code for code, label in variable_options if label in selected_labels
    ]
    if "revenue_growth" not in selected_codes:
        selected_codes.insert(0, "revenue_growth")
    monte["variables"] = selected_codes


def _extract_metric_pairs(summary) -> Sequence[Tuple[str, float]]:
    if isinstance(summary, Table):
        return list(zip([str(label) for label in summary.index], summary.column("Value")))

    if pd is not None and hasattr(summary, "reset_index"):
        try:
            frame = summary.reset_index()
        except Exception:
            frame = pd.DataFrame(summary)
        label_column = summary.index.name if getattr(summary, "index", None) is not None else None
        if not label_column or label_column not in frame.columns:
            label_column = frame.columns[0]
        value_column = "Value" if "Value" in frame.columns else frame.columns[-1]
        return list(zip(frame[label_column].astype(str), frame[value_column].astype(float)))

    if isinstance(summary, list):
        pairs: list[Tuple[str, float]] = []
        for position, row in enumerate(summary, start=1):
            if isinstance(row, Mapping):
                label = row.get("Metric") or row.get("Year") or f"Metric {position}"
                value = float(row.get("Value", float("nan")))
                pairs.append((str(label), value))
        return pairs

    if isinstance(summary, Mapping):
        value = summary.get("Value")
        if isinstance(value, Mapping):
            return [(str(name), float(val)) for name, val in value.items()]

    return []



UTILITY_EDITOR_COLUMNS = [
    ("label", "Year", "text"),
    ("electricity_per_day", "Electricity per day", "float"),
    ("electricity_rate", "Price per kWh", "float"),
    ("electricity_days", "Electricity operating days", "int"),
    ("water_per_day", "Water per day", "float"),
    ("water_rate", "Price per cubic meter", "float"),
    ("water_days", "Water operating days", "int"),
    ("steam_per_hour", "Steam per hour", "float"),
    ("steam_rate", "Price per steam hour", "float"),
    ("steam_days", "Steam operating days", "int"),
    ("steam_hours", "Steam operating hours", "int"),
]


def _default_utility_entry(index: int, label: str | None = None) -> dict:
    label_value = str(label) if label else f"Year {index + 1}"
    year_value = _parse_year_value(label_value, index + 1)
    entry = {
        "label": label_value,
        "year": year_value,
        "electricity_per_day": 1.0,
        "electricity_rate": 1.0,
        "electricity_days": 1,
        "water_per_day": 1.0,
        "water_rate": 1.0,
        "water_days": 1,
        "steam_per_hour": 1.0,
        "steam_rate": 1.0,
        "steam_days": 1,
        "steam_hours": 1,
    }
    return entry


def _normalise_utility_entry(
    row: Mapping | None, index: int, label: str | None = None
) -> dict:
    data = dict(_default_utility_entry(index, label))
    if not isinstance(row, Mapping):
        return data

    if not label:
        raw_label = row.get("label") or row.get("Year") or row.get("Years")
        if raw_label:
            data["label"] = str(raw_label)

    raw_year = row.get("year") if isinstance(row, Mapping) else None
    if raw_year is not None:
        try:
            data["year"] = int(float(raw_year))
        except Exception:  # pragma: no cover - defensive parsing
            data["year"] = _parse_year_value(data.get("label"), index + 1)
    else:
        data["year"] = _parse_year_value(data.get("label"), index + 1)

    for field in UTILITY_FLOAT_FIELDS:
        value = row.get(field)
        if value is None or value == "":
            continue
        try:
            data[field] = float(value)
        except (TypeError, ValueError):  # pragma: no cover - defensive parsing
            continue

    for field in UTILITY_INT_FIELDS:
        value = row.get(field)
        if value is None or value == "":
            continue
        try:
            data[field] = int(float(value))
        except (TypeError, ValueError):  # pragma: no cover - defensive parsing
            continue

    if not data["label"]:
        data["label"] = f"Year {index + 1}"
    return data


def _payload_to_utility_entries(payload: Mapping) -> list[dict]:
    utility = payload.get("utility_costs", {})
    rows: list[dict] = []
    if isinstance(utility, Mapping):
        stored_rows = utility.get("years")
        if isinstance(stored_rows, Sequence):
            for index, item in enumerate(stored_rows):
                rows.append(_normalise_utility_entry(item, index))
    return rows


def _utility_entries_to_payload(rows: Sequence[Mapping], payload: dict) -> None:
    utility = payload.setdefault("utility_costs", {})
    normalised: list[dict] = []
    for index, row in enumerate(rows):
        normalised.append(_normalise_utility_entry(row, index))
    if not normalised:
        normalised = [_default_utility_entry(0)]
    utility["years"] = normalised
    for legacy in ("electricity_per_day", "water_per_day", "steam_per_hour", "days", "hours"):
        utility.pop(legacy, None)


def _utility_entries_to_editor(rows: Sequence[Mapping]) -> list[dict]:
    table: list[dict] = []
    for entry in rows:
        normalised = _normalise_utility_entry(entry, len(table))
        record: dict[str, float | int | str] = {}
        for field, label, _ in UTILITY_EDITOR_COLUMNS:
            record[label] = normalised.get(field)
        table.append(record)
    return table


def _editor_rows_to_utility_entries(rows: Sequence[Mapping]) -> list[dict]:
    entries: list[dict] = []
    for index, row in enumerate(rows or []):
        if not isinstance(row, Mapping):
            continue
        payload_row = {}
        for field, label, _ in UTILITY_EDITOR_COLUMNS:
            payload_row[field] = row.get(label)
        entries.append(_normalise_utility_entry(payload_row, index))
    if not entries:
        entries = [_default_utility_entry(0)]
    return entries


def _utility_column_config():
    config = {}
    for field, label, kind in UTILITY_EDITOR_COLUMNS:
        if kind == "text":
            config[label] = st.column_config.TextColumn(label, required=True)
        elif kind == "int":
            config[label] = st.column_config.NumberColumn(label, min_value=0, step=1)
        else:
            config[label] = st.column_config.NumberColumn(label, min_value=0.0, format="%.4f")
    return config


def _extract_editor_rows(data) -> list[Mapping]:
    if data is None:
        return []
    if hasattr(data, "to_dict"):
        try:
            return data.to_dict(orient="records")  # type: ignore[attr-defined]
        except Exception:  # pragma: no cover - fallback when pandas not available
            pass
    if isinstance(data, list):
        return [row for row in data if isinstance(row, Mapping)]
    if isinstance(data, tuple):
        return [row for row in data if isinstance(row, Mapping)]
    return []


def _next_utility_entry(rows: Sequence[Mapping], payload_years: Sequence | None) -> dict:
    index = len(rows)
    label_override: str | None = None
    if isinstance(payload_years, Sequence) and index < len(payload_years):
        candidate = payload_years[index]
        if candidate is not None:
            label_override = str(candidate)
    elif rows:
        last = rows[-1]
        last_year = None
        if isinstance(last, Mapping):
            candidate = last.get("year")
            if isinstance(candidate, (int, float)):
                last_year = int(candidate)
            else:
                last_year = _parse_year_value(last.get("label"), index)
        if last_year is not None:
            label_override = str(last_year + 1)
    return _default_utility_entry(index, label_override)


def _resize_utility_entries(
    entries: Sequence[Mapping], target_length: int, labels: Sequence
) -> list[dict]:
    source = [_normalise_utility_entry(entry, idx) for idx, entry in enumerate(entries or [])]
    rows: list[dict] = []
    for index in range(target_length):
        if index < len(source):
            base = dict(source[index])
        elif source:
            base = dict(source[-1])
        else:
            base = _default_utility_entry(index)
        if index < len(labels):
            override = str(labels[index])
            base["label"] = override
            base["year"] = _parse_year_value(override, index + 1)
        rows.append(_normalise_utility_entry(base, index))
    if not rows:
        rows = [_default_utility_entry(0)]
    return rows




def _normalise_tax_entry(row: Mapping | None, index: int, base_rate: float) -> dict:
    label_value = f"Year {index + 1}"
    rate_value = float(base_rate)

    if isinstance(row, Mapping):
        raw_label = row.get("label") or row.get("Year")
        if raw_label not in (None, ""):
            label_value = str(raw_label).strip() or label_value

        raw_rate = row.get("rate") if "rate" in row else row.get("Rate")
        if raw_rate not in (None, ""):
            try:
                rate_value = float(raw_rate)
            except (TypeError, ValueError):  # pragma: no cover - defensive parsing
                rate_value = float(base_rate)

    return {"label": label_value, "rate": rate_value}


def _tax_entries_to_payload(
    rows: Sequence[Mapping], tax: Mapping, years: Sequence, base_rate: float
) -> None:
    lookup: dict[str, float] = {}
    normalised_rows: list[dict] = []

    for index, row in enumerate(rows or []):
        normalised = _normalise_tax_entry(row, index, base_rate)
        label = normalised["label"]
        rate = float(normalised["rate"])
        normalised_rows.append(normalised)

        lookup[label] = rate
        try:
            numeric_key = str(int(float(label)))
            lookup.setdefault(numeric_key, rate)
        except Exception:  # pragma: no cover - label is not numeric
            pass
        lookup.setdefault(f"Year {index + 1}", rate)

    resolved: list[float] = []
    for index, year in enumerate(years or []):
        key = str(year)
        fallback = f"Year {index + 1}"
        value = lookup.get(key)
        if value is None:
            value = lookup.get(fallback)
        if value is None and index < len(normalised_rows):
            value = float(normalised_rows[index]["rate"])
        if value is None:
            value = float(base_rate)
        resolved.append(float(value))

    if not resolved:
        resolved = [float(base_rate)]

    tax = cast(dict, tax)
    tax["schedule"] = resolved


def _next_tax_entry(
    rows: Sequence[Mapping], years: Sequence | None, base_rate: float
) -> dict:
    used_labels = {
        _normalise_tax_entry(row, index, base_rate)["label"]
        for index, row in enumerate(rows or [])
    }

    candidate_label: str | None = None
    if isinstance(years, Sequence):
        for raw in years:
            if raw is None:
                continue
            label = str(raw)
            if label not in used_labels:
                candidate_label = label
                break

    if candidate_label is None and rows:
        last = _normalise_tax_entry(rows[-1], len(rows) - 1, base_rate)["label"]
        try:
            candidate_label = str(int(float(last)) + 1)
        except Exception:  # pragma: no cover - fallback when last label not numeric
            candidate_label = f"Year {len(rows) + 1}"

    if candidate_label is None:
        candidate_label = f"Year {len(rows) + 1}"

    return {"label": candidate_label, "rate": float(base_rate)}


def _payload_to_depreciation_rows(payload: Mapping) -> list[dict]:
    depreciation = payload.get("depreciation", {})
    rows: list[dict] = []

    if isinstance(depreciation, Mapping):
        stored = depreciation.get("rows")
        if isinstance(stored, Sequence):
            for item in stored:
                if not isinstance(item, Mapping):
                    continue
                rows.append(
                    {
                        "asset_type": str(item.get("asset_type", item.get("asset", ""))),
                        "year": int(item.get("year", 0)),
                        "acquisition": float(item.get("acquisition", 0.0) or 0.0),
                        "depreciation_rate": float(item.get("depreciation_rate", item.get("rate", 0.0)) or 0.0),
                        "asset_life": int(item.get("asset_life", 0) or 0),
                        "method": str(item.get("method", "straight_line") or "straight_line").strip().lower(),
                        "opening_net_book": float(item.get("opening_net_book", 0.0) or 0.0),
                        "opening_cumulative": float(item.get("opening_cumulative", 0.0) or 0.0),
                        "override_net_book": bool(item.get("override_net_book", False)),
                        "override_cumulative": bool(item.get("override_cumulative", False)),
                    }
                )

    if rows:
        normalised: list[dict] = []
        for entry in rows:
            method_value = str(entry.get("method", "straight_line") or "straight_line").strip().lower()
            if method_value not in DEPRECIATION_METHOD_LABELS:
                method_value = "straight_line"
            entry["method"] = method_value
            normalised.append(entry)
        return normalised[:5]

    return _legacy_depreciation_rows(payload)


def _depreciation_rows_to_payload(rows: Sequence[Mapping], payload: dict) -> None:
    depreciation_rows: list[dict] = []
    for row in rows:
        asset = str(row.get("asset_type", "")).strip()
        if not asset:
            continue
        method_value = str(row.get("method", "straight_line") or "straight_line").strip().lower()
        if method_value not in DEPRECIATION_METHOD_LABELS:
            method_value = "straight_line"

        depreciation_rows.append(
            {
                "asset_type": asset,
                "year": int(row.get("year", 0)),
                "acquisition": float(row.get("acquisition", 0.0) or 0.0),
                "depreciation_rate": float(row.get("depreciation_rate", 0.0) or 0.0),
                "asset_life": int(row.get("asset_life", 0) or 0),
                "method": method_value,
                "opening_net_book": float(row.get("opening_net_book", 0.0) or 0.0),
                "opening_cumulative": float(row.get("opening_cumulative", 0.0) or 0.0),
                "override_net_book": bool(row.get("override_net_book", False)),
                "override_cumulative": bool(row.get("override_cumulative", False)),
            }
        )

    payload.setdefault("depreciation", {})["rows"] = depreciation_rows


def _collect_asset_type_options(payload: Mapping, rows: Sequence[Mapping]) -> list[str]:
    catalogue: dict[str, None] = {}

    depreciation = payload.get("depreciation")
    if isinstance(depreciation, Mapping):
        stored_options = depreciation.get("asset_types")
        if isinstance(stored_options, Sequence):
            for value in stored_options:
                label = str(value).strip()
                if label:
                    catalogue[label] = None

        stored_rows = depreciation.get("rows")
        if isinstance(stored_rows, Sequence):
            for item in stored_rows:
                if not isinstance(item, Mapping):
                    continue
                label = str(item.get("asset_type", item.get("asset", "")) or "").strip()
                if label:
                    catalogue[label] = None

    for row in rows:
        label = str(row.get("asset_type", "") or "").strip()
        if label:
            catalogue[label] = None

    return list(catalogue.keys())


def _initialise_session_payload(payload: dict) -> None:
    st.session_state["input_payload"] = payload
    st.session_state["core_assumption_rows"] = _payload_to_core_rows(payload)
    st.session_state["direct_labor_rows"] = _mapping_to_rows(
        payload.get("labor", {}).get("direct", {}),
        "Role",
        "Annual Cost",
    )
    st.session_state["indirect_labor_rows"] = _mapping_to_rows(
        payload.get("labor", {}).get("indirect", {}),
        "Role",
        "Annual Cost",
    )
    st.session_state["utility_entries"] = _payload_to_utility_entries(payload)
    st.session_state["receivable_rows"] = _payload_to_receivable_rows(payload)
    st.session_state["break_even_rows"] = _payload_to_break_even_rows(payload)
    st.session_state["inventory_rows"] = _payload_to_inventory_rows(payload)
    st.session_state["depreciation_rows"] = _payload_to_depreciation_rows(payload)
    st.session_state["sensitivity_rows"] = _payload_to_sensitivity_rows(payload)
    st.session_state["inflation_rows"] = _payload_to_inflation_rows(payload)
    st.session_state["risk_rows"] = _payload_to_risk_rows(payload)
    st.session_state["senior_debt_rows"] = _payload_to_debt_rows(payload, "senior_debt")
    st.session_state["revolver_rows"] = _payload_to_debt_rows(payload, "revolver")
    st.session_state["overdraft_rows"] = _payload_to_debt_rows(payload, "overdraft")
    st.session_state["ai_settings"] = _payload_to_ai_settings(payload)
    st.session_state["ai_api_key"] = st.session_state["ai_settings"].get("api_key", "")


def _payload_to_core_rows(payload: Mapping) -> list[dict]:
    unit_costs = payload.get("unit_costs", {})
    markup = payload.get("markup", {})
    totals = payload.get("total_production_units", {})
    capacities = payload.get("production_capacity", {})
    estimates = payload.get("production_estimate", {})
    years = payload.get("years", [])
    inflation_factors = _inflation_factors_from_payload(payload)
    risk_factors = _risk_factors_from_payload(payload)
    rows: list[dict] = []
    for name, values in unit_costs.items():
        production_cost = float(values.get("production", 0.0))
        selling_price = float(values.get("price", 0.0))
        freight_cost = float(values.get("freight", 0.0))
        markup_value = float(markup.get(name, 0.0))
        total_units = float(totals.get(name, 0.0))
        if total_units == 0.0 and isinstance(estimates, Mapping):
            estimate = estimates.get(name, [])
            total_units = sum(float(v) for v in estimate)
        max_capacity = float(capacities.get(name, 0.0))
        if max_capacity > 0.0 and total_units > max_capacity:
            total_units = max_capacity
        scaled_series = _scaled_production_series(name, total_units, years, estimates)
        first_year_units = scaled_series[0] if scaled_series else 0.0
        inflation_factor = inflation_factors[0] if inflation_factors else 1.0
        risk_factor = risk_factors[0] if risk_factors else 1.0
        total_revenue = first_year_units * selling_price * inflation_factor * risk_factor
        total_cost = (
            first_year_units
            * (production_cost + freight_cost + markup_value)
            * inflation_factor
            * risk_factor
        )
        rows.append(
            {
                "Product": str(name),
                "Production Cost": production_cost,
                "Selling Price": selling_price,
                "Freight Cost": freight_cost,
                "Markup": markup_value,
                "Total Production Units": total_units,
                "Max Capacity": max_capacity,
                "Total Revenue": total_revenue,
                "Total Cost": total_cost,
            }
        )
    return rows


def _payload_to_receivable_rows(payload: Mapping) -> list[dict]:
    years = list(payload.get("years", []))
    working = payload.get("working_capital", {})
    if not isinstance(working, Mapping):
        working = {}

    raw_days = working.get("days") if isinstance(working.get("days"), Mapping) else {}
    if not isinstance(raw_days, Mapping):
        raw_days = {}

    receivable_source = raw_days.get("accounts_receivable", [])
    receivable_days = (
        [int(float(value)) for value in receivable_source]
        if isinstance(receivable_source, Iterable)
        and not isinstance(receivable_source, (str, bytes))
        else []
    )

    prepaid_source = raw_days.get("prepaid_expenses", [])
    prepaid_days = (
        [int(float(value)) for value in prepaid_source]
        if isinstance(prepaid_source, Iterable)
        and not isinstance(prepaid_source, (str, bytes))
        else []
    )

    other_asset_source = raw_days.get("other_assets", [])
    other_asset_days = (
        [int(float(value)) for value in other_asset_source]
        if isinstance(other_asset_source, Iterable)
        and not isinstance(other_asset_source, (str, bytes))
        else []
    )

    calendar_source = working.get("calendar_days", [])
    calendar_days = (
        [int(float(value)) for value in calendar_source]
        if isinstance(calendar_source, Iterable)
        and not isinstance(calendar_source, (str, bytes))
        else []
    )

    max_length = max(
        len(years),
        len(receivable_days),
        len(prepaid_days),
        len(other_asset_days),
        len(calendar_days),
        1,
    )

    rows: list[dict] = []

    for index in range(max_length):
        if index < len(years):
            year_value = int(years[index])
            label = str(year_value)
        else:
            year_value = index + 1
            label = f"Year {index + 1}"

        if index < len(calendar_days):
            days_in_year = int(calendar_days[index])
        else:
            days_in_year = 366 if _is_leap_year(year_value) else 365

        receivable_day = int(receivable_days[index]) if index < len(receivable_days) else 0
        prepaid_day = int(prepaid_days[index]) if index < len(prepaid_days) else 0
        other_day = int(other_asset_days[index]) if index < len(other_asset_days) else 0

        rows.append(
            {
                "label": label,
                "year": year_value,
                "days_in_year": days_in_year,
                "accounts_receivable_days": receivable_day,
                "prepaid_expense_days": prepaid_day,
                "other_asset_days": other_day,
            }
        )

    return rows


def _receivable_rows_to_payload(rows: Sequence[Mapping], payload: dict) -> None:
    if rows is None:
        return

    working = payload.setdefault("working_capital", {})
    if not isinstance(working, dict):
        working = {}
        payload["working_capital"] = working

    days_mapping = working.get("days") if isinstance(working.get("days"), Mapping) else {}
    if not isinstance(days_mapping, dict):
        days_mapping = {}
    working["days"] = days_mapping

    if not rows:
        days_mapping["accounts_receivable"] = []
        days_mapping["prepaid_expenses"] = []
        days_mapping["other_assets"] = []
        working["calendar_days"] = []
        return

    labels: list[str] = []
    calendar_days: list[int] = []
    receivable_days: list[int] = []
    prepaid_days: list[int] = []
    other_asset_days: list[int] = []

    for index, row in enumerate(rows):
        label = str(row.get("label", row.get("Year", "")) or "").strip()
        if not label:
            label = f"Year {index + 1}"
        labels.append(label)

        try:
            calendar_value = int(float(row.get("days_in_year", row.get("Days in Year", 0)) or 0))
        except (TypeError, ValueError):
            calendar_value = 0
        calendar_days.append(calendar_value)

        try:
            receivable_value = int(
                float(
                    row.get(
                        "accounts_receivable_days",
                        row.get("Accounts Receivable Days", 0),
                    )
                    or 0
                )
            )
        except (TypeError, ValueError):
            receivable_value = 0
        receivable_days.append(receivable_value)

        try:
            prepaid_value = int(
                float(
                    row.get("prepaid_expense_days", row.get("Prepaid Expenses Days", 0))
                    or 0
                )
            )
        except (TypeError, ValueError):
            prepaid_value = 0
        prepaid_days.append(prepaid_value)

        try:
            other_value = int(
                float(row.get("other_asset_days", row.get("Other Assets Days", 0)) or 0)
            )
        except (TypeError, ValueError):
            other_value = 0
        other_asset_days.append(other_value)

    days_mapping["accounts_receivable"] = receivable_days
    days_mapping["prepaid_expenses"] = prepaid_days
    days_mapping["other_assets"] = other_asset_days
    working["calendar_days"] = calendar_days

    current_years = payload.get("years", [])
    target_length = max(len(current_years), len(rows))
    if target_length:
        _align_payload_horizon(payload, labels, target_length)

    try:
        st.session_state["receivable_rows"] = _payload_to_receivable_rows(payload)
        if "inventory_rows" in st.session_state:
            st.session_state["inventory_rows"] = _payload_to_inventory_rows(payload)
    except Exception:  # pragma: no cover - depends on Streamlit runtime
        pass


def _payload_to_inventory_rows(payload: Mapping) -> list[dict]:
    years = list(payload.get("years", []))
    working = payload.get("working_capital", {})
    if not isinstance(working, Mapping):
        working = {}

    raw_days = working.get("days") if isinstance(working.get("days"), Mapping) else {}
    if not isinstance(raw_days, Mapping):
        raw_days = {}

    inventory_source = raw_days.get("inventory", [])
    if isinstance(inventory_source, Iterable) and not isinstance(inventory_source, (str, bytes)):
        inventory_days = [int(float(value)) for value in inventory_source]
    else:
        inventory_days = []

    payable_source = raw_days.get("accounts_payable", [])
    if isinstance(payable_source, Iterable) and not isinstance(payable_source, (str, bytes)):
        payable_days = [int(float(value)) for value in payable_source]
    else:
        payable_days = []

    calendar_source = working.get("calendar_days", [])
    if isinstance(calendar_source, Iterable) and not isinstance(calendar_source, (str, bytes)):
        calendar_days = [int(float(value)) for value in calendar_source]
    else:
        calendar_days = []

    max_length = max(len(years), len(inventory_days), len(payable_days), len(calendar_days), 1)
    rows: list[dict] = []

    for index in range(max_length):
        if index < len(years):
            year_value = int(years[index])
            label = str(year_value)
        else:
            year_value = index + 1
            label = f"Year {index + 1}"

        if index < len(calendar_days):
            days_in_year = int(calendar_days[index])
        else:
            if year_value % 400 == 0 or (year_value % 4 == 0 and year_value % 100 != 0):
                days_in_year = 366
            else:
                days_in_year = 365

        if index < len(inventory_days):
            inventory_day = int(inventory_days[index])
        else:
            inventory_day = 0

        if index < len(payable_days):
            payable_day = int(payable_days[index])
        else:
            payable_day = 0

        rows.append(
            {
                "label": label,
                "year": year_value,
                "days_in_year": days_in_year,
                "inventory_days": inventory_day,
                "accounts_payable_days": payable_day,
            }
        )

    return rows


def _payload_to_break_even_rows(payload: Mapping) -> list[dict]:
    section = payload.get("break_even") if isinstance(payload, Mapping) else None
    if not isinstance(section, Mapping):
        return []

    rows: list[dict] = []
    for entry in section.get("rows", []):
        if not isinstance(entry, Mapping):
            continue
        product = str(entry.get("product") or entry.get("Product") or "").strip()
        if not product:
            continue
        rows.append(
            {
                "Product": product,
                "Fixed Cost": float(entry.get("fixed_cost", entry.get("Fixed Cost", 0.0)) or 0.0),
                "Selling Price": float(entry.get("selling_price", entry.get("Selling Price", 0.0)) or 0.0),
                "Variable Cost": float(entry.get("variable_cost", entry.get("Variable Cost", 0.0)) or 0.0),
                "Target Profit": float(entry.get("target_profit", entry.get("Target Profit", 0.0)) or 0.0),
                "Expected Volume": float(entry.get("expected_volume", entry.get("Expected Volume", 0.0)) or 0.0),
            }
        )

    return rows


def _break_even_rows_to_payload(rows: Sequence[Mapping], payload: dict) -> None:
    if rows is None:
        return

    section = payload.setdefault("break_even", {})
    if not isinstance(section, dict):
        section = {}
        payload["break_even"] = section

    serialised: list[dict] = []
    for row in rows:
        if not isinstance(row, Mapping):
            continue
        product = str(row.get("Product", "")).strip()
        if not product:
            continue
        serialised.append(
            {
                "product": product,
                "fixed_cost": float(row.get("Fixed Cost", 0.0) or 0.0),
                "selling_price": float(row.get("Selling Price", 0.0) or 0.0),
                "variable_cost": float(row.get("Variable Cost", 0.0) or 0.0),
                "target_profit": float(row.get("Target Profit", 0.0) or 0.0),
                "expected_volume": float(row.get("Expected Volume", 0.0) or 0.0),
            }
        )

    section["rows"] = serialised


def _default_break_even_rows(payload: Mapping) -> list[dict]:
    try:
        inputs = parse_inputs(payload)
        model = FinancialModel(inputs)
        table = model.break_even_analysis()
    except Exception:
        table = None
        inputs = None

    rows: list[dict] = []
    if table is not None:
        data = table.data
        fixed_values = data.get("Fixed Cost", [])
        selling = data.get("Selling Price", [])
        variable = data.get("Variable Cost per Unit", [])
        target = data.get("Target Profit", [])
        expected = data.get("Expected Volume", [])
        break_even_units = data.get("Break-even Units", data.get("Units", []))

        for idx, product in enumerate(table.index):
            price = float(selling[idx]) if idx < len(selling) else 0.0
            variable_cost = float(variable[idx]) if idx < len(variable) else 0.0
            target_profit = float(target[idx]) if idx < len(target) else 0.0
            expected_volume = float(expected[idx]) if idx < len(expected) else 0.0
            margin = price - variable_cost
            units = float(break_even_units[idx]) if idx < len(break_even_units) else float("nan")
            if idx < len(fixed_values):
                try:
                    fixed_cost = float(fixed_values[idx])
                except (TypeError, ValueError):
                    fixed_cost = 0.0
            elif units == units and margin > 0:
                fixed_cost = max(units * margin - target_profit, 0.0)
            else:
                fixed_cost = 0.0

            rows.append(
                {
                    "Product": product,
                    "Fixed Cost": fixed_cost,
                    "Selling Price": price,
                    "Variable Cost": variable_cost,
                    "Target Profit": target_profit,
                    "Expected Volume": expected_volume,
                }
            )

    if inputs is None and not rows:
        unit_costs = payload.get("unit_costs", {}) if isinstance(payload, Mapping) else {}
        totals = payload.get("total_production_units", {}) if isinstance(payload, Mapping) else {}
        if isinstance(unit_costs, Mapping):
            for product, values in unit_costs.items():
                if not isinstance(values, Mapping):
                    continue
                rows.append(
                    {
                        "Product": str(product),
                        "Fixed Cost": 0.0,
                        "Selling Price": float(values.get("price", 0.0) or 0.0),
                        "Variable Cost": float(values.get("production", 0.0) or 0.0)
                        + float(values.get("freight", 0.0) or 0.0),
                        "Target Profit": 0.0,
                        "Expected Volume": float(totals.get(product, 0.0) or 0.0),
                    }
                )

    if inputs is not None and isinstance(payload, Mapping):
        unit_costs = payload.get("unit_costs", {})
        totals = payload.get("total_production_units", {})
        known = {row["Product"] for row in rows}
        if isinstance(unit_costs, Mapping):
            for product, values in unit_costs.items():
                if product in known or not isinstance(values, Mapping):
                    continue
                rows.append(
                    {
                        "Product": str(product),
                        "Fixed Cost": 0.0,
                        "Selling Price": float(values.get("price", 0.0) or 0.0),
                        "Variable Cost": float(values.get("production", 0.0) or 0.0)
                        + float(values.get("freight", 0.0) or 0.0),
                        "Target Profit": 0.0,
                        "Expected Volume": float(totals.get(product, 0.0) or 0.0),
                    }
                )

    return rows




def _row_matches_default(row: Mapping[str, float], default_row: Mapping[str, float]) -> bool:
    keys = ("Fixed Cost", "Selling Price", "Variable Cost", "Target Profit", "Expected Volume")
    for key in keys:
        actual = float(row.get(key, 0.0) or 0.0)
        reference = float(default_row.get(key, 0.0) or 0.0)
        if not math.isclose(actual, reference, rel_tol=1e-9, abs_tol=1e-6):
            return False
    return True

def _calculate_break_even_metrics(row: Mapping[str, float]) -> dict[str, float]:
    fixed_cost = float(row.get("Fixed Cost", 0.0) or 0.0)
    selling_price = float(row.get("Selling Price", 0.0) or 0.0)
    variable_cost = float(row.get("Variable Cost", 0.0) or 0.0)
    target_profit = float(row.get("Target Profit", 0.0) or 0.0)
    expected_volume = float(row.get("Expected Volume", 0.0) or 0.0)

    contribution = selling_price - variable_cost
    ratio = contribution / selling_price if selling_price else float("nan")

    if contribution <= 0:
        break_even_units = float("nan")
        break_even_revenue = float("nan")
    else:
        break_even_units = (fixed_cost + target_profit) / contribution
        break_even_revenue = break_even_units * selling_price

    if expected_volume > 0 and break_even_units == break_even_units:
        margin_of_safety_units = expected_volume - break_even_units
        margin_of_safety_pct = margin_of_safety_units / expected_volume
    else:
        margin_of_safety_units = float("nan")
        margin_of_safety_pct = float("nan")

    return {
        "Contribution Margin": contribution,
        "Contribution Margin Ratio": ratio,
        "Break-even Units": break_even_units,
        "Break-even Revenue": break_even_revenue,
        "Margin of Safety (Units)": margin_of_safety_units,
        "Margin of Safety (%)": margin_of_safety_pct,
    }


def _aggregate_break_even_metrics(rows: Sequence[Mapping[str, float]]) -> dict[str, float]:
    if not rows:
        return {}

    total_fixed = 0.0
    total_expected_revenue = 0.0
    total_expected_contribution = 0.0
    aggregate_break_even_revenue = 0.0

    for row in rows:
        metrics = _calculate_break_even_metrics(row)
        fixed_cost = float(row.get("Fixed Cost", 0.0) or 0.0)
        selling_price = float(row.get("Selling Price", 0.0) or 0.0)
        variable_cost = float(row.get("Variable Cost", 0.0) or 0.0)
        expected_volume = float(row.get("Expected Volume", 0.0) or 0.0)

        total_fixed += fixed_cost
        total_expected_revenue += selling_price * expected_volume
        total_expected_contribution += (selling_price - variable_cost) * expected_volume

        if metrics["Break-even Revenue"] == metrics["Break-even Revenue"]:
            aggregate_break_even_revenue += metrics["Break-even Revenue"]

    weighted_margin_ratio = float("nan")
    if total_expected_revenue > 0:
        weighted_margin_ratio = total_expected_contribution / total_expected_revenue

    return {
        "total_fixed": total_fixed,
        "weighted_margin_ratio": weighted_margin_ratio,
        "aggregate_break_even_revenue": aggregate_break_even_revenue,
    }


def _inventory_rows_to_payload(rows: Sequence[Mapping], payload: dict) -> None:
    if rows is None:
        return

    working = payload.setdefault("working_capital", {})
    if not isinstance(working, dict):
        working = {}
        payload["working_capital"] = working

    days_mapping = working.get("days") if isinstance(working.get("days"), Mapping) else {}
    if not isinstance(days_mapping, dict):
        days_mapping = {}
    working["days"] = days_mapping

    if not rows:
        days_mapping["inventory"] = []
        days_mapping["accounts_payable"] = []
        working["calendar_days"] = []
        return

    labels: list[str] = []
    calendar_days: list[int] = []
    inventory_days: list[int] = []
    accounts_payable_days: list[int] = []

    for index, row in enumerate(rows):
        label = str(row.get("label", row.get("Year", "")) or "").strip()
        if not label:
            label = f"Year {index + 1}"
        labels.append(label)

        try:
            day_value = int(float(row.get("days_in_year", row.get("Days in Year", 0)) or 0))
        except (TypeError, ValueError):
            day_value = 0
        calendar_days.append(day_value)

        try:
            inventory_value = int(float(row.get("inventory_days", row.get("Inventory Days", 0)) or 0))
        except (TypeError, ValueError):
            inventory_value = 0
        inventory_days.append(inventory_value)

        try:
            payable_value = int(
                float(
                    row.get(
                        "accounts_payable_days",
                        row.get("Accounts Payable Days", 0),
                    )
                    or 0
                )
            )
        except (TypeError, ValueError):
            payable_value = 0
        accounts_payable_days.append(payable_value)

    days_mapping["inventory"] = inventory_days
    days_mapping["accounts_payable"] = accounts_payable_days
    working["calendar_days"] = calendar_days

    current_years = payload.get("years", [])
    target_length = max(len(current_years), len(rows))
    if target_length:
        _align_payload_horizon(payload, labels, target_length)

    try:
        st.session_state["inventory_rows"] = _payload_to_inventory_rows(payload)
    except Exception:  # pragma: no cover - depends on Streamlit runtime
        pass


def _prime_core_widget_state(rows: Sequence[Mapping]) -> None:
    """Synchronise core assumption widget keys with the supplied rows."""

    try:
        for index, row in enumerate(rows):
            _set_widget_value(f"core_desc_{index}", str(row.get("Product", "")))
            _set_widget_value(
                f"core_prod_{index}", float(row.get("Production Cost", 0.0))
            )
            _set_widget_value(
                f"core_sell_{index}", float(row.get("Selling Price", 0.0))
            )
            _set_widget_value(
                f"core_freight_{index}", float(row.get("Freight Cost", 0.0))
            )
            _set_widget_value(
                f"core_markup_{index}", float(row.get("Markup", 0.0))
            )
            _set_widget_value(
                f"core_units_{index}", float(row.get("Total Production Units", 0.0))
            )
            _set_widget_value(
                f"core_capacity_{index}", float(row.get("Max Capacity", 0.0))
            )
    except Exception:  # pragma: no cover - depends on Streamlit runtime
        pass


def _sync_core_rows_from_widgets(rows: Sequence[Mapping]) -> list[dict]:
    """Return core assumption rows updated from widget session state.

    When users edit the number inputs on the Streamlit page the widget values are
    stored in ``st.session_state`` before the next script run.  The model is
    parsed prior to rendering the input tab, so we normalise the stored rows
    against the widget state here to ensure the latest user edits feed straight
    into the payload digest that powers the financial outputs.
    """

    updated_rows: list[dict] = []
    for index, row in enumerate(rows):
        current = dict(row)

        description = st.session_state.get(f"core_desc_{index}")
        if description is not None:
            current["Product"] = str(description).strip()

        production = st.session_state.get(f"core_prod_{index}")
        if production is not None:
            current["Production Cost"] = float(production)

        selling = st.session_state.get(f"core_sell_{index}")
        if selling is not None:
            current["Selling Price"] = float(selling)

        freight = st.session_state.get(f"core_freight_{index}")
        if freight is not None:
            current["Freight Cost"] = float(freight)

        markup = st.session_state.get(f"core_markup_{index}")
        if markup is not None:
            current["Markup"] = float(markup)

        units = st.session_state.get(f"core_units_{index}")
        if units is not None:
            current["Total Production Units"] = float(units)

        capacity = st.session_state.get(f"core_capacity_{index}")
        if capacity is not None:
            current["Max Capacity"] = float(capacity)

        total_units = float(current.get("Total Production Units", 0.0))
        max_capacity = float(current.get("Max Capacity", 0.0))
        if max_capacity > 0.0 and total_units > max_capacity:
            total_units = max_capacity
        current["Total Production Units"] = total_units

        production_cost = float(current.get("Production Cost", 0.0))
        selling_price = float(current.get("Selling Price", 0.0))
        freight_cost = float(current.get("Freight Cost", 0.0))
        markup_value = float(current.get("Markup", 0.0))

        current["Total Revenue"] = total_units * selling_price
        current["Total Cost"] = total_units * (production_cost + freight_cost + markup_value)

        updated_rows.append(current)

    return updated_rows


def _core_rows_to_payload(rows: Sequence[Mapping], payload: dict) -> None:
    unit_costs: dict[str, dict[str, float]] = {}
    markup: dict[str, float] = {}
    years = payload.get("years", [])
    existing_estimate = payload.get("production_estimate", {})
    production_estimate: dict[str, list[float]] = {}
    total_units_map: dict[str, float] = {}
    capacity_map: dict[str, float] = {}

    for row in rows:
        name = str(row.get("Product", "")).strip()
        if not name:
            continue
        unit_costs[name] = {
            "production": float(row.get("Production Cost", 0.0)),
            "price": float(row.get("Selling Price", 0.0)),
            "freight": float(row.get("Freight Cost", 0.0)),
        }
        markup[name] = float(row.get("Markup", 0.0))
        max_capacity = float(row.get("Max Capacity", 0.0))
        total_units = float(row.get("Total Production Units", 0.0))
        if max_capacity > 0.0 and total_units > max_capacity:
            total_units = max_capacity
        total_units_map[name] = total_units
        capacity_map[name] = max_capacity

        scaled = _scaled_production_series(name, total_units, years, existing_estimate)
        production_estimate[name] = scaled

    payload["unit_costs"] = unit_costs
    payload["markup"] = markup
    if production_estimate:
        payload["production_estimate"] = production_estimate
    if total_units_map:
        payload["total_production_units"] = total_units_map
    else:
        payload.pop("total_production_units", None)
    if capacity_map:
        payload["production_capacity"] = capacity_map
    else:
        payload.pop("production_capacity", None)


def _get_widget_value(key: str, default: Any) -> Any:
    if key in st.session_state:
        return st.session_state[key]
    return default


def _get_widget_number(key: str, default: Any, cast: Callable[[Any], Any]) -> Any:
    value = _get_widget_value(key, default)
    try:
        return cast(value)
    except (TypeError, ValueError):
        return cast(default)


def _read_select_value(prefix: str, fallback: str | None = None) -> str:
    selection = _get_widget_value(f"{prefix}_select", fallback or "")
    if selection == "Add new…":
        custom = _get_widget_value(f"{prefix}_custom", fallback or "")
        return str(custom or "").strip()
    return str(selection or fallback or "").strip()


def _sync_commission_rows_from_widgets(rows: Sequence[Mapping]) -> list[dict]:
    updated: list[dict] = []
    visible = min(len(rows), MAX_VISIBLE_COMMISSION_ROWS)
    for index, row in enumerate(rows):
        current = dict(row)
        if index < visible:
            current["Year"] = int(
                _get_widget_number(f"commission_year_{index}", current.get("Year", 0), int)
            )
            current["Product"] = _read_select_value(
                f"commission_product_{index}", current.get("Product", "")
            )
            current["Yearly Commission %"] = float(
                _get_widget_number(
                    f"commission_increment_{index}",
                    current.get("Yearly Commission %", 0.0),
                    float,
                )
            )
            current["Revenue"] = float(
                _get_widget_number(
                    f"commission_revenue_{index}", current.get("Revenue", 0.0), float
                )
            )
        updated.append(current)
    return updated


def _sync_utility_entries_from_widgets(entries: Sequence[Mapping]) -> list[dict]:
    updated: list[dict] = []
    for index, entry in enumerate(entries or []):
        normalised = _normalise_utility_entry(entry, index)
        label = _read_select_value(
            f"utility_label_{index}", normalised.get("label", f"Year {index + 1}")
        )
        parsed_year = _parse_year_value(label, normalised.get("year", index + 1))
        updated.append(
            _normalise_utility_entry(
                {
                    "label": label,
                    "year": parsed_year,
                    "electricity_per_day": float(
                        _get_widget_number(
                            f"utility_elec_per_day_{index}",
                            normalised.get("electricity_per_day", 0.0),
                            float,
                        )
                    ),
                    "electricity_rate": float(
                        _get_widget_number(
                            f"utility_elec_rate_{index}",
                            normalised.get("electricity_rate", 0.0),
                            float,
                        )
                    ),
                    "electricity_days": int(
                        _get_widget_number(
                            f"utility_elec_days_{index}",
                            normalised.get("electricity_days", 0),
                            int,
                        )
                    ),
                    "water_per_day": float(
                        _get_widget_number(
                            f"utility_water_per_day_{index}",
                            normalised.get("water_per_day", 0.0),
                            float,
                        )
                    ),
                    "water_rate": float(
                        _get_widget_number(
                            f"utility_water_rate_{index}",
                            normalised.get("water_rate", 0.0),
                            float,
                        )
                    ),
                    "water_days": int(
                        _get_widget_number(
                            f"utility_water_days_{index}",
                            normalised.get("water_days", 0),
                            int,
                        )
                    ),
                    "steam_per_hour": float(
                        _get_widget_number(
                            f"utility_steam_per_hour_{index}",
                            normalised.get("steam_per_hour", 0.0),
                            float,
                        )
                    ),
                    "steam_rate": float(
                        _get_widget_number(
                            f"utility_steam_rate_{index}",
                            normalised.get("steam_rate", 0.0),
                            float,
                        )
                    ),
                    "steam_days": int(
                        _get_widget_number(
                            f"utility_steam_days_{index}",
                            normalised.get("steam_days", 0),
                            int,
                        )
                    ),
                    "steam_hours": int(
                        _get_widget_number(
                            f"utility_steam_hours_{index}",
                            normalised.get("steam_hours", 0),
                            int,
                        )
                    ),
                },
                index,
            )
        )
    return updated


def _sync_receivable_rows_from_widgets(
    rows: Sequence[Mapping], _payload: Mapping
) -> list[dict]:
    updated = [dict(row) for row in rows]
    for slot in range(MAX_VISIBLE_RECEIVABLE_ROWS):
        selector_key = f"receivable_row_selector_{slot}"
        selected_index = st.session_state.get(selector_key)
        if not isinstance(selected_index, int) or not (0 <= selected_index < len(updated)):
            continue
        current = dict(updated[selected_index])
        label = _read_select_value(
            f"receivable_label_{slot}_{selected_index}", current.get("label") or current.get("Year")
        )
        parsed_year = _parse_year_value(label, current.get("year") or current.get("Year") or slot + 1)
        current.update(
            {
                "label": label,
                "year": parsed_year,
                "days_in_year": int(
                    _get_widget_number(
                        f"receivable_days_in_year_{slot}_{selected_index}",
                        current.get("days_in_year", 365),
                        int,
                    )
                ),
                "accounts_receivable_days": int(
                    _get_widget_number(
                        f"receivable_accounts_receivable_days_{slot}_{selected_index}",
                        current.get("accounts_receivable_days", 0),
                        int,
                    )
                ),
                "prepaid_expense_days": int(
                    _get_widget_number(
                        f"receivable_prepaid_days_{slot}_{selected_index}",
                        current.get("prepaid_expense_days", 0),
                        int,
                    )
                ),
                "other_asset_days": int(
                    _get_widget_number(
                        f"receivable_other_asset_days_{slot}_{selected_index}",
                        current.get("other_asset_days", 0),
                        int,
                    )
                ),
            }
        )
        updated[selected_index] = current
    return updated


def _sync_inventory_rows_from_widgets(
    rows: Sequence[Mapping], _payload: Mapping
) -> list[dict]:
    updated = [dict(row) for row in rows]
    for slot in range(MAX_VISIBLE_INVENTORY_ROWS):
        selector_key = f"inventory_row_selector_{slot}"
        selected_index = st.session_state.get(selector_key)
        if not isinstance(selected_index, int) or not (0 <= selected_index < len(updated)):
            continue
        current = dict(updated[selected_index])
        label = _read_select_value(
            f"inventory_label_{slot}_{selected_index}", current.get("label") or current.get("Year")
        )
        parsed_year = _parse_year_value(label, current.get("year") or current.get("Year") or slot + 1)
        current.update(
            {
                "label": label,
                "year": parsed_year,
                "days_in_year": int(
                    _get_widget_number(
                        f"inventory_days_in_year_{slot}_{selected_index}",
                        current.get("days_in_year", 365),
                        int,
                    )
                ),
                "inventory_days": int(
                    _get_widget_number(
                        f"inventory_inventory_days_{slot}_{selected_index}",
                        current.get("inventory_days", 0),
                        int,
                    )
                ),
                "accounts_payable_days": int(
                    _get_widget_number(
                        f"inventory_accounts_payable_days_{slot}_{selected_index}",
                        current.get("accounts_payable_days", 0),
                        int,
                    )
                ),
            }
        )
        updated[selected_index] = current
    return updated


def _sync_labor_rows_from_widgets(state_key: str, rows: Sequence[Mapping]) -> list[dict]:
    updated: list[dict] = []
    for index, row in enumerate(rows):
        updated.append(
            {
                "Role": str(
                    _get_widget_value(f"{state_key}_role_{index}", row.get("Role", ""))
                ).strip(),
                "Annual Cost": float(
                    _get_widget_number(
                        f"{state_key}_cost_{index}", row.get("Annual Cost", 0.0), float
                    )
                ),
            }
        )
    return updated


def _labor_rows_to_payload(section: str, rows: Sequence[Mapping], payload: dict) -> None:
    labor = payload.setdefault("labor", {})
    if not isinstance(labor, dict):
        labor = {}
        payload["labor"] = labor
    labor[section] = {
        str(row.get("Role", "")).strip(): float(row.get("Annual Cost", 0.0) or 0.0)
        for row in rows
        if row.get("Role")
    }


def _sync_fixed_variable_rows_from_widgets(rows: Sequence[Mapping]) -> list[dict]:
    updated: list[dict] = []
    visible = min(len(rows), MAX_VISIBLE_COST_ROWS)
    for index, row in enumerate(rows):
        current = dict(row)
        if index < visible:
            current["Product"] = _read_select_value(
                f"fixed_variable_product_{index}", current.get("Product", "")
            )
            current["Fixed Cost"] = float(
                _get_widget_number(
                    f"fixed_variable_fixed_{index}", current.get("Fixed Cost", 0.0), float
                )
            )
            current["Variable Cost"] = float(
                _get_widget_number(
                    f"fixed_variable_variable_{index}", current.get("Variable Cost", 0.0), float
                )
            )
        updated.append(current)
    return updated


def _sync_break_even_rows_from_widgets(rows: Sequence[Mapping]) -> list[dict]:
    updated: list[dict] = []
    for index, row in enumerate(rows):
        current = dict(row)
        current["Product"] = _read_select_value(
            f"break_even_product_{index}", current.get("Product", "")
        )
        current["Selling Price"] = float(
            _get_widget_number(
                f"break_even_price_{index}", current.get("Selling Price", 0.0), float
            )
        )
        current["Fixed Cost"] = float(
            _get_widget_number(
                f"break_even_fixed_{index}", current.get("Fixed Cost", 0.0), float
            )
        )
        current["Variable Cost"] = float(
            _get_widget_number(
                f"break_even_variable_{index}", current.get("Variable Cost", 0.0), float
            )
        )
        current["Target Profit"] = float(
            _get_widget_number(
                f"break_even_target_{index}", current.get("Target Profit", 0.0), float
            )
        )
        current["Expected Volume"] = float(
            _get_widget_number(
                f"break_even_volume_{index}", current.get("Expected Volume", 0.0), float
            )
        )
        updated.append(current)
    return updated


def _sync_depreciation_rows_from_widgets(
    rows: Sequence[Mapping], payload: Mapping
) -> list[dict]:
    updated: list[dict] = []
    years = payload.get("years", []) or []
    for index, row in enumerate(rows):
        current = dict(row)
        current["asset_type"] = _read_select_value(
            f"dep_asset_{index}", current.get("asset_type", "")
        )
        method_label = _get_widget_value(f"dep_method_{index}", None)
        if isinstance(method_label, str):
            current["method"] = DEPRECIATION_LABEL_TO_VALUE.get(
                method_label, current.get("method", "straight_line")
            )
        current["year"] = int(
            _get_widget_number(
                f"dep_year_{index}", current.get("year", years[index] if index < len(years) else 0), int
            )
        )
        current["acquisition"] = float(
            _get_widget_number(f"dep_acq_{index}", current.get("acquisition", 0.0), float)
        )
        current["asset_life"] = int(
            _get_widget_number(f"dep_life_{index}", current.get("asset_life", 0), int)
        )
        current["depreciation_rate"] = float(
            _get_widget_number(f"dep_rate_{index}", current.get("depreciation_rate", 0.0), float)
        )
        current["opening_net_book"] = float(
            _get_widget_number(
                f"dep_open_nb_{index}", current.get("opening_net_book", 0.0), float
            )
        )
        current["total_asset_cost"] = float(
            _get_widget_number(
                f"dep_total_cost_{index}", current.get("total_asset_cost", 0.0), float
            )
        )
        current["total_depreciation"] = float(
            _get_widget_number(
                f"dep_total_dep_{index}", current.get("total_depreciation", 0.0), float
            )
        )
        current["cumulative_depreciation"] = float(
            _get_widget_number(
                f"dep_cum_dep_{index}", current.get("cumulative_depreciation", 0.0), float
            )
        )
        current["net_book_value"] = float(
            _get_widget_number(
                f"dep_net_book_{index}", current.get("net_book_value", 0.0), float
            )
        )
        updated.append(current)
    return updated


def _sync_inflation_rows_from_widgets(
    rows: Sequence[Mapping], payload: Mapping
) -> list[dict]:
    updated: list[dict] = []
    for index, row in enumerate(rows):
        current = dict(row)
        current["Year"] = _read_select_value(
            f"inflation_label_{index}", current.get("Year")
        )
        current["Rate"] = float(
            _get_widget_number(
                f"inflation_rate_{index}", current.get("Rate", 0.0), float
            )
        )
        updated.append(current)
    return updated


def _sync_risk_rows_from_widgets(
    rows: Sequence[Mapping], payload: Mapping
) -> list[dict]:
    updated: list[dict] = []
    categories = _risk_categories(payload, rows)
    for index, row in enumerate(rows):
        current = {"Year": _read_select_value(
            f"risk_label_{index}", row.get("Year")
        )}
        for category in categories:
            current[category] = float(
                _get_widget_number(
                    f"risk_{category}_{index}", row.get(category, 0.0), float
                )
            )
        updated.append(current)
    return updated


def _sync_sensitivity_rows_from_widgets(rows: Sequence[Mapping]) -> list[dict]:
    updated: list[dict] = []
    for index, row in enumerate(rows):
        variable = str(_get_widget_value(f"sensitivity_var_{index}", row.get("Variable", ""))).strip()
        values_text = _get_widget_value(
            f"sensitivity_vals_{index}", _format_float_list(row.get("Values", []))
        )
        try:
            values = _parse_float_list(values_text)
        except ValueError:
            values = row.get("Values", [])
        updated.append({"Variable": variable, "Values": values})
    return updated


def _sync_debt_rows_from_widgets(debt_type: str, rows: Sequence[Mapping]) -> list[dict]:
    session_key = {
        "senior_debt": "senior_debt_rows",
        "revolver": "revolver_rows",
        "overdraft": "overdraft_rows",
    }.get(debt_type, f"{debt_type}_rows")
    updated: list[dict] = []
    for index, row in enumerate(rows):
        current = dict(row)
        current["Year"] = int(
            _get_widget_number(f"{session_key}_year_{index}", current.get("Year", 0), int)
        )
        current["Duration"] = int(
            _get_widget_number(
                f"{session_key}_duration_{index}", current.get("Duration", row.get("Duration", 1) or 1), int
            )
        )
        current["Amount"] = float(
            _get_widget_number(f"{session_key}_amount_{index}", current.get("Amount", 0.0), float)
        )
        outstanding_key = f"{session_key}_outstanding_{index}"
        if outstanding_key in st.session_state:
            current["Outstanding"] = float(
                _get_widget_number(outstanding_key, current.get("Outstanding", current["Amount"]), float)
            )
        else:
            current.setdefault("Outstanding", float(current.get("Amount", 0.0)))

        interest_key = f"{session_key}_interest_{index}"
        if interest_key in st.session_state:
            current["Interest"] = float(
                _get_widget_number(interest_key, current.get("Interest", 0.0), float)
            )
        else:
            current.setdefault("Interest", float(current.get("Interest", 0.0)))
        updated.append(current)
    return updated


def _payload_to_tax_entries(payload: Mapping) -> list[dict]:
    tax = payload.get("tax", {}) if isinstance(payload, Mapping) else {}
    years = payload.get("years", []) if isinstance(payload, Mapping) else []
    if not isinstance(tax, Mapping):
        tax = {}
    base_rate = float(tax.get("rate", 0.0) or 0.0)
    schedule = tax.get("schedule", []) if isinstance(tax.get("schedule"), Sequence) else []
    return [
        {
            "label": str(year) if year is not None else f"Year {index + 1}",
            "rate": float(schedule[index]) if index < len(schedule) else base_rate,
        }
        for index, year in enumerate(years or [])
    ]


def _sync_tax_entries_from_widgets(rows: Sequence[Mapping]) -> list[dict]:
    updated: list[dict] = []
    for index, row in enumerate(rows):
        label = str(_get_widget_value(f"tax_year_label_{index}", row.get("label", ""))).strip()
        if not label:
            label = f"Year {index + 1}"
        rate = float(_get_widget_number(f"tax_rate_value_{index}", row.get("rate", 0.0), float))
        updated.append({"label": label, "rate": rate})
    return updated


def _sensitivity_rows_to_payload(rows: Sequence[Mapping], payload: dict) -> None:
    variables = {
        str(row.get("Variable", "")).strip(): list(row.get("Values", []))
        for row in rows
        if row.get("Variable") and row.get("Values")
    }
    payload.setdefault("sensitivity", {})["variables"] = variables

def _payload_to_debt_rows(payload: Mapping, key: str) -> list[dict]:
    financing = payload.get("financing", {}) if isinstance(payload, Mapping) else {}
    raw_rows = financing.get(key, []) if isinstance(financing, Mapping) else []

    if raw_rows is None:
        iterable: Iterable[Mapping] = []
    elif isinstance(raw_rows, Mapping):
        iterable = raw_rows.values()  # type: ignore[assignment]
    else:
        iterable = raw_rows  # type: ignore[assignment]

    rows: list[dict] = []
    for item in iterable:
        if not isinstance(item, Mapping):
            continue
        year_value = item.get("year")
        try:
            year = int(year_value)
        except (TypeError, ValueError):
            continue
        amount = float(item.get("amount", 0.0))
        outstanding = float(item.get("outstanding", amount))
        duration_value = item.get("duration", 1)
        try:
            duration = int(duration_value)
        except (TypeError, ValueError):
            duration = 1
        rows.append(
            {
                "Year": year,
                "Amount": amount,
                "Outstanding": outstanding,
                "Duration": max(1, duration),
            }
        )

    rows.sort(key=lambda row: row.get("Year", 0))
    return rows


def _debt_rows_to_payload(rows: Sequence[Mapping], payload: dict, key: str) -> None:
    cleaned: list[dict] = []
    for row in rows:
        year_value = row.get("Year")
        try:
            year = int(year_value)
        except (TypeError, ValueError):
            continue
        amount = float(row.get("Amount", 0.0))
        outstanding = float(row.get("Outstanding", amount))
        duration_value = row.get("Duration", 1)
        try:
            duration = int(duration_value)
        except (TypeError, ValueError):
            duration = 1
        cleaned.append(
            {
                "year": year,
                "amount": amount,
                "outstanding": outstanding,
                "duration": max(1, duration),
            }
        )

    cleaned.sort(key=lambda item: item["year"])
    financing = payload.setdefault("financing", {})
    financing[key] = cleaned
    if key == "senior_debt":
        financing.pop("senior_debt_schedule", None)
    if key == "revolver":
        financing.pop("revolver_initial", None)


def _payload_to_ai_settings(payload: Mapping) -> dict:
    ai = payload.get("ai", {}) if isinstance(payload, Mapping) else {}
    if not isinstance(ai, Mapping):
        ai = {}

    provider = str(ai.get("provider", "OpenAI") or "OpenAI")
    model = str(ai.get("model", "gpt-4") or "gpt-4")

    horizon_value = ai.get("forecast_horizon", 3)
    try:
        horizon = int(float(horizon_value))
    except (TypeError, ValueError):
        horizon = 3
    horizon = max(horizon, 0)

    ml_methods = [
        str(method).strip().lower()
        for method in ai.get("ml_methods", ["linear_regression"])
        if str(method).strip()
    ]
    if not ml_methods:
        ml_methods = ["linear_regression"]

    features = [
        str(feature).strip().lower()
        for feature in ai.get("generative_features", ["summary"])
        if str(feature).strip()
    ]
    if not features:
        features = ["summary"]

    api_key_value = ai.get("api_key", "")
    if isinstance(api_key_value, (str, bytes)):
        api_key = api_key_value.strip()
    else:
        api_key = ""

    return {
        "enabled": bool(ai.get("enabled", False)),
        "provider": provider,
        "model": model,
        "forecast_horizon": horizon,
        "ml_methods": ml_methods,
        "generative_features": features,
        "api_key": api_key,
    }


def _ai_settings_to_payload(settings: Mapping[str, object], payload: dict) -> None:
    if settings is None:
        return

    ai = payload.setdefault("ai", {})
    if not isinstance(ai, dict):
        ai = {}
        payload["ai"] = ai

    ai["enabled"] = bool(settings.get("enabled", False))
    ai["provider"] = str(settings.get("provider", "OpenAI") or "OpenAI")
    ai["model"] = str(settings.get("model", "gpt-4") or "gpt-4")

    horizon_value = settings.get("forecast_horizon", 3)
    try:
        horizon = int(float(horizon_value))
    except (TypeError, ValueError):
        horizon = 3
    ai["forecast_horizon"] = max(horizon, 0)

    ml_methods_raw = settings.get("ml_methods", ["linear_regression"])
    if isinstance(ml_methods_raw, Iterable) and not isinstance(ml_methods_raw, (str, bytes)):
        ml_methods = [
            str(method).strip().lower()
            for method in ml_methods_raw
            if str(method).strip()
        ]
    else:
        ml_methods = [str(ml_methods_raw).strip().lower()]
    ai["ml_methods"] = ml_methods or ["linear_regression"]

    features_raw = settings.get("generative_features", ["summary"])
    if isinstance(features_raw, Iterable) and not isinstance(features_raw, (str, bytes)):
        features = [
            str(feature).strip().lower()
            for feature in features_raw
            if str(feature).strip()
        ]
    else:
        features = [str(features_raw).strip().lower()]
    ai["generative_features"] = features or ["summary"]

    api_key = str(settings.get("api_key", "") or "")
    ai["api_key"] = api_key


def _render_ai_settings(payload: dict, container: Optional[DeltaGenerator] = None) -> None:
    target = container or st
    settings = st.session_state.setdefault("ai_settings", _payload_to_ai_settings(payload))
    st.session_state.setdefault("ai_api_key", settings.get("api_key", ""))

    provider_options = list(AI_PROVIDER_OPTIONS)
    if settings.get("provider") not in provider_options:
        provider_options.append(settings.get("provider"))

    current_provider = settings.get("provider", "OpenAI")
    try:
        provider_index = provider_options.index(current_provider)
    except ValueError:
        provider_index = 0

    ml_defaults = [
        ML_METHOD_LABELS.get(code, code.replace("_", " ").title())
        for code in settings.get("ml_methods", ["linear_regression"])
    ]
    feature_defaults = [
        GEN_AI_FEATURE_LABELS.get(code, code.replace("_", " ").title())
        for code in settings.get("generative_features", ["summary"])
    ]

    form = target.form("ai_settings_form")
    with form:
        enabled = form.checkbox(
            "Enable AI Enhancements",
            value=bool(settings.get("enabled", False)),
            help="Toggle machine-learning forecasts and generative commentary.",
        )
        provider = form.selectbox(
            "Provider",
            provider_options,
            index=provider_index,
            help="Select the API provider powering generative insights.",
        )
        model = form.text_input(
            "Model",
            value=settings.get("model", "gpt-4"),
            help="Name of the deployed model (for example `gpt-4o-mini`).",
        )
        horizon = form.number_input(
            "Forecast Horizon (years)",
            min_value=0,
            max_value=20,
            value=int(settings.get("forecast_horizon", 3)),
            step=1,
            help="Number of additional years used for machine-learning revenue forecasts.",
        )

        ml_selection = form.multiselect(
            "Machine Learning Methods",
            list(ML_METHOD_LABELS.values()),
            default=ml_defaults,
            help="Choose algorithms applied to projected net revenue.",
        )
        feature_selection = form.multiselect(
            "Generative Features",
            list(GEN_AI_FEATURE_LABELS.values()),
            default=feature_defaults,
            help="Pick the narrative focus areas generated by the AI summary.",
        )
        api_key = form.text_input(
            "API Key",
            value=st.session_state.get("ai_api_key", ""),
            type="password",
            help="Store your provider API key securely. Keys are retained only for the current session.",
        )

        submitted = form.form_submit_button("Save AI Configuration")

    if submitted:
        ml_codes = [ML_LABEL_TO_CODE.get(label, label.replace(" ", "_").lower()) for label in ml_selection]
        feature_codes = [
            GEN_AI_LABEL_TO_CODE.get(label, label.replace(" ", "_").lower())
            for label in feature_selection
        ]

        settings.update(
            {
                "enabled": enabled,
                "provider": provider,
                "model": model.strip() or "gpt-4",
                "forecast_horizon": int(horizon),
                "ml_methods": ml_codes or ["linear_regression"],
                "generative_features": feature_codes or ["summary"],
                "api_key": api_key.strip(),
            }
        )
        st.session_state["ai_settings"] = settings
        st.session_state["ai_api_key"] = settings.get("api_key", "")
        _ai_settings_to_payload(settings, payload)
        st.success("AI configuration updated. Rerunning the model with the new settings.")
        _rerun()


def _render_ai_summary(payload: Mapping) -> None:
    settings = _payload_to_ai_settings(payload)
    st.caption(
        "Adjust these settings from the Input Landing Page above the projection horizon controls."
    )

    rows = [
        {"Setting": "Enabled", "Value": "Yes" if settings.get("enabled") else "No"},
        {"Setting": "Provider", "Value": settings.get("provider", "OpenAI")},
        {"Setting": "Model", "Value": settings.get("model", "gpt-4")},
        {
            "Setting": "Forecast Horizon (years)",
            "Value": settings.get("forecast_horizon", 3),
        },
        {
            "Setting": "ML Methods",
            "Value": ", ".join(settings.get("ml_methods", [])) or "linear_regression",
        },
        {
            "Setting": "Generative Features",
            "Value": ", ".join(settings.get("generative_features", [])) or "summary",
        },
    ]

    if pd is not None:
        summary_frame = pd.DataFrame(rows).astype(str)
        st.table(summary_frame)
    else:
        for row in rows:
            st.write(f"**{row['Setting']}**: {row['Value']}")

def _payload_to_inflation_rows(payload: Mapping) -> list[dict]:
    years = list(payload.get("years", []))
    series = list(payload.get("inflation_series", []))
    default_rate = float(payload.get("inflation_rate", 0.0))

    rows: list[dict] = []
    if years:
        values = _ensure_schedule_length(series, len(years), fill=default_rate)
        for position, year in enumerate(years):
            rows.append({"Year": str(year), "Rate": float(values[position])})
    elif series:
        for index, value in enumerate(series, start=1):
            rows.append({"Year": f"Year {index}", "Rate": float(value)})
    else:
        rows.append({"Year": "Year 1", "Rate": default_rate})

    return rows


def _inflation_rows_to_payload(rows: Sequence[Mapping], payload: dict) -> None:
    if rows is None:
        return

    rates: list[float] = []
    labels: list[str] = []
    for index, row in enumerate(rows):
        label = str(row.get("Year", "")).strip()
        if not label:
            label = f"Year {index + 1}"
        labels.append(label)
        try:
            rate = float(row.get("Rate", 0.0))
        except (TypeError, ValueError):
            rate = 0.0
        rates.append(rate)

    if not rates:
        payload["inflation_series"] = []
        payload["inflation_labels"] = []
        return

    payload["inflation_series"] = list(rates)
    payload["inflation_labels"] = labels
    current_years = payload.get("years", [])
    target_length = max(len(current_years), len(rates))
    if target_length:
        _align_payload_horizon(payload, labels, target_length)


def _inflation_factors_from_payload(payload: Mapping) -> list[float]:
    series = payload.get("inflation_series", []) if isinstance(payload, Mapping) else []
    factors: list[float] = []
    running = 1.0
    if isinstance(series, Iterable):
        for value in series:
            try:
                rate = float(value)
            except (TypeError, ValueError):
                rate = 0.0
            running *= 1.0 + rate
            factors.append(running)
    return factors


def _risk_factors_from_payload(payload: Mapping) -> list[float]:
    if not isinstance(payload, Mapping):
        return []

    schedule = payload.get("risk", {})
    if not isinstance(schedule, Mapping):
        schedule = {}

    years = payload.get("years", [])
    target_length = len(years)
    if target_length == 0 and schedule:
        target_length = max((len(values) for values in schedule.values()), default=0)

    factors: list[float] = []
    for index in range(target_length):
        factor = 1.0
        for values in schedule.values():
            if not isinstance(values, Iterable) or isinstance(values, (str, bytes)):
                continue
            if not values:
                continue
            raw = values[index] if index < len(values) else values[-1]
            try:
                rate = float(raw)
            except (TypeError, ValueError):
                rate = 0.0
            factor *= max(0.0, 1.0 - rate)
        factors.append(factor)
    return factors


def _scaled_production_series(
    product: str,
    total_units: float,
    years: Sequence[Any],
    existing_estimate: Mapping[str, Sequence[Any]] | Sequence[Any] | None,
) -> list[float]:
    if isinstance(existing_estimate, Mapping) and product in existing_estimate:
        series = [float(value) for value in existing_estimate.get(product, [])]
    elif isinstance(existing_estimate, Sequence) and not isinstance(
        existing_estimate, (str, bytes)
    ):
        # Legacy payloads may store a simple list when only one product exists.
        series = [float(value) for value in existing_estimate]
    else:
        series = []

    target_length = len(years)
    if target_length == 0:
        return []

    if len(series) < target_length:
        series = series + [0.0] * (target_length - len(series))
    elif len(series) > target_length:
        series = series[:target_length]

    current_total = sum(series)
    if current_total > 0:
        factor = total_units / current_total if current_total else 0.0
        return [value * factor for value in series]

    per_year = total_units / target_length if target_length else 0.0
    return [per_year for _ in range(target_length)]


def _payload_to_fixed_variable_rows(payload: Mapping) -> list[dict]:
    section = payload.get("fixed_variable_costs") if isinstance(payload, Mapping) else None
    rows: list[dict] = []

    raw_rows: Iterable | None
    if isinstance(section, Mapping):
        raw_rows = section.get("rows", section.get("data", []))
    else:
        raw_rows = section if isinstance(section, Iterable) else None

    if isinstance(raw_rows, Iterable) and not isinstance(raw_rows, (str, bytes)):
        for entry in raw_rows:
            if not isinstance(entry, Mapping):
                continue
            product = str(entry.get("product") or entry.get("Product") or "").strip()
            if not product:
                continue
            has_fixed = "fixed_cost" in entry or "Fixed Cost" in entry
            fixed_cost = float(entry.get("fixed_cost", entry.get("Fixed Cost", 0.0)) or 0.0)
            variable_cost = float(
                entry.get("variable_cost", entry.get("Variable Cost", 0.0)) or 0.0
            )
            rows.append(
                {
                    "Product": product,
                    "Fixed Cost": fixed_cost,
                    "Variable Cost": variable_cost,
                    "__has_fixed__": has_fixed,
                }
            )

    return rows


def _fixed_variable_rows_to_payload(rows: Sequence[Mapping], payload: dict) -> None:
    if rows is None:
        return

    section = payload.setdefault("fixed_variable_costs", {})
    if not isinstance(section, dict):
        section = {}
        payload["fixed_variable_costs"] = section

    serialised: list[dict] = []
    mapping: dict[str, tuple[Optional[float], float]] = {}
    for row in rows:
        if not isinstance(row, Mapping):
            continue
        product = str(row.get("Product", "") or "").strip()
        if not product:
            continue
        fixed_cost = float(row.get("Fixed Cost", 0.0) or 0.0)
        variable_cost = float(row.get("Variable Cost", 0.0) or 0.0)
        has_fixed = bool(row.get("__has_fixed__", False))
        entry: dict[str, float] = {"product": product, "variable_cost": variable_cost}
        if has_fixed:
            entry["fixed_cost"] = fixed_cost
        serialised.append(entry)
        mapping[product] = (fixed_cost if has_fixed else None, variable_cost)

    section["rows"] = serialised

    break_even = payload.get("break_even")
    if isinstance(break_even, dict):
        raw_rows = break_even.setdefault("rows", [])
        if isinstance(raw_rows, list):
            for entry in raw_rows:
                if not isinstance(entry, dict):
                    continue
                product = str(entry.get("product") or entry.get("Product") or "").strip()
                if product in mapping:
                    fixed_cost, variable_cost = mapping[product]
                    if fixed_cost is not None:
                        entry["fixed_cost"] = fixed_cost
                    elif "fixed_cost" in entry:
                        entry.pop("fixed_cost", None)
                    entry["variable_cost"] = variable_cost

    if "break_even_rows" in st.session_state and isinstance(
        st.session_state["break_even_rows"], list
    ):
        updated: list[dict] = []
        for entry in st.session_state["break_even_rows"]:
            product = str(entry.get("Product", "") or "").strip()
            if product in mapping:
                fixed_cost, variable_cost = mapping[product]
                new_entry = dict(entry)
                if fixed_cost is not None:
                    new_entry["Fixed Cost"] = fixed_cost
                elif "Fixed Cost" in new_entry:
                    new_entry["Fixed Cost"] = entry.get("Fixed Cost", 0.0)
                new_entry["Variable Cost"] = variable_cost
                new_entry["__has_fixed__"] = fixed_cost is not None
                updated.append(new_entry)
            else:
                updated.append(entry)
        st.session_state["break_even_rows"] = updated


def _payload_to_commission_rows(payload: Mapping) -> list[dict]:
    section = payload.get("distributor_commission") if isinstance(payload, Mapping) else None
    if isinstance(section, Mapping):
        raw_rows = section.get("rows", section)
    else:
        raw_rows = section

    rows: list[dict] = []
    if isinstance(raw_rows, Iterable) and not isinstance(raw_rows, (str, bytes)):
        for item in raw_rows:
            if not isinstance(item, Mapping):
                continue
            product = str(item.get("product", "")).strip()
            if not product:
                continue
            year_value = item.get("year")
            try:
                year = int(year_value)
            except (TypeError, ValueError):
                continue
            rate = float(item.get("rate", 0.0) or 0.0) * 100.0
            share = float(item.get("revenue_share", 1.0) or 0.0) * 100.0
            payment_days = int(float(item.get("payment_days", 0) or 0))
            revenue_estimate = _commission_revenue_estimate(payload, year, product)
            revenue_value = revenue_estimate * (share / 100.0) if revenue_estimate else 0.0
            rows.append(
                {
                    "Year": year,
                    "Product": product,
                    "Commission (%)": rate,
                    "Revenue Share (%)": share if share > 0 else 100.0,
                    "Revenue": revenue_value,
                    "Payment Days": max(payment_days, 0),
                }
            )

    if rows:
        rows.sort(key=lambda row: (row["Year"], str(row["Product"]).lower()))
        grouped: dict[str, list[dict]] = {}
        for row in rows:
            grouped.setdefault(str(row["Product"]), []).append(row)
        for entries in grouped.values():
            entries.sort(key=lambda entry: entry["Year"])
            previous_rate: float | None = None
            for entry in entries:
                current_rate = float(entry.get("Commission (%)", 0.0))
                if previous_rate and previous_rate > 0:
                    increment = (current_rate / previous_rate - 1.0) * 100.0
                else:
                    increment = 0.0
                entry["Yearly Commission %"] = increment
                previous_rate = current_rate
        return rows

    years = payload.get("years", []) if isinstance(payload, Mapping) else []
    unit_costs = payload.get("unit_costs", {}) if isinstance(payload, Mapping) else {}
    fallback: list[dict] = []
    if isinstance(unit_costs, Mapping):
        for year in years:
            for product in unit_costs.keys():
                # Default distributor commission rate: 5%
                rate = 0.05
                revenue_estimate = _commission_revenue_estimate(payload, int(year), str(product))
                fallback.append(
                    {
                        "Year": int(year),
                        "Product": str(product),
                        "Yearly Commission %": 0.0,
                        "Revenue Share (%)": 100.0,
                        "Revenue": revenue_estimate,
                        "Payment Days": 30,
                    }
                )

    fallback.sort(key=lambda row: (row["Year"], str(row["Product"]).lower()))
    return fallback


def _commission_rows_to_payload(rows: Sequence[Mapping], payload: dict) -> None:
    cleaned: list[dict] = []
    grouped: dict[str, list[dict[str, object]]] = {}
    for row in rows or []:
        if not isinstance(row, Mapping):
            continue
        year_value = row.get("Year")
        try:
            year = int(year_value)
        except (TypeError, ValueError):
            continue
        product = str(row.get("Product", "")).strip()
        if not product:
            continue
        grouped.setdefault(product, []).append(
            {
                "Year": year,
                "Product": product,
                "Yearly Commission %": row.get("Yearly Commission %", 0.0),
                "Revenue": row.get("Revenue"),
                "Revenue Share (%)": row.get("Revenue Share (%)", row.get("Revenue Share", 100.0)),
                "Payment Days": row.get("Payment Days", 0),
            }
        )

    base_rates = _commission_base_rates(payload)
    for product, entries in grouped.items():
        entries.sort(key=lambda entry: entry["Year"])
        rate = base_rates.get(product, 0.05)
        for entry in entries:
            increment_value = entry.get("Yearly Commission %", 0.0)
            try:
                increment = float(increment_value)
            except (TypeError, ValueError):
                increment = 0.0
            rate = rate * (1 + increment / 100.0)
            revenue_value = entry.get("Revenue")
            share_value = entry.get("Revenue Share (%)", 100.0)
            share = None
            if revenue_value is not None:
                revenue_estimate = _commission_revenue_estimate(payload, entry["Year"], product)
                if revenue_estimate > 0:
                    try:
                        share = max(float(revenue_value) / float(revenue_estimate), 0.0)
                    except (TypeError, ValueError):
                        share = None
            if share is None:
                try:
                    share = max(float(share_value) / 100.0, 0.0)
                except (TypeError, ValueError):
                    share = 1.0
            payment_value = entry.get("Payment Days", 0)
            try:
                payment_days = max(int(float(payment_value)), 0)
            except (TypeError, ValueError):
                payment_days = 0
            cleaned.append(
                {
                    "year": int(entry["Year"]),
                    "product": product,
                    "rate": rate,
                    "revenue_share": share if share > 0 else 1.0,
                    "payment_days": payment_days,
                }
            )

    cleaned.sort(key=lambda entry: (entry["year"], entry["product"].lower()))
    if cleaned:
        payload["distributor_commission"] = {"rows": cleaned}
    else:
        payload.pop("distributor_commission", None)

def _payload_to_risk_rows(payload: Mapping) -> list[dict]:
    source: Mapping[str, Sequence[float]] = payload.get("risk", {}) or {}
    risk: dict[str, list[float]] = {}
    for name, values in source.items():
        key = str(name).strip().lower()
        if not key:
            continue
        risk[key] = [float(value) for value in values]

    labels = list(payload.get("inflation_labels") or payload.get("years", []))
    categories = _risk_categories(payload)

    max_length = max([len(labels)] + [len(values) for values in risk.values()] or [0])
    if max_length == 0:
        max_length = 1

    if not labels:
        labels = [f"Year {index + 1}" for index in range(max_length)]
    elif len(labels) < max_length:
        labels = labels + [f"Year {index + 1}" for index in range(len(labels), max_length)]

    rows: list[dict] = []
    for index in range(max_length):
        label = labels[index] if index < len(labels) else f"Year {index + 1}"
        row = {"Year": str(label)}
        for category in categories:
            values = risk.get(category, [])
            row[category] = float(values[index]) if index < len(values) else 0.0
        rows.append(row)

    return rows


def _risk_rows_to_payload(rows: Sequence[Mapping], payload: dict) -> None:
    if rows is None:
        return

    categories = _risk_categories(payload, rows)
    if not rows:
        payload["risk"] = {category: [] for category in categories}
        return

    labels: list[str] = []
    risk_payload: dict[str, list[float]] = {category: [] for category in categories}

    for index, row in enumerate(rows):
        label = str(row.get("Year", "")).strip()
        if not label:
            label = f"Year {index + 1}"
        labels.append(label)
        for category in categories:
            try:
                value = float(row.get(category, 0.0))
            except (TypeError, ValueError):
                value = 0.0
            risk_payload[category].append(min(max(value, 0.0), 1.0))

    payload["risk"] = risk_payload
    current_years = payload.get("years", [])
    target_length = max(len(current_years), len(rows))
    if target_length:
        _align_payload_horizon(payload, labels, target_length)

    if "risk_rows" in st.session_state:
        st.session_state["risk_rows"] = _payload_to_risk_rows(payload)


def _risk_categories(payload: Mapping | None = None, rows: Sequence[Mapping] | None = None) -> list[str]:
    categories: list[str] = []
    seen: set[str] = set()

    def _add(name: str | None) -> None:
        if not name:
            return
        key = str(name).strip().lower()
        if not key or key == "year" or key in seen:
            return
        seen.add(key)
        categories.append(key)

    for default in DEFAULT_RISK_CATEGORIES:
        _add(default)

    if payload and isinstance(payload.get("risk"), Mapping):
        for name in payload["risk"].keys():
            _add(str(name))

    if rows:
        for row in rows:
            for key in row.keys():
                _add(str(key))

    return categories


def _align_payload_horizon(
    payload: dict,
    labels: Sequence[str],
    target_length: int,
    *,
    update_years: bool = False,
) -> None:
    if target_length <= 0:
        return

    years = list(payload.get("years", []))
    derived_years = _derive_years_from_labels(labels)

    base_years = years
    if derived_years:
        # Avoid replacing calendar years with relative labels (e.g. "Year 1")
        # unless the model lacks any existing horizon or the user explicitly
        # updated the start/end years via the horizon controls.
        if update_years:
            base_years = derived_years
        elif not years and all(year >= 1900 for year in derived_years):
            base_years = derived_years

    payload["years"] = _resize_years(years, target_length, base_years)

    payload["inflation_series"] = _resize_sequence(
        payload.get("inflation_series", []), target_length
    )

    production = payload.get("production_estimate", {})
    for name, series in list(production.items()):
        production[name] = _resize_sequence(series, target_length)

    utility = payload.setdefault("utility_costs", {})
    existing_rows: Sequence[Mapping] | list[dict] = []
    if isinstance(utility.get("years"), Sequence):
        existing_rows = list(utility.get("years", []))

    resized_rows = _resize_utility_entries(existing_rows, target_length, labels)
    utility["years"] = resized_rows
    for legacy in ("electricity_per_day", "water_per_day", "steam_per_hour", "days", "hours"):
        utility.pop(legacy, None)

    tax = payload.setdefault("tax", {})
    schedule = tax.get("schedule", [])
    fill_rate = tax.get("rate", schedule[-1] if schedule else 0.0)
    tax["schedule"] = _resize_sequence(schedule, target_length, fill=fill_rate)

    risk = payload.setdefault("risk", {})
    for category, values in list(risk.items()):
        risk[category] = _resize_sequence(values, target_length)

    working_capital = payload.setdefault("working_capital", {})
    days_mapping = working_capital.get("days") if isinstance(working_capital.get("days"), Mapping) else {}
    if not isinstance(days_mapping, Mapping):
        days_mapping = {}
    working_capital["days"] = days_mapping
    for key, values in list(days_mapping.items()):
        days_mapping[key] = _resize_sequence(values, target_length)

    calendar_values = working_capital.get("calendar_days")
    if isinstance(calendar_values, Iterable) and not isinstance(calendar_values, (str, bytes)):
        fill = calendar_values[-1] if calendar_values else None
        if fill is None:
            derived = payload.get("years", [])
            fill = 366 if derived and _is_leap_year(int(derived[-1])) else 365
        working_capital["calendar_days"] = _resize_sequence(calendar_values, target_length, fill=fill)
    else:
        years_for_calendar = payload.get("years", [])[:target_length]
        if years_for_calendar:
            working_capital["calendar_days"] = [
                366 if _is_leap_year(int(year)) else 365 for year in years_for_calendar
            ]
        else:
            working_capital["calendar_days"] = [365 for _ in range(target_length)]

    scenarios = payload.get("scenarios", {})
    for scenario in scenarios.values():
        if not isinstance(scenario, Mapping):
            continue
        if "inflation" in scenario:
            scenario["inflation"] = _resize_sequence(scenario.get("inflation", []), target_length)
        if "interest" in scenario:
            scenario["interest"] = _resize_sequence(scenario.get("interest", []), target_length)


def _is_leap_year(year: int) -> bool:
    return year % 400 == 0 or (year % 4 == 0 and year % 100 != 0)


def _resize_sequence(values: Iterable, target_length: int, fill=None) -> list:
    items = list(values)
    if target_length <= 0:
        return []
    if len(items) >= target_length:
        return items[:target_length]
    if fill is None:
        fill = items[-1] if items else 0
    items.extend([fill for _ in range(target_length - len(items))])
    return items


def _resize_years(current: Sequence[int], target_length: int, derived: Sequence[int]) -> list[int]:
    if target_length <= 0:
        return []
    if derived and len(derived) == target_length:
        return [int(value) for value in derived]
    existing = list(current)
    if len(existing) >= target_length:
        return [int(value) for value in existing[:target_length]]
    if existing:
        if len(existing) >= 2:
            step = existing[1] - existing[0]
        else:
            step = 1
        base = existing[-1]
        extension = [int(base + step * (index + 1)) for index in range(target_length - len(existing))]
        return [int(value) for value in existing + extension]
    return [index + 1 for index in range(target_length)]


def _derive_years_from_labels(labels: Sequence[str]) -> list[int]:
    derived: list[int] = []
    for label in labels:
        value = _parse_year_number(label)
        if value is None:
            return []
        derived.append(value)
    return derived


def _parse_year_number(label: str) -> int | None:
    if not label:
        return None
    try:
        return int(float(label))
    except ValueError:
        match = re.search(r"-?\d+(?:\.\d+)?", label)
        if match:
            try:
                return int(float(match.group()))
            except ValueError:
                return None
    return None


if __name__ == "__main__":  # pragma: no cover - Streamlit executes the script directly
    if _streamlit_runtime_exists():
        main()
    else:  # pragma: no cover - guidance for incorrect invocation
        raise SystemExit(
            "This module is a Streamlit application. Launch it with "
            "`streamlit run streamlit_app.py` instead of executing it directly."
        )
